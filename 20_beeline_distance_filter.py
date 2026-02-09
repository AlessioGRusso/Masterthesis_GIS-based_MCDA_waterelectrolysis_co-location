from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------- Pfade ----------------
ROOT = Path(__file__).resolve().parent
DATA = ROOT / "Daten"
OUT  = ROOT / "Output"

XLSX_IN     = OUT / "UWWTD_TP_Database.xlsx"
EHB_TX_JSON = DATA / "Transmission.json"
EHB_HP_JSON = DATA / "High Pressure Distribiution.json"

# ---------------- Excel / Spalten ----------------
SHEET_GENERAL  = "General Data"
SHEET_HYDROGEN = "H2 Logistics"

COL_CODE = "UWWTD Code"
COL_NAME = "Name"
COL_LAT  = "Latitude"
COL_LON  = "Longitude"
COL_PE   = "Capacity/PE"

COL_BUILT_EHB = "Built Scenario 1 (EHB)"
COL_DIST_EHB  = "Direct Distance to EHB [km]"

# ---------------- Ökonomie ----------------
THRESHOLD = 0.0  # "bauen", wenn decision(dist_km, PE) >= THRESHOLD

try:
    # decision(dist_km, pe) -> float oder (…, €/kg)
    from electrolysis_decision import decision
except Exception as e:
    raise ImportError(
        "Konnte 'electrolysis_decision.decision' nicht importieren. "
        "Lege 'electrolysis_decision.py' ins Projekt-Root."
    ) from e


def econ_value(dist_km: float, pe: float) -> float:
    """Gibt den relevanten Wert aus decision() als float zurück (NaN bei Fehlern)."""
    try:
        v = decision(float(dist_km), float(pe))
        if isinstance(v, (list, tuple)) and len(v) >= 2:
            return float(v[1])
        return float(v)
    except Exception:
        return float("nan")


# ---------------- Geo/CRS Helfer ----------------
def utm_crs_from_latlon(lat: float, lon: float) -> str:
    """Wählt eine passende UTM-Zone als metrisches CRS (für saubere Distanzberechnung)."""
    zone = int((lon + 180) / 6) + 1
    south = lat < 0
    return f"+proj=utm +zone={zone} +datum=WGS84 +units=m +no_defs {'+south' if south else ''}"


def _read_lines(path: Path) -> gpd.GeoDataFrame:
    """Liest Linien (GeoJSON/GPKG) robust ein und filtert auf gültige LineStrings."""
    if not path.exists():
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    gdf = gpd.read_file(path)
    if gdf.empty:
        return gdf.set_crs("EPSG:4326")

    if gdf.crs is None:
        gdf = gdf.set_crs("EPSG:4326")

    gdf = gdf[gdf.geometry.notnull() & gdf.geometry.is_valid]
    gdf = gdf[gdf.geometry.geom_type.isin(["LineString", "MultiLineString"])]

    if gdf.empty:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    return gdf


def load_line_sources(paths: List[Path], extras: Optional[List[gpd.GeoDataFrame]] = None) -> gpd.GeoDataFrame:
    """Lädt mehrere Linienquellen und führt sie zusammen."""
    parts: List[gpd.GeoDataFrame] = []
    for p in paths:
        try:
            g = _read_lines(p)
            if not g.empty:
                parts.append(g)
        except Exception:
            continue

    if extras:
        parts.extend([e for e in extras if e is not None and not e.empty])

    if not parts:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    return gpd.GeoDataFrame(pd.concat(parts, ignore_index=True), crs=parts[0].crs)


def compute_min_distance_to_lines_km(
    df_points: pd.DataFrame, lat_col: str, lon_col: str, lines: gpd.GeoDataFrame
) -> pd.Series:
    """Minimale Distanz eines Punktes zur nächsten Linie (km). Index bleibt wie df_points."""
    out = pd.Series(np.nan, index=df_points.index, dtype=float)
    if lines.empty or df_points.empty:
        return out

    lat = pd.to_numeric(df_points[lat_col], errors="coerce")
    lon = pd.to_numeric(df_points[lon_col], errors="coerce")
    valid = lat.notna() & lon.notna() & np.isfinite(lat) & np.isfinite(lon)
    if not valid.any():
        return out

    pts = gpd.GeoDataFrame(
        geometry=[Point(xy) for xy in zip(lon[valid], lat[valid])],
        crs="EPSG:4326",
        index=df_points.index[valid],
    )

    # metrisches CRS rund um den Schwerpunkt der Punkte
    try:
        cen = pts.geometry.union_all().centroid
    except Exception:
        cen = pts.geometry.unary_union.centroid

    crs_m = utm_crs_from_latlon(cen.y, cen.x)

    joined = gpd.sjoin_nearest(
        pts.to_crs(crs_m),
        lines.to_crs(crs_m)[["geometry"]],
        how="left",
        distance_col="_dist_m",
    )

    out.loc[joined.index] = (joined["_dist_m"] / 1000.0).astype(float)
    return out


# ---------------- Output: Export kept plants als GPKG ----------------
def export_kept_gpkg(kept_df: pd.DataFrame) -> None:
    """Exportiert die behaltenen Anlagen als GeoPackage (Punkte in EPSG:4326)."""
    out_dir = OUT / "WWTP Geopackages"
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / "WWTPS_kept_after_BDF.gpkg"
    layer_name = "WWTPS_kept_after_BDF"

    df = kept_df.dropna(subset=["_lat", "_lon"]).copy()
    if df.empty:
        print("ℹ️ Keine behaltenen Anlagen mit Koordinaten → kein GPKG erzeugt.")
        return

    geom = [Point(lon, lat) for lon, lat in zip(df["_lon"], df["_lat"])]
    gdf = gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")

    cols = [c for c in [COL_CODE, COL_NAME, COL_LAT, COL_LON, COL_PE] if c in gdf.columns]
    gdf = gdf[cols + ["geometry"]]

    if out_path.exists():
        out_path.unlink()

    gdf.to_file(out_path, layer=layer_name, driver="GPKG")
    print(f"✓ GPKG exportiert: {len(gdf)} Anlagen → {out_path.name}")


# (Optional) Excel: Spaltenbreite anpassen (wird aktuell nicht automatisch genutzt)
def autosize_sheet(path: Path, sheet_name: str, min_w: float = 8.0, max_w: float = 80.0) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        max_len = 0
        for c in col:
            if c.value is not None:
                s = str(c.value)
                max_len = max(max_len, max(len(line) for line in s.splitlines()))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, max_len + 2))

    wb.save(path)
    wb.close()


# ---------------- Hauptablauf ----------------
def main() -> None:
    if not XLSX_IN.exists():
        raise FileNotFoundError(f"Excel nicht gefunden: {XLSX_IN}")

    sheets: Dict[str, pd.DataFrame] = pd.read_excel(XLSX_IN, sheet_name=None, engine="openpyxl")
    if SHEET_GENERAL not in sheets:
        raise KeyError(f"Sheet '{SHEET_GENERAL}' fehlt in {XLSX_IN.name}")

    df_gen = sheets[SHEET_GENERAL].copy()

    # Pflichtspalten
    for c in (COL_CODE, COL_NAME, COL_LAT, COL_LON, COL_PE):
        if c not in df_gen.columns:
            raise KeyError(f"Spalte '{c}' fehlt in '{SHEET_GENERAL}'")

    work = df_gen.copy().reset_index(drop=True)
    work["_lat"] = pd.to_numeric(work[COL_LAT], errors="coerce")
    work["_lon"] = pd.to_numeric(work[COL_LON], errors="coerce")
    work["_pe"]  = pd.to_numeric(work[COL_PE],  errors="coerce")

    # 1) EHB laden
    ehb_lines = load_line_sources([EHB_TX_JSON, EHB_HP_JSON])

    # 2) Distanz zur nächsten EHB-Leitung
    work[COL_DIST_EHB] = compute_min_distance_to_lines_km(work, "_lat", "_lon", ehb_lines).round(3)

    # 3) Entscheidung (ökonomisch sinnvoll?)
    econ = pd.Series(
        [econ_value(d, p) for d, p in zip(work[COL_DIST_EHB], work["_pe"])],
        index=work.index,
        dtype=float,
    )

    built_ehb = (econ >= THRESHOLD).astype(int)

    # 4) Filter: nur behalten, wenn Built=1
    keep_mask = built_ehb.eq(1)
    kept = work.loc[keep_mask].copy()
    kept_codes = set(kept[COL_CODE].astype(str))

    # 5) H2 Logistics neu bauen (nur kept)
    hydrogen = pd.DataFrame(
        {
            COL_CODE: kept[COL_CODE],
            COL_NAME: kept[COL_NAME],
            COL_DIST_EHB: kept[COL_DIST_EHB],
            COL_BUILT_EHB: built_ehb.loc[kept.index],
        }
    )[[COL_CODE, COL_NAME, COL_DIST_EHB, COL_BUILT_EHB]]

    # 6) Alle Sheets auf kept_codes filtern
    new_sheets: Dict[str, pd.DataFrame] = {}
    for name, df in sheets.items():
        if COL_CODE in df.columns:
            new_sheets[name] = df[df[COL_CODE].astype(str).isin(kept_codes)].copy()
        else:
            new_sheets[name] = df.copy()

    new_sheets[SHEET_HYDROGEN] = hydrogen

    # 7) Sync: Sheets dürfen nur Codes enthalten, die im General-Tab existieren
    general_codes = set(new_sheets[SHEET_GENERAL][COL_CODE].astype(str).str.strip())
    for name, df in list(new_sheets.items()):
        if name != SHEET_GENERAL and COL_CODE in df.columns:
            before = len(df)
            new_sheets[name] = df[df[COL_CODE].astype(str).str.strip().isin(general_codes)].copy()
            after = len(new_sheets[name])
            if before != after:
                print(f"  → {name}: {before} → {after} Zeilen (Sync mit General Data)")

    # 8) Workbook komplett neu schreiben
    XLSX_IN.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(XLSX_IN, engine="openpyxl", mode="w") as xw:
        for name, df in new_sheets.items():
            df.to_excel(xw, sheet_name=name, index=False)

    print(f"✓ Fertig: {len(kept)} von {len(df_gen)} Anlagen behalten (EHB built=1)")
    export_kept_gpkg(kept)


if __name__ == "__main__":
    main()
