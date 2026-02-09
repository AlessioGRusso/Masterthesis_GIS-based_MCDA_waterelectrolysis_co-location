import os
import sys
import time
import re
import threading
import subprocess
from pathlib import Path
from queue import Queue, Empty
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

try:
    from tksheet import Sheet  # optional
except ImportError:
    Sheet = None


# -----------------------------------------------------------------------------
# Mini-Logger (keine externen Abhängigkeiten, optional still)
# -----------------------------------------------------------------------------
DEBUG = False  # auf True setzen, wenn du beim Debuggen mehr sehen willst


def _dbg(msg: str) -> None:
    if DEBUG:
        print(f"[00_database] {msg}")


def center_on_screen(win: tk.Tk | tk.Toplevel):
    """Zentriert ein Tkinter-Fenster auf dem Bildschirm."""
    try:
        win.update_idletasks()
        w = win.winfo_width()
        h = win.winfo_height()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = int((sw - w) / 2)
        y = int((sh - h) / 2)
        win.geometry(f"+{x}+{y}")
    except Exception as e:
        # Zentrierung ist nicht kritisch → im Zweifel ignorieren
        _dbg(f"center_on_screen failed: {e}")


# Pfade und wichtige Konstanten
BASE = Path(__file__).resolve().parent
OUT_XLSX_DEFAULT = BASE / "Output" / "UWWTD_TP_Database.xlsx"

# Alle Skripte die nacheinander ausgeführt werden
SCRIPTS = [
    # Datenbank aufbauen und Geopackages erstellen
    BASE / "10_build_database.py",           # Datenbank wird gebaut; KO-Kriterium: >300k PE
    BASE / "11_patch_database.py",           # Falsche Koordinaten werden korrigiert

    # KO-Kriterien (Ausschlusskriterien)
    BASE / "20_beeline_distance_filter.py",  # KO-Kriterium: Luftlinien-Entfernung
    BASE / "21_available_area_filter.py",    # KO-Kriterium: Verfügbarkeit freier Fläche nahe der Anlage
    BASE / "22_risk_assessment.py",          # Risikobewertung (erstellt risks_cache.csv)
    BASE / "23_pa_filter.py",                # KO-Kriterium: POI in streng geschütztem Gebiet + erstellt Available_Area.gpkg

    # H2 Logistik
    BASE / "24_least_cost_path.py",          # KO-Kriterium: Entfernung einer berechneten "realen" Pipeline-Route (nutzt Available_Area.gpkg)
    BASE / "25_lcp_filter.py",

    BASE / "30_hydrogen_offtakers.py",

    # H2 Erneuerbare Energien
    BASE / "40_renewable_energy_profiles.py",
    BASE / "41_LCOH_optimization.py",

    # Netzanschluss Energie
    BASE / "50_energy_network.py",

    # Fernwärme
    BASE / "60_district_heating_networks.py",  # Fernwärmenetz vorhanden oder geplant + technische Parameter

    # Zusätzliche Daten
    BASE / "70_energy_prices.py",
    BASE / "80_calculation_oxygen_demand.py",
]


# -----------------------------------------------------------------------------
# Excel-Formatierung: Spaltenbreite automatisch + Zahlen rechtsbündig
# -----------------------------------------------------------------------------
def _text_len(val) -> int:
    s = str(val)
    return max(len(part) for part in s.splitlines()) if "\n" in s else len(s)


_NUMERIC_LIKE_RE = re.compile(
    r"""^\s*[-+]?
        (?:
          \d+(?:[.,]\d+)? |
          \d{1,3}(?:[.,]\d{3})+(?:[.,]\d+)?
        )\s*$""",
    re.X
)


def _is_numeric_like(value) -> bool:
    if isinstance(value, (int, float)):
        return True
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    return bool(_NUMERIC_LIKE_RE.match(s))


def autosize_excel(path: Path, min_w: float = 8.0, max_w: float = 80.0,
                   retries: int = 8, wait_s: float = 1.0) -> None:
    """Passt Spaltenbreiten an + richtet Zahlen rechts aus. Robust gegen Excel-locks."""
    if not path.exists():
        return

    for attempt in range(1, retries + 1):
        try:
            wb_write = load_workbook(path, data_only=False)
            wb_vals = load_workbook(path, data_only=True, read_only=True)

            for ws in wb_write.worksheets:
                ws_vals = wb_vals[ws.title] if ws.title in wb_vals.sheetnames else None
                max_row = ws.max_row or 1

                for col in ws.iter_cols(min_row=1, max_row=max_row):
                    max_len = 0
                    col_idx = col[0].column

                    for c in col:
                        # 1) Auto-Breite
                        if c.value is not None:
                            max_len = max(max_len, _text_len(c.value))

                        # 2) Numerik (Zeile 1 = Header überspringen)
                        if c.row == 1:
                            continue

                        numeric = False
                        if c.data_type == "n" or isinstance(c.value, (int, float)):
                            numeric = True
                        elif c.data_type == "f" and ws_vals is not None:
                            # Formel → versuche gecachten Wert aus data_only workbook
                            try:
                                cached = ws_vals.cell(row=c.row, column=c.column).value
                                numeric = isinstance(cached, (int, float)) or _is_numeric_like(cached)
                            except Exception as e:
                                _dbg(f"autosize formula cached lookup failed: {e}")
                        else:
                            numeric = _is_numeric_like(c.value)

                        if numeric:
                            old = c.alignment or Alignment()
                            c.alignment = Alignment(
                                horizontal="right",
                                vertical=old.vertical,
                                text_rotation=old.text_rotation,
                                wrap_text=old.wrap_text,
                                shrink_to_fit=old.shrink_to_fit,
                                indent=old.indent
                            )

                    # 3) Breite setzen
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w, max(min_w, max_len + 2))

            # Sheets nach bevorzugter Reihenfolge sortieren (wenn möglich)
            try:
                desired_order = [
                    "Summary",
                    "General Data",
                    "Technical Data - Plant Metrics",
                    "H2 Logistics",
                    "H2 Renewables",
                    "Grid Energy Connection",
                    "Risks",
                    "District Heating",
                    "Additional Data",
                ]
                priority = {name: i for i, name in enumerate(desired_order)}
                orig_pos = {id(ws): i for i, ws in enumerate(wb_write.worksheets)}
                wb_write._sheets.sort(
                    key=lambda _ws: (priority.get(_ws.title, len(desired_order)), orig_pos[id(_ws)])
                )
            except Exception as e:
                _dbg(f"sheet sort failed: {e}")

            wb_write.save(path)
            wb_write.close()
            wb_vals.close()
            break

        except PermissionError:
            # Excel hat die Datei offen → retry
            if attempt == retries:
                raise
            time.sleep(wait_s)


# -----------------------------------------------------------------------------
# Anzeige-Hilfen für Viewer
# -----------------------------------------------------------------------------
def _compute_col_pixel_widths(headers, data, char_px=7, pad=2, min_px=80, max_px=520):
    cols = len(headers)
    maxlens = [len(str(h or "")) for h in headers]
    for row in data:
        for i in range(min(cols, len(row))):
            v = "" if row[i] is None else str(row[i])
            vlen = max((len(p) for p in v.splitlines()), default=0) if "\n" in v else len(v)
            if vlen > maxlens[i]:
                maxlens[i] = vlen
    return [min(max_px, max(min_px, (m + pad) * char_px)) for m in maxlens]


def _infer_numeric_columns(headers, data, min_ratio: float = 0.9):
    n = len(headers)
    ok = [0] * n
    tot = [0] * n
    for row in data[1:]:
        for i in range(min(n, len(row))):
            val = row[i]
            if val is None or val == "":
                continue
            tot[i] += 1
            if _is_numeric_like(val):
                ok[i] += 1
    return [(tot[i] > 0 and ok[i] / max(1, tot[i]) >= min_ratio) for i in range(n)]


# -----------------------------------------------------------------------------
# Excel-Viewer (read-only)
# -----------------------------------------------------------------------------
class ExcelViewer(tk.Toplevel):
    def __init__(self, master, xlsx_path: Path, title: str = "Database Viewer (read-only)", mode: str = "all"):
        super().__init__(master)
        self.title(title)

        # Fenster zentriert öffnen
        w, h = 1100, 700
        self.withdraw()
        self.update_idletasks()
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.deiconify()

        self.xlsx_path = xlsx_path
        self.wb = None

        # vorher wurde mode ignoriert → das ist safe, weil default gleich bleibt
        self.mode = mode or "all"

        if Sheet is None:
            messagebox.showwarning(
                "tksheet optional",
                "Optional: 'pip install tksheet' für schnellere Tabellenansicht."
            )

        # Topbar
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Label(top, text=str(xlsx_path), font=("", 9, "italic")).pack(side="left")
        ttk.Button(top, text="Reload", command=self._reload).pack(side="right")

        # Notebook
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._load_workbook()

    # ---------- Helpers ----------
    def _normkey(self, s: str | None) -> str:
        s = (s or "").lower()
        s = "".join(ch for ch in s if ch.isalnum() or ch.isspace() or ch == ":")
        return " ".join(s.split())

    def _find_sheet(self, name_like: str):
        key = self._normkey(name_like)
        for ws in self.wb.worksheets:
            if self._normkey(ws.title) == key:
                return ws
        for ws in self.wb.worksheets:
            if key in self._normkey(ws.title):
                return ws
        return None

    def _find_header_index(self, headers, *cands):
        hmap = {self._normkey(h): i for i, h in enumerate(headers)}
        for c in cands:
            k = self._normkey(c)
            if k in hmap:
                return hmap[k]
        for i, h in enumerate(headers):
            hk = self._normkey(h)
            if any(self._normkey(c) in hk for c in cands):
                return i
        return None

    # --- Hooks: bewusst im Code gelassen, um spätere Filter-Logik zu ermöglichen ---
    # Wichtig: Verhalten bleibt wie vorher (keine Filterung aktiv).
    def _category_keep_predicate(self):
        return lambda _h: True

    def _compute_allowed_codes(self):
        return None

    def _apply_district_heating_column_hides(self, headers, data):
        n = len(headers)
        drop_idx = set()

        targets = {"quellen", "entfernung klasse", "unnamed: 15", "unnamed: 16", "unnamed: 17", "unnamed: 18"}
        norm = [self._normkey(h) for h in headers]
        for i, hnorm in enumerate(norm):
            if hnorm in targets:
                drop_idx.add(i)

        # P–S = 16..19 -> zero-based 15..18
        for pos in (15, 16, 17, 18):
            if pos < n:
                drop_idx.add(pos)

        if not drop_idx:
            return headers, data

        keep_indices = [i for i in range(n) if i not in drop_idx]
        new_headers = [headers[i] for i in keep_indices]
        new_data = [[row[i] if i < len(row) else "" for i in keep_indices] for row in data]
        return new_headers, new_data

    # ---------- Workbook laden & Tabs bauen ----------
    def _on_close(self):
        try:
            if self.wb is not None:
                self.wb.close()
        except Exception as e:
            _dbg(f"workbook close failed: {e}")
        self.destroy()

    def _load_workbook(self):
        try:
            if self.wb is not None:
                try:
                    self.wb.close()
                except Exception:
                    pass
            self.wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            messagebox.showerror("Öffnen fehlgeschlagen", f"Konnte Datei nicht lesen:\n{e}")
            return

        # (Derzeit nicht genutzt – bleibt für spätere Erweiterung)
        _allowed_codes = self._compute_allowed_codes()
        _keep_col = self._category_keep_predicate()

        for ws in self.wb.worksheets:
            name = ws.title
            is_summary = self._normkey(name) == "summary"
            is_dh = self._normkey(name) == "district heating"

            frame = ttk.Frame(self.nb)
            self.nb.add(frame, text=name)
            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)

            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = [str(x) if x is not None else "" for x in next(rows_iter)]
            except StopIteration:
                headers = []

            if not headers:
                headers = [get_column_letter(i + 1) for i in range(ws.max_column or 1)]

            data = []
            max_cols = len(headers)
            for row in rows_iter:
                vals = ["" if v is None else v for v in row]
                if len(vals) < max_cols:
                    vals += [""] * (max_cols - len(vals))
                elif len(vals) > max_cols:
                    headers += [get_column_letter(i + 1) for i in range(max_cols, len(vals))]
                    max_cols = len(headers)
                    for r in range(len(data)):
                        if len(data[r]) < max_cols:
                            data[r] += [""] * (max_cols - len(data[r]))
                data.append(vals)

            # District Heating – bestimmte Spalten ausblenden
            if is_dh:
                headers, data = self._apply_district_heating_column_hides(headers, data)

            col_widths = _compute_col_pixel_widths(headers, data)
            numeric_cols = _infer_numeric_columns(headers, data)

            # Treeview fallback (Summary + wenn tksheet fehlt)
            if is_summary or Sheet is None:
                tree = ttk.Treeview(frame, show="headings")
                vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
                hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
                tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

                tree.grid(row=0, column=0, sticky="nsew")
                vsb.grid(row=0, column=1, sticky="ns")
                hsb.grid(row=1, column=0, sticky="ew")

                tree["columns"] = [f"c{i}" for i in range(len(headers))]
                for i, h in enumerate(headers):
                    cid = f"c{i}"
                    tree.heading(cid, text=h)
                    anchor = "e" if numeric_cols[i] else "w"
                    tree.column(cid, width=col_widths[i], stretch=True, anchor=anchor)

                it = iter(data)

                def insert_chunk(chunk_size=800):
                    for _ in range(chunk_size):
                        try:
                            vals = next(it)
                        except StopIteration:
                            return
                        tree.insert("", "end", values=vals)
                    frame.after(1, insert_chunk)

                insert_chunk()
                continue

            # tksheet
            sheet = Sheet(frame, data=data, headers=headers, show_x_scrollbar=True, show_y_scrollbar=True)
            sheet.grid(row=0, column=0, sticky="nsew")

            sheet.enable_bindings(
                "single_select", "row_select", "column_select",
                "row_height_resize", "column_width_resize",
                "drag_select", "select_all", "copy"
            )
            sheet.disable_bindings("edit_cell", "paste", "cut", "delete")

            sheet.set_options(
                show_vertical_grid=True,
                show_horizontal_grid=False,
                table_grid_fg="#D0D0D0"
            )

            freeze_cols = min(2, len(headers))
            sheet.set_options(freeze_columns=freeze_cols, freeze_rows=0)
            if freeze_cols >= 1:
                sheet.highlight_columns(columns=list(range(freeze_cols)), bg="#F2F2F2", fg=None, redraw=False)

            for c, w in enumerate(col_widths):
                try:
                    sheet.column_width(column=c, width=w)
                except Exception as e:
                    _dbg(f"tksheet set width failed: {e}")

            # Numerische Spalten rechts ausrichten
            try:
                for ci, isnum in enumerate(numeric_cols):
                    if not isnum:
                        continue
                    col_label = get_column_letter(ci + 1)
                    span = sheet[col_label]
                    if hasattr(span, "align"):
                        span.align("right")
                sheet.refresh()
            except Exception as e:
                _dbg(f"tksheet numeric align failed: {e}")

    def _reload(self):
        for _ in range(len(self.nb.tabs())):
            self.nb.forget(0)
        try:
            if self.wb is not None:
                self.wb.close()
        except Exception:
            pass
        self._load_workbook()


# -----------------------------------------------------------------------------
# GUI Haupt-App
# -----------------------------------------------------------------------------
_RUNNING_SENTINEL = object()  # ersetzt das alte "self.proc = ..." (gleiches Verhalten, weniger weird)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Database Creator")

        # zentriert öffnen
        w, h = 980, 680
        self.withdraw()
        self.update_idletasks()
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.deiconify()

        self.proc = None
        self.cancel_requested = False
        self.log_queue = Queue()
        self.sub_visible = False

        # Datei-Zeile
        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=(12, 6))
        ttk.Label(top, text="Excel-Datei (xlsx):").pack(side="left")
        self.excel_var = tk.StringVar(value=str(OUT_XLSX_DEFAULT))
        ent = ttk.Entry(top, textvariable=self.excel_var)
        ent.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(top, text="Durchsuchen…", command=self._browse_excel).pack(side="left")
        self.excel_var.trace_add("write", lambda *_: self._update_buttons_for_path())

        # Button-Zeile
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=12, pady=6)
        self.btn_view = ttk.Button(btns, text="View Database", command=self._open_viewer_simple)
        self.btn_start = ttk.Button(btns, text="Create Database", command=self.start_run)
        self.btn_abort = ttk.Button(btns, text="Stop", command=self.request_cancel)

        # Fortschritt
        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=12, pady=(6, 6))
        prog.columnconfigure(1, weight=1)

        ttk.Label(prog, text="Progress:").grid(row=0, column=0, sticky="w")
        self.progress = ttk.Progressbar(prog, mode="determinate", maximum=max(1, len(SCRIPTS)))
        self.progress.grid(row=0, column=1, sticky="ew", padx=(6, 0))

        self.sub_lbl = ttk.Label(prog, text="Sub-progress:")
        self.sub_bar = ttk.Progressbar(prog, mode="determinate", maximum=100)
        self.sub_info_var = tk.StringVar(value="")
        self.sub_info = ttk.Label(prog, textvariable=self.sub_info_var)
        self._sub_hide(init=True)

        # Log
        logf = ttk.Frame(self)
        logf.pack(fill="both", expand=True, padx=12, pady=(0, 6))
        self.txt = tk.Text(logf, wrap="word", height=20)
        vsb = ttk.Scrollbar(logf, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=vsb.set)
        self.txt.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")
        self.txt.tag_configure("royal", foreground="#4169E1")

        # Status
        self.status_var = tk.StringVar(value="Bereit.")
        ttk.Label(self, textvariable=self.status_var, anchor="w").pack(fill="x", padx=12, pady=(0, 8))

        # Bottom-right "Back to Overview"
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=12, pady=(0, 10))
        ttk.Label(bottom, text="").pack(side="left", fill="x", expand=True)
        ttk.Button(bottom, text="Back to Overview", command=self.back_to_overview).pack(side="right")

        # Log-Poller
        self.after(80, self._poll_log_queue)

        self._buttons_frame = btns
        self.after(0, lambda: center_on_screen(self))
        self._update_buttons_for_path()

    # ---------- Viewer ----------
    def _open_viewer_simple(self):
        path = Path(self.excel_var.get()).expanduser()
        if not path.exists():
            messagebox.showwarning("Keine Datei", "Die angegebene Excel-Datei existiert nicht.")
            return
        ExcelViewer(self, path, title="Database Viewer (read-only)", mode="all")

    # ---------- Buttons Dynamik ----------
    def _update_buttons_for_path(self):
        for w in self._buttons_frame.winfo_children():
            w.pack_forget()

        if self.proc is not None:
            self.btn_abort.config(state="normal")
            self.btn_abort.pack(side="left")
            return

        xlsx = Path(self.excel_var.get()).expanduser()
        self.btn_start.config(state="normal")
        self.btn_start.pack(side="left")
        if xlsx.exists():
            self.btn_view.config(state="normal")
            self.btn_view.pack(side="left", padx=(8, 0))

    # ---------- UI helpers ----------
    def _sub_show(self):
        if not self.sub_visible:
            self.sub_lbl.grid(row=1, column=0, sticky="w", pady=(8, 0))
            self.sub_bar.grid(row=1, column=1, sticky="ew", padx=(6, 0), pady=(8, 0))
            self.sub_info.grid(row=2, column=1, sticky="w", padx=(6, 0), pady=(2, 0))
            self.sub_visible = True

    def _sub_hide(self, init: bool = False):
        if init or self.sub_visible:
            self.sub_lbl.grid_remove()
            self.sub_bar.grid_remove()
            self.sub_info.grid_remove()
            self.sub_visible = False
        self.sub_info_var.set("")
        self.sub_bar["value"] = 0

    # ---------- Callbacks ----------
    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title="Excel-Datei wählen",
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")]
        )
        if path:
            self.excel_var.set(path)

    def start_run(self):
        if self.proc is not None:
            return

        self.cancel_requested = False
        self.txt.delete("1.0", "end")
        self.progress["value"] = 0
        self._sub_hide()
        self.status_var.set("Starte Pipeline …")

        # Buttons sofort umschalten (gleiches Verhalten wie vorher, nur ohne Ellipsis)
        self.proc = _RUNNING_SENTINEL
        self._update_buttons_for_path()
        self.proc = None

        threading.Thread(target=self._worker, daemon=True).start()

    def request_cancel(self):
        self.cancel_requested = True
        self.status_var.set("Abbruch angefordert …")
        if self.proc and hasattr(self.proc, "poll") and self.proc.poll() is None:
            try:
                self.proc.kill()
            except Exception:
                pass

    # ---------- Progress-Erkennung ----------
    ANSI_RE = re.compile(r"\x1b\[[0-9;]*[A-Za-z]")
    BRACKET_BAR_RE = re.compile(r"^\s*\[[^\]]+\]\s*")
    TQDM_PREFIX_RE = re.compile(r"^\s*\d{1,3}%\|\S+\|\s*\d+/\d+")

    @classmethod
    def _strip_ansi(cls, s: str) -> str:
        return cls.ANSI_RE.sub("", s)

    @classmethod
    def _parse_percent(cls, line: str):
        line = cls._strip_ansi(line)
        m = re.search(r'(\d{1,3}(?:[.,]\d+)?)\s*%', line)
        if m:
            try:
                val = float(m.group(1).replace(',', '.'))
                return None if val < 0 else min(val, 100.0)
            except ValueError:
                pass
        m = re.search(r'(\d+)\s*/\s*(\d+)', line)
        if m:
            num, den = int(m.group(1)), int(m.group(2))
            if den > 0:
                return max(0.0, min(100.0, (num / den) * 100.0))
        m = re.search(r'\b(\d{1,3})%\|\S+\|\s*\d+/\d+', line)  # tqdm
        if m:
            return float(m.group(1))
        return None

    @classmethod
    def _clean_progress_text(cls, line: str) -> str:
        s = cls._strip_ansi(line)
        s = cls.BRACKET_BAR_RE.sub("", s)
        s = cls.TQDM_PREFIX_RE.sub("", s)
        return s.strip().replace("\r", "")

    @classmethod
    def _is_progress_update(cls, line: str):
        s = cls._strip_ansi(line).strip()
        if cls._parse_percent(s) is not None:
            return True
        bar_chars = set("[]()=-><#|/\\_.:%▏▎▍▌▋▊▉█░▒")
        if s and len(s) >= 10 and (sum(ch in s for ch in bar_chars) / len(s) > 0.55):
            return True
        if re.match(r'^\s*[\[\|\(].*[\]\|\)]\s*$', s):
            return True
        return False

    # ---------- Worker ----------
    def _worker(self):
        self.after(0, self._update_buttons_for_path)

        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"

        excel_path = Path(self.excel_var.get())

        try:
            for idx, script in enumerate(SCRIPTS, 1):
                if self.cancel_requested:
                    break

                self._log(f"\n▶️  Starte {script.name} …\n")
                self._status("Läuft: " + script.name)
                self._sub_hide()

                self.proc = subprocess.Popen(
                    [sys.executable, "-u", str(script)],
                    cwd=str(BASE),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    env=env,
                    encoding="utf-8",
                    errors="replace",
                )
                self.after(0, self._update_buttons_for_path)

                assert self.proc.stdout is not None
                buf = ""
                while True:
                    ch = self.proc.stdout.read(1)
                    if ch == "" and self.proc.poll() is not None:
                        if buf:
                            if self._is_progress_update(buf):
                                self._update_subprogress(buf)
                            else:
                                self._flush_line(buf)
                            buf = ""
                        break
                    if not ch:
                        time.sleep(0.01)
                        continue

                    if ch == "\r":
                        line = buf
                        buf = ""
                        if line:
                            self._update_subprogress(line)
                        continue

                    if ch == "\n":
                        line = buf
                        buf = ""
                        if self._is_progress_update(line):
                            self._update_subprogress(line)
                        else:
                            self._flush_line(line)
                        continue

                    buf += ch
                    if len(buf) > 4000:
                        if self._is_progress_update(buf):
                            self._update_subprogress(buf)
                        else:
                            self._flush_line(buf)
                        buf = ""

                ret = self.proc.wait()
                self.proc = None
                if self.cancel_requested:
                    break

                if ret != 0:
                    self._log(f"\n❌ {script.name} beendet sich mit Code {ret}\n")
                    self._status("Fehler – siehe Log")
                    break

                self._log(f"✅ {script.name} fertig\n")

                try:
                    autosize_excel(excel_path)
                except Exception as e:
                    self._log(f"⚠️  Konnte Excel-Formatierung nicht anwenden: {e}\n")

                self._sub_hide()
                self._set_progress(idx)

            if self.cancel_requested:
                self._log("\n⛔ Abgebrochen.\n")
                self._status("Abgebrochen.")
            else:
                self._log("\n🎉 Pipeline abgeschlossen.\n")
                self._status("Fertig.")
        finally:
            self.proc = None
            self.after(0, self._update_buttons_for_path)

    # ---------- Thread-sichere UI-Helfer ----------
    def _flush_line(self, line: str):
        if line:
            clean = self._strip_ansi(line)
            self._log(clean + ("\n" if not clean.endswith("\n") else ""))

    def _log(self, msg: str):
        self.log_queue.put(msg)

    def _poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                line = msg.lstrip()
                is_start = line.startswith("▶️  Starte ")
                is_end_ok = (line.startswith("✅ ") and " fertig" in line)
                is_end_err = (line.startswith("❌ ") and "beendet sich mit Code" in line)

                if is_start or is_end_ok or is_end_err:
                    start_idx = self.txt.index("end")
                    self.txt.insert("end", msg)
                    end_idx = self.txt.index("end")
                    self.txt.tag_add("royal", start_idx, end_idx)
                else:
                    self.txt.tag_remove("royal", "end-1c", "end")
                    self.txt.insert("end", msg)

                self.txt.see("end")
        except Empty:
            pass
        self.after(80, self._poll_log_queue)

    def _set_progress(self, value: int):
        self.after(0, lambda: self.progress.configure(value=value))

    def _status(self, text: str):
        self.after(0, lambda: self.status_var.set(text))

    def _update_subprogress(self, line: str):
        pct = self._parse_percent(line)
        info = self._clean_progress_text(line)
        self.after(0, lambda: self._apply_subprogress(pct, info))

    def back_to_overview(self):
        # Schließe nur das aktuelle Fenster, ohne ##_main.py neu zu starten
        self.destroy()

    def _apply_subprogress(self, pct: float | None, info: str):
        self._sub_show()
        if info:
            self.sub_info_var.set(info[:300])
        if pct is not None:
            if str(self.sub_bar.cget("mode")) != "determinate":
                self.sub_bar.configure(mode="determinate", maximum=100)
            self.sub_bar["value"] = pct


if __name__ == "__main__":
    App().mainloop()
