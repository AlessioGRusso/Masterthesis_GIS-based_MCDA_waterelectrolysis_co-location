
import os
import sys
import math
import time
import glob
import requests
from typing import List, Tuple, Dict, Optional

import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, Polygon, MultiPolygon
from shapely.ops import unary_union
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pyproj import Transformer

# osmnx Kompatibilität - verschiedene Versionen haben unterschiedliche APIs
import osmnx as ox

try:
    from osmnx.features import features_from_point as ox_features_from_point
except Exception:
    try:
        from osmnx.geometries import geometries_from_point as ox_features_from_point
    except Exception:
        ox_features_from_point = None

# Konfiguration
EXCEL_CANDIDATES = [
    os.path.join("Output", "WWTP_TP_Database.xlsx"),
    os.path.join("Output", "UWWTD_TP_Database.xlsx"),
    "WWTP_TP_Database.xlsx",
    "UWWTD_TP_Database.xlsx",
]
SOURCE_SHEET_NAME = "General Data"
TARGET_SHEET_NAME = "Risks"

# CSV-Cache für schnellere Wiederholung
CACHE_DIR = os.path.join("Output", "Risks")
CACHE_FILE = os.path.join(CACHE_DIR, "risks_cache.csv")

# GPKG-Export (alle Anlagen, ohne Löschung)
GPKG_DIR = os.path.join("Output", "WWTP Geopackages")
GPKG_FILE = os.path.join(GPKG_DIR, "WWTPS_after_RAF.gpkg")

# EEA-Services für Hochwasser-Daten
EEA_POTENTIAL_QUERY = (
    "https://water.discomap.eea.europa.eu/arcgis/rest/services/"
    "Flood/Potential_flood_prone_area/MapServer/0/query"
)
# KNOCKOUT-Kriterium: FloodsRiskZone_WM
EEA_FLOODS_RISK_ZONE_QUERY = (
    "https://water.discomap.eea.europa.eu/arcgis/rest/services/"
    "FloodsDirective/FloodsRiskZone_WM/MapServer/2/query"
)

# EEA-Services (Protected Areas)
PA_SRC_NATDA = {
    "tag": "NATDA",
    "service": "https://bio.discomap.eea.europa.eu/arcgis/rest/services/ProtectedSites/NatDAv22_Dyna_WM/MapServer",
    "layers": None,
    "layer_names": None,
}
PA_SRC_N2K = {
    "tag": "N2K",
    "service": "http://bio.discomap.eea.europa.eu/arcgis/rest/services/ProtectedSites/Natura2000_Dyna_WM/MapServer",
    "layers": None,
    "layer_names": None,
}

REQUEST_SLEEP = 0.05

# *** 5-km Suchradien ***
FLOOD_SEARCH_RADIUS_M = 5000
PA_SEARCH_RADIUS_M = 5000
OSM_POINT_DIST_M = 5000
# OSM-Kachelgröße (~5–6 km Kante)
OSM_TILE_SIZE_DEG = 0.05

HEADERS = {"User-Agent": "wwtp-risks/4.8 (+python-requests)"}

CRS_WGS84 = "EPSG:4326"
CRS_EUROPE = "EPSG:3035"


# =========================
# Utils
# =========================
def detect_excel_path() -> Optional[str]:
    for p in EXCEL_CANDIDATES:
        if os.path.exists(p):
            return p
    return None


def normalize_header(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def find_header_map(ws) -> Dict[str, int]:
    hmap = {}
    for col in range(1, ws.max_column + 1):
        v = normalize_header(ws.cell(row=1, column=col).value)
        if v:
            hmap[v.lower()] = col
    return hmap


def to_float_or_none(x):
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None


_transformers_cache: Dict[Tuple[str, str], Transformer] = {}


def get_transformer(src: str, dst: str) -> Transformer:
    key = (src, dst)
    if key not in _transformers_cache:
        _transformers_cache[key] = Transformer.from_crs(src, dst, always_xy=True)
    return _transformers_cache[key]


def reproject_geom(geom, src: str = CRS_WGS84, dst: str = CRS_EUROPE):
    from shapely.ops import transform as s_transform
    t = get_transformer(src, dst)
    return s_transform(t.transform, geom)


def min_distance_geom_to_union(geom_wgs84, union_europe):
    if union_europe is None or geom_wgs84 is None:
        return math.nan
    ge_eu = reproject_geom(geom_wgs84, CRS_WGS84, CRS_EUROPE)
    if ge_eu.intersects(union_europe) or union_europe.contains(ge_eu):
        return 0.0
    return float(ge_eu.distance(union_europe))


# =========================
# ArcGIS Helpers
# =========================
def arcgis_query_point_buffer(query_url: str, lat: float, lon: float, radius_m: int,
                              session: requests.Session, out_sr: int = 4326,
                              out_fields: str = "*"):
    params = {
        "where": "1=1",
        "geometry": f"{lon},{lat}",
        "geometryType": "esriGeometryPoint",
        "inSR": 4326,
        "spatialRel": "esriSpatialRelIntersects",
        "distance": radius_m,
        "units": "esriSRUnit_Meter",
        "outFields": out_fields,
        "returnGeometry": "true",
        "outSR": out_sr,
        "f": "json",
    }
    r = session.get(query_url, params=params, timeout=40)
    r.raise_for_status()
    js = r.json()
    if "error" in js:
        return []
    return js.get("features", []) or []


def arcgis_service_info(service_url: str, session: requests.Session):
    url = f"{service_url}?f=json"
    r = session.get(url, timeout=40)
    r.raise_for_status()
    return r.json()


def arcgis_polygon_layers(service_url: str, session: requests.Session):
    js = arcgis_service_info(service_url, session)
    layers = js.get("layers", []) or []
    poly_ids = []
    names = {}
    for lyr in layers:
        if isinstance(lyr, dict):
            if lyr.get("geometryType") == "esriGeometryPolygon":
                lid = int(lyr["id"])
                poly_ids.append(lid)
                names[lid] = str(lyr.get("name", ""))
    return poly_ids, names


def arcgis_polygon_layers_cached(src_def: dict, session: requests.Session):
    if src_def.get("layers") is None or src_def.get("layer_names") is None:
        lids, names = arcgis_polygon_layers(src_def["service"], session)
        src_def["layers"] = lids
        src_def["layer_names"] = names
    return src_def["layers"] or []


def esri_polygon_to_shapely(esri_geom: dict):
    if not esri_geom:
        return None
    rings = esri_geom.get("rings")
    if not rings:
        return None
    try:
        poly = Polygon(rings[0], holes=rings[1:] if len(rings) > 1 else None)
        if not poly.is_valid:
            poly = poly.buffer(0)
        if poly.is_empty:
            return None
        return poly
    except Exception:
        return None


def fetch_arcgis_polygons_union(query_url: str, lat: float, lon: float, radius_m: int,
                                session: requests.Session):
    feats = arcgis_query_point_buffer(query_url, lat, lon, radius_m, session, out_sr=4326)
    geoms = []
    for f in feats:
        g = esri_polygon_to_shapely(f.get("geometry"))
        if g is not None:
            geoms.append(g)
    if not geoms:
        return None
    geoms_eu = [reproject_geom(g, CRS_WGS84, CRS_EUROPE) for g in geoms]
    return unary_union(geoms_eu)


def prepare_n2k_sources(session: requests.Session) -> List[dict]:
    lids, names = arcgis_polygon_layers(PA_SRC_N2K["service"], session)
    spa_layers = []
    hab_layers = []
    for lid in lids:
        nm = names.get(lid, "").lower()
        if any(k in nm for k in ["birds", "spa", "special protection area"]):
            spa_layers.append(lid)
        elif any(k in nm for k in ["habitat", "sac", "sci", "habitats directive"]):
            hab_layers.append(lid)
    srcs = []
    if spa_layers:
        srcs.append({"tag": "N2K SPA", "service": PA_SRC_N2K["service"], "layers": spa_layers, "layer_names": names})
    if hab_layers:
        srcs.append(
            {"tag": "N2K SAC/SCI", "service": PA_SRC_N2K["service"], "layers": hab_layers, "layer_names": names})
    if not srcs:
        srcs.append({"tag": "N2K", "service": PA_SRC_N2K["service"], "layers": lids, "layer_names": names})
    return srcs


# =========================
# Typableitung (nur Info/Report)
# =========================
def _get(attrs: dict, *names: str) -> str:
    for n in names:
        if n in attrs and attrs[n] not in (None, ""):
            return str(attrs[n]).strip()
        if n.lower() in attrs and attrs[n.lower()] not in (None, ""):
            return str(attrs[n.lower()]).strip()
    return ""


IUCN_CATS = ["IA", "IB", "I", "II", "III", "IV", "V", "VI"]


def derive_pa_type(src_tag: str, attrs: dict) -> str:
    tag = (src_tag or "").strip()
    if tag == "NATDA":
        abbr = _get(attrs, "designationTypeCode", "DESIG_ABBR", "DESIG_ABBREV")
        iucn = _get(attrs, "iucnCategory", "iucnCatCod", "IUCN_CAT")
        descr = _get(attrs, "iucnDescri", "DESIG_ENG", "DESIGNATION")
        parts = [tag]
        if abbr: parts.append(abbr)
        if iucn:
            parts.append(iucn)
        elif descr:
            parts.append(descr)
        return " ".join(parts).strip()
    if tag.startswith("N2K"):
        if tag in ("N2K SPA", "N2K SAC/SCI"):
            return tag
        st = (_get(attrs, "SITETYPE") or "").upper()
        if st.startswith("A"): return "N2K SPA"
        if st.startswith("B"): return "N2K SAC/SCI"
        desig = " ".join((_get(attrs, "DESIG_ABBR"), _get(attrs, "DESIGNATION"), _get(attrs, "DESIG_ENG"))).upper()
        if "SPA" in desig: return "N2K SPA"
        if "SAC" in desig or "SCI" in desig: return "N2K SAC/SCI"
        return "N2K"
    return tag or "PA"


# =========================
# Schutzgebiete: Union + Attribute
# =========================
def fetch_protected_areas_union_and_attr(lat: float, lon: float, radius_m: int,
                                         src_def: dict, session: requests.Session):
    layers = arcgis_polygon_layers_cached(src_def, session)
    if not layers:
        return None, []

    all_geoms_eu = []
    geom_attr_pairs: List[Tuple[Polygon, str, str]] = []

    for lyr_id in layers:
        qurl = f"{src_def['service'].rstrip('/')}/{lyr_id}/query"
        feats = arcgis_query_point_buffer(qurl, lat, lon, radius_m, session, out_sr=4326, out_fields="*")
        for f in feats:
            poly = esri_polygon_to_shapely(f.get("geometry"))
            if poly is None:
                continue
            poly_eu = reproject_geom(poly, CRS_WGS84, CRS_EUROPE)
            all_geoms_eu.append(poly_eu)

            attrs = f.get("attributes", {}) or {}
            label = _get(attrs, "SITENAME", "SITE_NAME", "NAME", "DESIGNATION", "DESIG_ENG", "SITECODE")
            pa_type_full = derive_pa_type(src_def.get("tag", ""), attrs)
            geom_attr_pairs.append((poly_eu, label, pa_type_full))
        time.sleep(REQUEST_SLEEP)

    if not all_geoms_eu:
        return None, []
    return unary_union(all_geoms_eu), geom_attr_pairs


# =========================
# Tile-Cache (Flood + PA + OSM)
# =========================
def tile_key(lat: float, lon: float, size_deg: float = OSM_TILE_SIZE_DEG) -> Tuple[float, float]:
    ky = round(lat / size_deg) * size_deg
    kx = round(lon / size_deg) * size_deg
    return ky, kx


_OSM_CACHE: Dict[Tuple[float, float], Tuple[object, object]] = {}
_FLOOD_TILE_CACHE: Dict[Tuple[float, float], object] = {}
_FLOOD_RISK_ZONE_CACHE: Dict[Tuple[float, float], object] = {}
_PA_TILE_CACHE: Dict[Tuple[float, float], Tuple[object, List[Tuple[Polygon, str, str]]]] = {}


def get_osm_unions_for_tile(lat: float, lon: float, radius_m: int):
    if ox_features_from_point is None:
        raise ImportError(
            "osmnx 'features_from_point' / 'geometries_from_point' nicht verfügbar. Bitte osmnx aktualisieren.")
    k = tile_key(lat, lon)
    if k in _OSM_CACHE:
        return _OSM_CACHE[k]

    tags_res = {"landuse": ["residential"]}
    gdf_res = ox_features_from_point((lat, lon), tags_res, dist=radius_m)
    if (gdf_res is not None) and (len(gdf_res) > 0):
        gdf_res = gdf_res[gdf_res.geometry.type.isin(["Polygon", "MultiPolygon"])].copy()
        res_union = unary_union([reproject_geom(g, CRS_WGS84, CRS_EUROPE) for g in gdf_res.geometry])
    else:
        res_union = None

    tags_leisure = {"leisure": [
        "park", "pitch", "stadium", "playground", "garden", "recreation_ground",
        "sports_centre", "swimming_pool", "golf_course", "marina", "beach_resort",
        "track", "ice_rink",
    ]}
    gdf_lea = ox_features_from_point((lat, lon), tags_leisure, dist=radius_m)
    if (gdf_lea is not None) and (len(gdf_lea) > 0):
        gdf_lea = gdf_lea[gdf_lea.geometry.type.isin(["Polygon", "MultiPolygon"])].copy()
        lea_union = unary_union([reproject_geom(g, CRS_WGS84, CRS_EUROPE) for g in gdf_lea.geometry])
    else:
        lea_union = None

    _OSM_CACHE[k] = (res_union, lea_union)
    return _OSM_CACHE[k]


def get_flood_union_for_tile(lat: float, lon: float, session: requests.Session):
    """Holt nur Potential_flood_prone_area für Distanzberechnung"""
    k = tile_key(lat, lon)
    if k in _FLOOD_TILE_CACHE:
        return _FLOOD_TILE_CACHE[k]
    qlat, qlon = k
    radius = 15000
    union = None
    try:
        union = fetch_arcgis_polygons_union(EEA_POTENTIAL_QUERY, qlat, qlon, radius, session)
    except Exception:
        union = None
    _FLOOD_TILE_CACHE[k] = union
    return union


def get_flood_risk_zone_union_for_tile(lat: float, lon: float, session: requests.Session):
    """Holt FloodsRiskZone_WM für Knockout-Kriterium"""
    k = tile_key(lat, lon)
    if k in _FLOOD_RISK_ZONE_CACHE:
        return _FLOOD_RISK_ZONE_CACHE[k]
    qlat, qlon = k
    radius = 15000
    union = None
    try:
        union = fetch_arcgis_polygons_union(EEA_FLOODS_RISK_ZONE_QUERY, qlat, qlon, radius, session)
    except Exception:
        union = None
    _FLOOD_RISK_ZONE_CACHE[k] = union
    return union


def get_pa_union_and_pairs_for_tile(lat: float, lon: float, session: requests.Session, n2k_sources: List[dict]):
    k = tile_key(lat, lon)
    if k in _PA_TILE_CACHE:
        return _PA_TILE_CACHE[k]
    qlat, qlon = k
    radius = 15000

    all_geoms = []
    all_pairs: List[Tuple[Polygon, str, str]] = []

    try:
        nat_union, nat_pairs = fetch_protected_areas_union_and_attr(qlat, qlon, radius, PA_SRC_NATDA, session)
        if nat_union is not None:
            all_geoms.append(nat_union);
            all_pairs.extend(nat_pairs)
    except Exception:
        pass

    for src_n2k in n2k_sources:
        try:
            n2k_union, n2k_pairs = fetch_protected_areas_union_and_attr(qlat, qlon, radius, src_n2k, session)
            if n2k_union is not None:
                all_geoms.append(n2k_union);
                all_pairs.extend(n2k_pairs)
        except Exception:
            pass

    union_all = unary_union(all_geoms) if all_geoms else None
    _PA_TILE_CACHE[k] = (union_all, all_pairs)
    return _PA_TILE_CACHE[k]


# =========================
# Distanz + Label/Typ zu PA
# =========================
def nearest_distance_and_label_from_geom(src_geom_wgs84,
                                         union_eu,
                                         geom_attr_pairs: List[Tuple[Polygon, str, str]]):
    """
    Liefert (Distanz in m, Label, Typ) zur nächstgelegenen/überdeckten Schutzgebiets-Geometrie.
    Distanz = 0, wenn die Geometrie im Union liegt oder es eine echte Schnittmenge gibt.
    """
    if union_eu is None or src_geom_wgs84 is None:
        return math.nan, "", ""
    src_eu = reproject_geom(src_geom_wgs84, CRS_WGS84, CRS_EUROPE)
    # liegt drin / schneidet
    if src_eu.intersects(union_eu) or union_eu.contains(src_eu):
        for geom, lab, ptype in geom_attr_pairs:
            if geom.intersects(src_eu) or geom.contains(src_eu):
                return 0.0, lab or "", ptype or ""
        return 0.0, "", ""
    # sonst: nächste Distanz + zugehöriges Label/Typ
    min_d = float("inf")
    min_label = ""
    min_type = ""
    for geom, lab, ptype in geom_attr_pairs:
        d = src_eu.distance(geom)
        if d < min_d:
            min_d = d
            min_label = lab or ""
            min_type = ptype or ""
    if min_d == float("inf"):
        return math.nan, "", ""
    return float(min_d), min_label, min_type


# =========================
# Fortschritt
# =========================
def _render_bar(progress: float, width: int = 28) -> str:
    progress = max(0.0, min(1.0, progress))
    filled = int(round(width * progress))
    return "[" + "#" * filled + "-" * (width - filled) + "]"


def _print_progress(done_full: int, total: int, stage_done: int,
                    flood_done: bool, res_done: bool, pa_done: bool,
                    extra_msg: str = ""):
    part = stage_done / 3.0
    progress = ((done_full + part) / total) if total > 0 else 1.0
    bar = _render_bar(progress)
    pct = int(progress * 100)
    s_f = "ok" if flood_done else "-"
    s_r = "ok" if res_done else "-"
    s_p = "ok" if pa_done else "-"
    line = f"{bar} {pct:3d}%  {done_full}/{total}  Flood:{s_f}  Residential:{s_r}  PA:{s_p}  {extra_msg}"
    sys.stdout.write("\r" + line[:160]);
    sys.stdout.flush()


def _print_progress_newline():
    sys.stdout.write("\n");
    sys.stdout.flush()


# =========================
# WWTP GPKG
# =========================
def autodetect_wwtp_gpkg() -> Optional[str]:
    patterns = [os.path.join("Output", "WWTP Geopackages", "WWTPS_Shapes.gpkg")]
    candidates = []
    for pat in patterns:
        candidates.extend(glob.glob(pat))
    if not candidates:
        return None
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def load_wwtp_polygons(gpkg_path: str, layer: Optional[str] = None) -> gpd.GeoDataFrame:
    if layer is None:
        import fiona
        poly_layers = []
        for lyr in fiona.listlayers(gpkg_path):
            try:
                g = gpd.read_file(gpkg_path, layer=lyr)
                if not g.empty and g.geometry.type.isin(["Polygon", "MultiPolygon"]).any():
                    poly_layers.append(lyr)
            except Exception:
                pass
        if not poly_layers:
            raise ValueError("Kein Polygon-Layer in GPKG gefunden.")
        layer = poly_layers[0]
    gdf = gpd.read_file(gpkg_path, layer=layer)
    gdf = gdf[gdf.geometry.type.isin(["Polygon", "MultiPolygon"])].copy()
    gdf = gdf.to_crs(4326)  # WGS84
    return gdf, layer


# =========================
# MAIN
# =========================
def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--wwtp_gpkg", default=None, help="Pfad zur GPKG mit WWTP-Polygonen")
    ap.add_argument("--wwtp_layer", default=None, help="Layername (optional)")
    args = ap.parse_args()

    xlsx_path = detect_excel_path()
    if not xlsx_path:
        raise FileNotFoundError("Excel-Datei nicht gefunden (Output/*_TP_Database.xlsx).")

    gpkg_path = args.wwtp_gpkg or autodetect_wwtp_gpkg()
    if not gpkg_path or not os.path.exists(gpkg_path):
        raise FileNotFoundError("WWTP-GPKG nicht gefunden. Übergib --wwtp_gpkg <pfad>.")

    wb = load_workbook(xlsx_path)

    src = None
    for name in wb.sheetnames:
        if name.strip().lower() == SOURCE_SHEET_NAME.strip().lower():
            src = wb[name]
            break
    if src is None:
        raise ValueError(f"Sheet '{SOURCE_SHEET_NAME}' nicht gefunden.")

    hmap = find_header_map(src)
    if not hmap:
        raise ValueError("Keine Kopfzeile gefunden (Zeile 1).")
    lat_col = hmap.get("latitude")
    lon_col = hmap.get("longitude")
    if not lat_col or not lon_col:
        raise ValueError("Header 'Latitude' und/oder 'Longitude' nicht gefunden.")
    code_col = hmap.get("uwwtd code") or hmap.get("uwwtd_code") or hmap.get("code") or hmap.get("id")

    name_col = (
            hmap.get("name") or
            hmap.get("plant name") or
            hmap.get("wwtp name") or
            hmap.get("facility name") or
            hmap.get("facility") or
            hmap.get("agglomeration") or
            hmap.get("uwwtp name")
    )

    wwtp_gdf, used_layer = load_wwtp_polygons(gpkg_path, args.wwtp_layer)

    # Cache laden
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_exists = os.path.exists(CACHE_FILE)
    cache_map: Dict[str, Tuple[object, object, object, str]] = {}
    if cache_exists:
        try:
            df_cache = pd.read_csv(CACHE_FILE)
            cols = {c.strip(): c for c in df_cache.columns}
            c_lat = cols.get("Latitude") or cols.get("latitude") or "Latitude"
            c_lon = cols.get("Longitude") or cols.get("longitude") or "Longitude"
            c_flood = cols.get("Distance to possible flood area [m]")
            c_res = cols.get("Distance to residential area [m]")
            c_pa = cols.get("Distance to protected area [m]")
            c_ptype = cols.get("PA_Type") or "PA_Type"
            if c_flood and c_res and c_pa and c_ptype:
                for _, row in df_cache.iterrows():
                    try:
                        lat_v = float(str(row[c_lat]).replace(",", "."))
                        lon_v = float(str(row[c_lon]).replace(",", "."))
                        key = f"{lat_v:.6f}|{lon_v:.6f}"
                        cache_map[key] = (row[c_flood], row[c_res], row[c_pa], str(row.get(c_ptype, "")))
                    except Exception:
                        continue
            else:
                cache_exists = False
        except Exception:
            cache_exists = False

    session = requests.Session()
    session.headers.update(HEADERS)

    _ = arcgis_polygon_layers_cached(PA_SRC_NATDA, session)
    n2k_sources = prepare_n2k_sources(session)

    total_rows = max(src.max_row - 1, 0)

    result_map: Dict[str, Tuple[float, float, float, str]] = dict(cache_map)
    records_list = []
    knockout_rows = []  # Zeilen, die wegen FloodsRiskZone gelöscht werden

    done_full = 0
    _print_progress(done_full, total_rows, 0, False, False, False, "Starte …")

    def nearest_polygon(pt_wgs84: Point) -> Optional[Polygon]:
        if wwtp_gdf.empty:
            return None
        try:
            idx = list(wwtp_gdf.sindex.nearest(pt_wgs84.bounds, return_all=False))[0]
        except Exception:
            cand_idx = list(wwtp_gdf.sindex.intersection(pt_wgs84.bounds))
            if not cand_idx:
                return None
            dists = [(i, pt_wgs84.distance(wwtp_gdf.geometry.iloc[i])) for i in cand_idx]
            idx = min(dists, key=lambda x: x[1])[0]
        return wwtp_gdf.geometry.iloc[idx]

    for r in range(2, src.max_row + 1):
        lat = to_float_or_none(src.cell(row=r, column=lat_col).value)
        lon = to_float_or_none(src.cell(row=r, column=lon_col).value)

        row_name = ""
        if name_col:
            nv = src.cell(row=r, column=name_col).value
            if nv is not None:
                row_name = str(nv).strip()

        flood_done = res_done = pa_done = False
        stage_done = 0

        if (lat is None) or (lon is None):
            records_list.append({
                "Latitude": lat,
                "Longitude": lon,
                "Distance to possible flood area [m]": math.nan,
                "Distance to residential area [m]": math.nan,
                "Distance to protected area [m]": math.nan,
                "PA_Type": "",
                "Name": row_name,
            })
            done_full += 1
            _print_progress(done_full, total_rows, 0, False, False, False, f"Zeile {r - 1} ohne Koordinaten")
            continue

        key = f"{float(lat):.6f}|{float(lon):.6f}"
        if key in result_map:
            dflood, dres, dpa_cached, ptype_cached = result_map[key]
            records_list.append({
                "Latitude": lat,
                "Longitude": lon,
                "Distance to possible flood area [m]": dflood,
                "Distance to residential area [m]": dres,
                "Distance to protected area [m]": dpa_cached,
                "PA_Type": ptype_cached,
                "Name": row_name,  # Name im Cache
            })
            flood_done = res_done = pa_done = True
            stage_done = 3
            _print_progress(done_full, total_rows, stage_done, flood_done, res_done, pa_done, f"Zeile {r - 1} (Cache)")

            done_full += 1
            _print_progress(done_full, total_rows, 0, False, False, False, f"Abgeschlossen {done_full}/{total_rows}")
            continue

        # --- Tile-Caches ---
        ky, kx = tile_key(lat, lon)
        flood_union = get_flood_union_for_tile(ky, kx, session)
        flood_risk_zone_union = get_flood_risk_zone_union_for_tile(ky, kx, session)
        pa_union, pa_pairs = get_pa_union_and_pairs_for_tile(ky, kx, session, n2k_sources)
        res_union, lea_union = get_osm_unions_for_tile(ky, kx, OSM_POINT_DIST_M)

        pt = Point(float(lon), float(lat))
        src_poly = nearest_polygon(pt) or pt
        src_geom = src_poly

        # KNOCKOUT-CHECK: Liegt die Anlage in FloodsRiskZone_WM?
        is_in_flood_risk_zone = False
        if flood_risk_zone_union is not None:
            src_geom_eu = reproject_geom(src_geom, CRS_WGS84, CRS_EUROPE)
            if src_geom_eu.intersects(flood_risk_zone_union) or flood_risk_zone_union.contains(src_geom_eu):
                is_in_flood_risk_zone = True
                knockout_rows.append(r)
                done_full += 1
                _print_progress(done_full, total_rows, 0, False, False, False, f"Zeile {r - 1} KNOCKOUT (FloodsRiskZone)")
                continue  # Diese Anlage wird nicht in die Ergebnisse aufgenommen

        dist_flood = min_distance_geom_to_union(src_geom, flood_union) if flood_union is not None else math.nan
        flood_done = True;
        stage_done = 1
        _print_progress(done_full, total_rows, stage_done, flood_done, res_done, pa_done, f"Zeile {r - 1} Flood")

        d_res = min_distance_geom_to_union(src_geom, res_union) if res_union is not None else math.nan
        d_lea = min_distance_geom_to_union(src_geom, lea_union) if lea_union is not None else math.nan
        candidates = [d for d in (d_res, d_lea) if not math.isnan(d)]
        dist_res_lea = min(candidates) if candidates else math.nan
        res_done = True;
        stage_done = 2
        _print_progress(done_full, total_rows, stage_done, flood_done, res_done, pa_done, f"Zeile {r - 1} Residential")

        d_pa = math.nan
        pa_type = ""
        if pa_union is not None:
            d_pa, lab, pa_type = nearest_distance_and_label_from_geom(src_geom, pa_union, pa_pairs)
        pa_done = True;
        stage_done = 3
        _print_progress(done_full, total_rows, stage_done, flood_done, res_done, pa_done, f"Zeile {r - 1} PA")

        result_map[key] = (dist_flood, dist_res_lea, d_pa, pa_type)
        records_list.append({
            "Latitude": lat,
            "Longitude": lon,
            "Distance to possible flood area [m]": dist_flood,
            "Distance to residential area [m]": dist_res_lea,
            "Distance to protected area [m]": d_pa,
            "PA_Type": pa_type,
            "Name": row_name,
        })

        done_full += 1
        _print_progress(done_full, total_rows, 0, False, False, False, f"Abgeschlossen {done_full}/{total_rows}")

    _print_progress_newline()

    # Knockout-Zusammenfassung
    plants_after = total_rows - len(knockout_rows)
    print(f"Plants: {total_rows} → {plants_after}")

    # 1) CSV-Cache schreiben
    try:
        os.makedirs(CACHE_DIR, exist_ok=True)
        df_out = pd.DataFrame(records_list)
        df_out.to_csv(CACHE_FILE, index=False)
    except Exception as e:
        pass

    # 2) RISKS-Blatt

    if TARGET_SHEET_NAME in wb.sheetnames:
        del wb[TARGET_SHEET_NAME]
    dst = wb.create_sheet(TARGET_SHEET_NAME)

    src = wb[SOURCE_SHEET_NAME]
    hmap = find_header_map(src)
    src_headers_all = []
    for c in range(1, src.max_column + 1):
        v = normalize_header(src.cell(row=1, column=c).value)
        if v:
            src_headers_all.append(v)

    EXCLUDE = {"latitude", "longitude", "capacity/pe"}
    src_headers = [h for h in src_headers_all if h.lower() not in EXCLUDE]

    new_cols = [
        "Distance to possible flood area [m]",
        "Distance to residential area [m]",
        "Distance to protected area [m]",
        "PA_Type",
    ]

    for j, h in enumerate(src_headers + new_cols, start=1):
        dst.cell(row=1, column=j, value=h)

    lat_col = hmap.get("latitude")
    lon_col = hmap.get("longitude")
    out_row = 2
    right_align = Alignment(horizontal="right")

    def _fmt(d):
        try:
            if isinstance(d, (int, float)) and not math.isnan(d) and d <= 5000:
                return d
        except Exception:
            pass
        return "> 5000"

    knockout_set = set(knockout_rows)  # Für schnellere Lookup
    for r in range(2, src.max_row + 1):
        # Überspringe Zeilen, die im Knockout sind
        if r in knockout_set:
            continue

        for j, h in enumerate(src_headers, start=1):
            src_col = hmap.get(h.lower())
            val = src.cell(row=r, column=src_col).value if src_col else None
            dst.cell(row=out_row, column=j, value=val)

        lat = to_float_or_none(src.cell(row=r, column=lat_col).value)
        lon = to_float_or_none(src.cell(row=r, column=lon_col).value)
        key = f"{float(lat):.6f}|{float(lon):.6f}" if (lat is not None and lon is not None) else None

        dist_flood = dist_res = dist_pa = math.nan
        pa_type = ""
        if key and key in result_map:
            v = result_map[key]
            try:
                dist_flood = float(v[0])
            except Exception:
                pass
            try:
                dist_res = float(v[1])
            except Exception:
                pass
            try:
                dist_pa = float(v[2])
            except Exception:
                pass
            pa_type = str(v[3] or "")

        col_off = len(src_headers)
        c1 = dst.cell(row=out_row, column=col_off + 1, value=_fmt(dist_flood));
        c1.alignment = right_align
        c2 = dst.cell(row=out_row, column=col_off + 2, value=_fmt(dist_res));
        c2.alignment = right_align
        c3 = dst.cell(row=out_row, column=col_off + 3, value=_fmt(dist_pa));
        c3.alignment = right_align
        dst.cell(row=out_row, column=col_off + 4, value=pa_type)

        out_row += 1

    # 3) Knockout-Zeilen aus ALLEN Sheets löschen
    if knockout_rows:
        # Sortiere absteigend, damit  von unten nach oben löschen
        sorted_knockout = sorted(knockout_rows, reverse=True)
        
        for sheet_name in wb.sheetnames:
            # Überspringe das neu erstellte Risks-Sheet
            if sheet_name.strip().lower() == TARGET_SHEET_NAME.strip().lower():
                continue
            
            sheet = wb[sheet_name]
            for row_idx in sorted_knockout:
                # Prüfe, ob die Zeile existiert (manche Sheets könnten weniger Zeilen haben)
                if row_idx <= sheet.max_row:
                    sheet.delete_rows(row_idx, 1)

    # 4) Speichern (Excel)
    try:
        wb.save(xlsx_path)
    except Exception as e:
        raise RuntimeError("Could not save Excel file. Is it open?\n" f"Error: {e}")

if __name__ == "__main__":
    main()
