
from __future__ import annotations

# ---------- stdlib ----------
import argparse, math, sys, time, subprocess, re
from typing import Dict, Optional, Tuple, List, Set
from pathlib import Path

# Automatische Abhängigkeiten installieren falls sie fehlen
def ensure_deps():
    """Installiert fehlende Python-Pakete automatisch"""
    required = ["pandas","geopandas","shapely","pyproj","requests","openpyxl","fiona","rtree"]
    import importlib, pkgutil
    missing = [p for p in required if pkgutil.find_loader(p) is None]
    if not missing:
        return
    print("[setup] Installiere fehlende Pakete:", ", ".join(missing))
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", *missing], check=True)
    except Exception as e:
        print("[setup] pip install fehlgeschlagen:", e)

ensure_deps()

# ---------- 3rd-party ----------
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, Polygon
from shapely.ops import unary_union
from shapely.ops import transform as shp_transform
from shapely.prepared import prep
from pyproj import CRS, Transformer
import requests
from openpyxl import load_workbook

# Pfade und Debug-Einstellungen
BASE = Path(__file__).resolve().parent
CACHE_FILE = BASE / "Output" / "Risks" / "risks_cache.csv"
OUT_CSV    = BASE / "Output" / "Risks" / "near_pa_secondary_check.csv"
GPKG_DIR   = BASE / "Output" / "WWTP Geopackages"
GPKG_PATH  = GPKG_DIR / "WWTP_after_secondary_check.gpkg"
AVAILABLE_AREA_GPKG = BASE / "Output" / "Geopackages" / "Available_Area.gpkg"

EXCEL_CANDIDATES = [
    BASE / "Output" / "WWTP_TP_Database.xlsx",
    BASE / "Output" / "UWWTD_TP_Database.xlsx",
    BASE / "WWTP_TP_Database.xlsx",
    BASE / "UWWTD_TP_Database.xlsx",
]
DEFAULT_FREE_GPKG = BASE / "Daten" / "WWTP_Free_Area.gpkg"

DEBUG_EXPORT_KO = True
DEBUG_DIR = BASE / "Output" / "Debug" / "PA_Check"

# ---------- EEA ArcGIS services ----------
PA_SRC_NATDA = {
    "tag": "NATDA",
    "service": "https://bio.discomap.eea.europa.eu/arcgis/rest/services/ProtectedSites/NatDAv22_Dyna_WM/MapServer",
    "layers": None,
    "layer_names": None,
}
PA_SRC_N2K = {
    "tag": "N2K",
    "service": "http://bio.discomap.eea.europa.eu/arcgis/rest/services/ProtectedSites/Natura2000_Dyna_WM/MapServer",
    "layers": None,
    "layer_names": None,
}
REQUEST_SLEEP = 0.05

# ---------- STRICT PA FILTER ----------
ROMAN_IUCN = {"I","IA","IB","II","III"}
IUCN_KEYS_LOWER = {
    "iucncategory","iucn_cat","iucn_category","iucn","iucncat",
    "iucn_catg","iucn_cat_n","iucn_cat_txt","cdda_iucncat","cdda_iucn_cat"
}

def _is_strict(attrs: dict, src_tag: str) -> bool:
    """N2K: always strict; NatDA/CDDA: only IUCN I–III (roman) or exactly numeric 1/2/3."""
    tag = (src_tag or "").upper()
    if tag == "N2K":
        return True
    if not attrs:
        return False
    for k, v in (attrs or {}).items():
        if k is None or v is None:
            continue
        if str(k).strip().lower() in IUCN_KEYS_LOWER:
            raw = str(v).strip()
            if raw.upper() in ROMAN_IUCN:
                return True
            if re.fullmatch(r"\s*[123]\s*", raw):  # strictly 1 or 2 or 3
                return True
    return False

# ---------- ArcGIS helpers ----------
def arcgis_service_info(service_url: str, session: requests.Session):
    r = session.get(f"{service_url}?f=json", timeout=40); r.raise_for_status(); return r.json()

def arcgis_polygon_layers(service_url: str, session: requests.Session):
    js = arcgis_service_info(service_url, session)
    layers = js.get("layers", []) or []
    poly_ids, names = [], {}
    for lyr in layers:
        if isinstance(lyr, dict) and lyr.get("geometryType") == "esriGeometryPolygon":
            lid = int(lyr["id"]); poly_ids.append(lid); names[lid] = str(lyr.get("name", ""))
    return poly_ids, names

def arcgis_polygon_layers_cached(src_def: dict, session: requests.Session):
    if src_def.get("layers") is None or src_def.get("layer_names") is None:
        lids, names = arcgis_polygon_layers(src_def["service"], session)
        src_def["layers"] = lids; src_def["layer_names"] = names
    return src_def["layers"] or []

def arcgis_query_point_buffer(query_url: str, lat: float, lon: float, radius_m: int,
                              session: requests.Session, out_sr: int = 4326, out_fields: str = "*"):
    params = dict(
        where="1=1",
        geometry=f"{lon},{lat}",
        geometryType="esriGeometryPoint",
        inSR=4326,
        spatialRel="esriSpatialRelIntersects",
        distance=radius_m,
        units="esriSRUnit_Meter",
        outFields=out_fields,
        returnGeometry="true",
        outSR=out_sr,
        f="json",
    )
    r = session.get(query_url, params=params, timeout=40); r.raise_for_status()
    js = r.json();
    if "error" in js: return []
    return js.get("features", []) or []

# ---------- ESRI rings -> Shapely (multipart-fest) ----------
def _ring_signed_area(coords):
    a = 0.0
    for i in range(len(coords)-1):
        x1, y1 = coords[i]; x2, y2 = coords[i+1]; a += (x1*y2 - x2*y1)
    return a*0.5

def esri_polygon_to_shapely(esri_geom: dict):
    if not esri_geom: return None
    rings = esri_geom.get("rings")
    if not rings: return None
    outers: List[Tuple[Polygon, List]] = []; hole_polys: List[Polygon] = []
    for coords in rings:
        if not coords or len(coords) < 4: continue
        try: poly = Polygon(coords)
        except Exception: continue
        if not poly.is_valid: poly = poly.buffer(0)
        if poly.is_empty: continue
        sa = _ring_signed_area(coords)
        if sa < 0: outers.append((poly, []))   # CW outer
        else:     hole_polys.append(poly)      # CCW hole
    if not outers and hole_polys:
        u = unary_union(hole_polys).buffer(0); return None if u.is_empty else u
    for h in hole_polys:
        rp = h.representative_point()
        for i,(op,holes) in enumerate(outers):
            try:
                if op.contains(rp):
                    holes.append(list(h.exterior.coords)); break
            except Exception: pass
    polys = []
    for op, holes in outers:
        try:
            p = Polygon(list(op.exterior.coords), holes=holes if holes else None)
            if not p.is_valid: p = p.buffer(0)
            if not p.is_empty: polys.append(p)
        except Exception:
            if not op.is_empty: polys.append(op)
    if not polys: return None
    return polys[0] if len(polys)==1 else unary_union(polys)

def fetch_protected_areas_union_and_attr(lat: float, lon: float, radius_m: int,
                                         src_def: dict, session: requests.Session):
    """Strict-filtered *service-internal* union (WGS84). Returns (union_geom, kept_attrs_list)."""
    lids = arcgis_polygon_layers_cached(src_def, session)
    if not lids: return None, []
    all_geoms = []; kept_attrs = []
    tag = src_def.get("tag","")
    for lyr_id in lids:
        qurl = f"{src_def['service'].rstrip('/')}/{lyr_id}/query"
        feats = arcgis_query_point_buffer(qurl, lat, lon, radius_m, session, out_sr=4326, out_fields="*")
        for f in feats:
            attrs = (f.get("attributes") or {})
            if not _is_strict(attrs, tag): continue
            poly = esri_polygon_to_shapely(f.get("geometry"))
            if poly is None: continue
            if not poly.is_valid: poly = poly.buffer(0)
            if poly.is_empty: continue
            all_geoms.append(poly); kept_attrs.append(attrs)
        time.sleep(REQUEST_SLEEP)
    if not all_geoms: return None, []
    return unary_union(all_geoms), kept_attrs

# ---------- CRS / reprojection ----------
def utm_crs_for_lonlat(lon: float, lat: float) -> CRS:
    zone = int(math.floor((lon + 180) / 6) + 1)
    epsg = 32600 + zone if lat >= 0 else 32700 + zone
    return CRS.from_epsg(epsg)

def transformer(src_epsg: int | CRS, dst_crs: CRS) -> Transformer:
    src = CRS.from_epsg(src_epsg) if isinstance(src_epsg, int) else src_epsg
    return Transformer.from_crs(src, dst_crs, always_xy=True)

def reproject_geom(geom, src: CRS | int, dst: CRS):
    if geom is None: return None
    tr = transformer(src, dst)
    return shp_transform(lambda x,y,z=None: tr.transform(x,y), geom)

# ---------- data/io helpers ----------
def detect_cache_file(explicit: Optional[Path]) -> Path:
    if explicit: return explicit
    if CACHE_FILE.exists(): return CACHE_FILE
    if (BASE / "risks_cache.csv").exists(): return BASE / "risks_cache.csv"
    if Path("risks_cache.csv").exists(): return Path("risks_cache.csv")
    raise FileNotFoundError("risks_cache.csv not found. Expected under Output/Risks/.")

def normalize_cache_columns(df: pd.DataFrame) -> Dict[str, str]:
    cmap = {str(c).strip(): str(c).strip() for c in df.columns}
    lower = {c.lower(): c for c in cmap}
    need = {
        "lat": ["Latitude","latitude","LAT"],
        "lon": ["Longitude","longitude","LON"],
        "name":["Name","name","WWTP name","Plant name","facility"],
        "pa_dist":["Distance to protected area [m]","distance to protected area [m]","PA_Dist_m"],
        "code":["UWWTD code","uwwtd code","UWWTD_Code","Code","ID"],
    }
    out: Dict[str,str] = {}
    for key, options in need.items():
        for opt in options:
            if opt in cmap: out[key] = cmap[opt]; break
            if opt.lower() in lower: out[key] = lower[opt.lower()]; break
    missing = [k for k in ("lat","lon","pa_dist") if k not in out]
    if missing:
        raise KeyError(f"Missing required columns in cache: {', '.join(missing)}. Available: {list(df.columns)}")
    return out

def load_free_area_gdf(path: Path, layer: Optional[str]) -> Tuple[gpd.GeoDataFrame, str]:
    if not path.exists():
        raise FileNotFoundError(f"Free-area GPKG not found: {path}")
    lyr = layer
    if lyr is None:
        import fiona
        polys = []
        for L in fiona.listlayers(str(path)):
            try:
                g = gpd.read_file(path, layer=L)
                if not g.empty and g.geometry.type.isin(["Polygon","MultiPolygon"]).any():
                    polys.append(L)
            except Exception: pass
        if not polys:
            raise ValueError("No polygon layer in free-area GPKG.")
        lyr = polys[0]
    gdf = gpd.read_file(path, layer=lyr)
    if gdf.empty:
        raise ValueError(f"Layer '{lyr}' is empty.")
    gdf = gdf[~gdf.geometry.is_empty & gdf.geometry.notnull()].copy()
    gdf.reset_index(drop=True, inplace=True)
    return gdf, lyr

def detect_excel_path() -> Optional[str]:
    for p in EXCEL_CANDIDATES:
        if Path(p).exists(): return str(p)
    for p in ["Output/WWTP_TP_Database.xlsx","Output/UWWTD_TP_Database.xlsx",
              "WWTP_TP_Database.xlsx","UWWTD_TP_Database.xlsx"]:
        if Path(p).exists(): return p
    return None

# ---------- progress bar ----------
def _format_eta(start_ts: float, done: int, total: int) -> str:
    if done<=0: return "ETA --:--"
    elapsed = max(0.0, time.time()-start_ts)
    rate = done/elapsed if elapsed>0 else 0
    remain = (total-done)/rate if rate>0 else 0
    mm, ss = int(remain//60), int(remain%60)
    return f"ETA {mm:02d}:{ss:02d}"

def _progress_line(done: int, total: int, ko: int, start_ts: float, label: str = "", width: int = 28) -> str:
    ratio = 0.0 if total==0 else done/total
    filled = int(ratio*width)
    bar = "#"*filled + "-"*(width-filled)
    return f"[{bar}] {done}/{total} {ratio*100:5.1f}% | KO {ko} | {_format_eta(start_ts,done,total)} | {(label or '')[:40]}"

# ---------- robust clipped metrics WITHOUT cross-service union ----------
def free_area_metrics_clipped_no_union(lat: float, lon: float, radius_m: int,
                                       free_gdf: gpd.GeoDataFrame,
                                       session: requests.Session) -> Tuple[float, float, int,
                                                                           object, object, object]:
    """
    Returns:
      total_outside_m2, max_single_outside_m2, n_polys,
      pa_natda_utm, pa_n2k_utm, buf_utm  (for debug)
    """
    # UTM & Buffer
    utm = utm_crs_for_lonlat(lon, lat)
    pt_ll = Point(lon, lat)
    pt_utm = reproject_geom(pt_ll, 4326, utm)
    buf = pt_utm.buffer(radius_m).buffer(0)

    # Free areas to UTM
    if free_gdf.crs is None:
        raise ValueError("Free-area GPKG lacks CRS.")
    free_local = free_gdf if str(free_gdf.crs).lower()==str(utm).lower() else free_gdf.to_crs(utm)

    # Candidates via sindex
    sidx = free_local.sindex
    cand_idx = list(sidx.intersection(buf.bounds))
    if not cand_idx:
        return 0.0, 0.0, 0, None, None, buf
    subset = free_local.iloc[cand_idx].copy()

    # strict per-service unions (but no cross-service union)
    pa_nat_ll, _  = fetch_protected_areas_union_and_attr(lat, lon, radius_m, PA_SRC_NATDA, session)
    pa_n2k_ll, _  = fetch_protected_areas_union_and_attr(lat, lon, radius_m, PA_SRC_N2K, session)
    pa_natda_utm  = reproject_geom(pa_nat_ll, 4326, utm) if pa_nat_ll else None
    pa_n2k_utm    = reproject_geom(pa_n2k_ll, 4326, utm) if pa_n2k_ll else None
    if pa_natda_utm is not None and not pa_natda_utm.is_empty:
        try: pa_natda_utm = pa_natda_utm.buffer(0)
        except Exception: pass
        pre_nat = prep(pa_natda_utm)
    else:
        pre_nat = None
    if pa_n2k_utm is not None and not pa_n2k_utm.is_empty:
        try: pa_n2k_utm = pa_n2k_utm.buffer(0)
        except Exception: pass
        pre_n2k = prep(pa_n2k_utm)
    else:
        pre_n2k = None

    total = 0.0; max_single = 0.0; n = 0
    for g in subset.geometry:
        if g is None or g.is_empty: continue
        try: g = g.buffer(0)
        except Exception: pass
        if not g.intersects(buf): continue
        clip = g.intersection(buf)
        if clip.is_empty: continue
        try: clip = clip.buffer(0)
        except Exception: pass

        # 1) primär: sequentielle Differenzen (ohne Cross-Union)
        try:
            rem = clip
            if pre_nat is not None and pre_nat.intersects(rem):
                rem = rem.difference(pa_natda_utm)
                try: rem = rem.buffer(0)
                except Exception: pass
            if pre_n2k is not None and pre_n2k.intersects(rem):
                rem = rem.difference(pa_n2k_utm)
                try: rem = rem.buffer(0)
                except Exception: pass
            outside_area = float(rem.area) if not rem.is_empty else 0.0
        except Exception:
            # 2) robustes Fallback: Inclusion–Exclusion (ohne Union)
            clip_area = float(clip.area)
            inter_nat = 0.0
            inter_n2k = 0.0
            inter_both = 0.0
            if pre_nat is not None and pre_nat.intersects(clip):
                tmp = clip.intersection(pa_natda_utm)
                if not tmp.is_empty:
                    try: tmp = tmp.buffer(0)
                    except Exception: pass
                    inter_nat = float(tmp.area)
            if pre_n2k is not None and pre_n2k.intersects(clip):
                tmp2 = clip.intersection(pa_n2k_utm)
                if not tmp2.is_empty:
                    try: tmp2 = tmp2.buffer(0)
                    except Exception: pass
                    inter_n2k = float(tmp2.area)
                if inter_nat > 0.0:
                    both = tmp2.intersection(clip.intersection(pa_natda_utm)) if 'tmp' in locals() else tmp2.intersection(pa_natda_utm).intersection(clip)
                    if not both.is_empty:
                        try: both = both.buffer(0)
                        except Exception: pass
                        inter_both = float(both.area)
            outside_area = max(0.0, clip_area - inter_nat - inter_n2k + inter_both)

        total += outside_area
        if outside_area > max_single:
            max_single = outside_area
        n += 1

    return total, max_single, n, pa_natda_utm, pa_n2k_utm, buf

# ---------- debug export ----------
def _safe_name(s: Optional[str]) -> str:
    if not s: return "unknown"
    return re.sub(r"[^A-Za-z0-9_-]+", "_", str(s))[:60]

def export_debug_case_no_union(lat: float, lon: float, name: Optional[str], code: Optional[str],
                               radius_m: int, free_gdf: gpd.GeoDataFrame, session: requests.Session):
    try:
        tot, mx, n, pa_nat, pa_n2k, buf = free_area_metrics_clipped_no_union(lat, lon, radius_m, free_gdf, session)
        utm = utm_crs_for_lonlat(lon, lat)
        # Für Visualisierung: free_clip + outside erneut bestimmen (wie in Funktion)
        free_local = free_gdf if str(free_gdf.crs).lower()==str(utm).lower() else free_gdf.to_crs(utm)
        sidx = free_local.sindex
        cand_idx = list(sidx.intersection(buf.bounds))
        subset = free_local.iloc[cand_idx].copy() if cand_idx else free_local.iloc[[]].copy()
        subset["free_clip"] = subset.geometry.apply(lambda g: g.intersection(buf) if g.intersects(buf) else None)
        subset = subset[subset["free_clip"].notnull() & (~subset["free_clip"].is_empty)]
        pre_nat = prep(pa_nat) if pa_nat is not None and not pa_nat.is_empty else None
        pre_n2k = prep(pa_n2k) if pa_n2k is not None and not pa_n2k.is_empty else None
        def outside_geom(c):
            rem = c
            if pre_nat is not None and pre_nat.intersects(rem):
                rem = rem.difference(pa_nat)
            if pre_n2k is not None and pre_n2k.intersects(rem):
                rem = rem.difference(pa_n2k)
            return rem
        subset["outside"] = subset["free_clip"].apply(outside_geom)

        DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        fname = f"{_safe_name(code)}_{_safe_name(name)}_{lat:.5f}_{lon:.5f}.gpkg"
        pth = DEBUG_DIR / fname
        if pth.exists():
            try: pth.unlink()
            except Exception: pass

        gpd.GeoDataFrame([{"geometry": buf}], crs=utm).to_file(pth, layer="buffer", driver="GPKG")
        if pa_nat is not None and not pa_nat.is_empty:
            gpd.GeoDataFrame([{"geometry": pa_nat}], crs=utm).to_file(pth, layer="pa_natda", driver="GPKG")
        if pa_n2k is not None and not pa_n2k.is_empty:
            gpd.GeoDataFrame([{"geometry": pa_n2k}], crs=utm).to_file(pth, layer="pa_n2k", driver="GPKG")
        gpd.GeoDataFrame(geometry=subset["free_clip"], crs=utm).to_file(pth, layer="free_clip", driver="GPKG")
        outs = subset["outside"][~subset["outside"].is_empty]
        if not outs.empty:
            gpd.GeoDataFrame(geometry=outs, crs=utm).to_file(pth, layer="outside", driver="GPKG")
        print(f"[debug] Exported KO debug GPKG -> {pth}")
    except Exception as e:
        print(f"[debug] Export failed: {e}")

# ---------- GPKG helpers ----------
def read_excel_remaining_as_gdf(xlsx_path: str) -> gpd.GeoDataFrame:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    frames = []
    for ws in wb.worksheets:
        header = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            if v is None: continue
            header[str(v).strip().lower()] = col
        col_lat = header.get('latitude'); col_lon = header.get('longitude')
        if not col_lat or not col_lon: continue
        col_name = header.get('name') or header.get('plant name') or header.get('wwtp name')
        col_code = header.get('uwwtd code') or header.get('uwwtd_code') or header.get('code') or header.get('id')
        rows = []
        for r in range(2, ws.max_row + 1):
            try:
                lat_val = ws.cell(row=r, column=col_lat).value
                lon_val = ws.cell(row=r, column=col_lon).value
                if lat_val is None or lon_val is None: continue
                lat = float(str(lat_val).replace(',', '.'))
                lon = float(str(lon_val).replace(',', '.'))
                name = ws.cell(row=r, column=col_name).value if col_name else None
                code = ws.cell(row=r, column=col_code).value if col_code else None
                rows.append({'Name': str(name) if name is not None else None,
                             'Code': str(code) if code is not None else None,
                             'Latitude': lat, 'Longitude': lon, 'Sheet': ws.title})
            except Exception: continue
        if rows: frames.append(pd.DataFrame(rows))
    if not frames:
        return gpd.GeoDataFrame(columns=['Name','Code','Latitude','Longitude','Sheet','geometry'], geometry=[], crs='EPSG:4326')
    df_all = pd.concat(frames, ignore_index=True)
    df_all.drop_duplicates(subset=['Code','Latitude','Longitude','Name'], inplace=True)
    df_all['geometry'] = df_all.apply(lambda r: Point(float(r['Longitude']), float(r['Latitude'])), axis=1)
    return gpd.GeoDataFrame(df_all, geometry='geometry', crs='EPSG:4326')

def remaining_from_cache_after_deletions(cache_df: pd.DataFrame, out_df: pd.DataFrame,
                                         lat_col: str, lon_col: str, name_col: str, code_col: str) -> gpd.GeoDataFrame:
    df_all = cache_df.copy()
    if 'RowsDeletedInExcel' not in out_df.columns:
        out_df['RowsDeletedInExcel'] = 0
    out_df['RowsDeletedInExcel'] = pd.to_numeric(out_df['RowsDeletedInExcel'], errors='coerce').fillna(0).astype(int)
    dels = out_df[(out_df['SecondCheck_Status']=='KO') & (out_df['RowsDeletedInExcel']>0)].copy()
    del_codes: Set[str] = set(); del_triplets: Set[tuple] = set()
    if not dels.empty:
        if code_col in dels.columns: del_codes = set(dels[code_col].dropna().astype(str))
        for _, r in dels.iterrows():
            try:
                nm = str(r[name_col]).strip().lower() if name_col in dels.columns and pd.notna(r[name_col]) else None
                lt = float(r[lat_col]); ln = float(r[lon_col])
                if nm is not None: del_triplets.add((nm, lt, ln))
            except Exception: continue
    if code_col in df_all.columns:
        df_all = df_all[~df_all[code_col].astype(str).isin(del_codes)]
    if name_col in df_all.columns:
        def not_deleted_row(r):
            try: return (str(r[name_col]).strip().lower(), float(r[lat_col]), float(r[lon_col])) not in del_triplets
            except Exception: return True
        df_all = df_all[df_all.apply(not_deleted_row, axis=1)]
    cols = [c for c in [name_col, code_col, lat_col, lon_col] if c in df_all.columns]
    df_all = df_all[cols].dropna()
    df_all['geometry'] = df_all.apply(lambda r: Point(float(r[lon_col]), float(r[lat_col])), axis=1)
    return gpd.GeoDataFrame(df_all, geometry='geometry', crs='EPSG:4326')

def write_gpkg(deleted_gdf: gpd.GeoDataFrame, remaining_gdf: gpd.GeoDataFrame, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        try: path.unlink()
        except Exception: pass
    wrote_any = False
    if deleted_gdf is not None and not deleted_gdf.empty:
        deleted_gdf.to_file(path, layer='deleted', driver='GPKG'); wrote_any = True
    if remaining_gdf is not None and not remaining_gdf.empty:
        remaining_gdf.to_file(path, layer='remaining', driver='GPKG'); wrote_any = True
    if not wrote_any:
        print("[info] No features for GPKG export (both layers empty). Skipping GPKG.")

def collect_outside_polygons(lat: float, lon: float, name: Optional[str], code: Optional[str],
                             radius_m: int, free_gdf: gpd.GeoDataFrame, session: requests.Session) -> List[dict]:
    """Sammelt alle Outside-Polygone für eine Anlage mit Metadaten."""
    try:
        utm = utm_crs_for_lonlat(lon, lat)
        pt_ll = Point(lon, lat)
        pt_utm = reproject_geom(pt_ll, 4326, utm)
        buf = pt_utm.buffer(radius_m).buffer(0)
        
        free_local = free_gdf if str(free_gdf.crs).lower()==str(utm).lower() else free_gdf.to_crs(utm)
        sidx = free_local.sindex
        cand_idx = list(sidx.intersection(buf.bounds))
        if not cand_idx:
            return []
        subset = free_local.iloc[cand_idx].copy()
        
        pa_nat_ll, _ = fetch_protected_areas_union_and_attr(lat, lon, radius_m, PA_SRC_NATDA, session)
        pa_n2k_ll, _ = fetch_protected_areas_union_and_attr(lat, lon, radius_m, PA_SRC_N2K, session)
        pa_natda_utm = reproject_geom(pa_nat_ll, 4326, utm) if pa_nat_ll else None
        pa_n2k_utm = reproject_geom(pa_n2k_ll, 4326, utm) if pa_n2k_ll else None
        
        if pa_natda_utm is not None and not pa_natda_utm.is_empty:
            try: pa_natda_utm = pa_natda_utm.buffer(0)
            except Exception: pass
            pre_nat = prep(pa_natda_utm)
        else:
            pre_nat = None
        if pa_n2k_utm is not None and not pa_n2k_utm.is_empty:
            try: pa_n2k_utm = pa_n2k_utm.buffer(0)
            except Exception: pass
            pre_n2k = prep(pa_n2k_utm)
        else:
            pre_n2k = None
        
        results = []
        for idx, g in enumerate(subset.geometry):
            if g is None or g.is_empty: continue
            try: g = g.buffer(0)
            except Exception: pass
            if not g.intersects(buf): continue
            clip = g.intersection(buf)
            if clip.is_empty: continue
            try: clip = clip.buffer(0)
            except Exception: pass
            
            rem = clip
            if pre_nat is not None and pre_nat.intersects(rem):
                rem = rem.difference(pa_natda_utm)
                try: rem = rem.buffer(0)
                except Exception: pass
            if pre_n2k is not None and pre_n2k.intersects(rem):
                rem = rem.difference(pa_n2k_utm)
                try: rem = rem.buffer(0)
                except Exception: pass
            
            if not rem.is_empty:
                area_m2 = float(rem.area)
                # Zurück zu WGS84 für GPKG
                rem_wgs84 = reproject_geom(rem, utm, 4326)
                results.append({
                    'Name': name,
                    'Code': code,
                    'Latitude': lat,
                    'Longitude': lon,
                    'Area_m2': round(area_m2, 2),
                    'geometry': rem_wgs84
                })
        return results
    except Exception as e:
        print(f"[warning] Fehler beim Sammeln der Outside-Polygone für {name}: {e}")
        return []

# ---------- Excel deletion helper ----------
def delete_rows_in_excel_for_facility(xlsx_path: str, name: Optional[str], code: Optional[str],
                                      lat: float, lon: float, lat_lon_tol_deg: float = 1e-4) -> int:
    wb = load_workbook(xlsx_path); deleted = 0
    for ws in wb.worksheets:
        header = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            if v is None: continue
            header[str(v).strip().lower()] = col
        col_code = None
        for key in ["uwwtd code","uwwtd_code","code","id"]:
            if key in header: col_code = header[key]; break
        col_name = header.get("name") or header.get("plant name") or header.get("wwtp name")
        col_lat = header.get("latitude"); col_lon = header.get("longitude")
        matches: List[int] = []
        for r in range(2, ws.max_row + 1):
            try:
                ok = False
                if code and col_code:
                    cell = ws.cell(row=r, column=col_code).value
                    if cell is not None and str(cell).strip() == str(code).strip(): ok = True
                else:
                    nm_ok = False
                    if name and col_name:
                        cell = ws.cell(row=r, column=col_name).value
                        if cell is not None and str(cell).strip().lower() == str(name).strip().lower(): nm_ok = True
                    lat_ok = lon_ok = False
                    if col_lat:
                        try:
                            lat_val = float(str(ws.cell(row=r, column=col_lat).value).replace(',', '.'))
                            lat_ok = abs(lat_val - lat) <= lat_lon_tol_deg
                        except Exception: pass
                    if col_lon:
                        try:
                            lon_val = float(str(ws.cell(row=r, column=col_lon).value).replace(',', '.'))
                            lon_ok = abs(lon_val - lon) <= lat_lon_tol_deg
                        except Exception: pass
                    ok = nm_ok and lat_ok and lon_ok
                if ok: matches.append(r)
            except Exception: continue
        for r in sorted(set(matches), reverse=True):
            ws.delete_rows(idx=r, amount=1); deleted += 1
    wb.save(xlsx_path); return deleted

# ---------- MAIN ----------
def main() -> None:
    _ = argparse.ArgumentParser(add_help=False).parse_args([])
    print("NO cross-service union. Sequence: subtract NatDA (I–III) then N2K (strict).")
    print("OK if ANY clipped free-area polygon has >= 6000 m² outside BOTH services.")

    cache_path = detect_cache_file(None)
    df = pd.read_csv(cache_path)
    colmap = normalize_cache_columns(df)
    lat_c = colmap["lat"]; lon_c = colmap["lon"]; dist_c = colmap["pa_dist"]
    name_c = colmap.get("name") or "Name"; code_c = colmap.get("code") or "Code"

    def _f(x):
        try: return float(str(x).replace(',', '.'))
        except Exception: return math.nan
    df[dist_c] = df[dist_c].map(_f); df[lat_c] = df[lat_c].map(_f); df[lon_c] = df[lon_c].map(_f)

    threshold_m = 500.0
    radius_m = 2500
    area_threshold = 6000.0
    EPS = 1.0

    cand = df[(df[dist_c].notna()) & (df[dist_c] >= 0) & (df[dist_c] <= threshold_m)].copy()
    total = len(cand)
    if total == 0:
        print(f"No candidates within 0–{int(threshold_m)} m."); return

    free_gdf, _ = load_free_area_gdf(DEFAULT_FREE_GPKG, None)
    session = requests.Session()

    rows = []; ko_count = 0; err_count = 0; deleted_records = []
    all_outside_polygons = []  # Sammelt alle Outside-Polygone
    start_ts = time.time(); processed = 0
    print(_progress_line(processed, total, ko_count, start_ts, label='starting...'), end='\r', flush=True)

    for _, row in cand.iterrows():
        lat = float(row[lat_c]); lon = float(row[lon_c])
        name = str(row[name_c]) if name_c in cand.columns else None
        code = str(row[code_c]) if code_c in cand.columns and not pd.isna(row[code_c]) else None

        status = "ERROR"; notes = None
        total_out_m2 = math.nan; max_single_m2 = math.nan; n_polys = 0
        try:
            total_out_m2, max_single_m2, n_polys, pa_nat, pa_n2k, buf = free_area_metrics_clipped_no_union(
                lat, lon, radius_m, free_gdf, session
            )
            status = "OK" if (max_single_m2 + EPS) >= area_threshold else "KO"
            
            # Sammle Outside-Polygone für alle Anlagen
            outside_polys = collect_outside_polygons(lat, lon, name, code, radius_m, free_gdf, session)
            all_outside_polygons.extend(outside_polys)
            
        except Exception as e:
            notes = f"Fehler bei Prüfung: {e}"; err_count += 1

        # Debug export for KO
        if DEBUG_EXPORT_KO and status == "KO":
            export_debug_case_no_union(lat, lon, name, code, radius_m, free_gdf, session)

        del_rows = 0
        if status == "KO":
            xlsx_path = detect_excel_path()
            if not xlsx_path:
                notes = (notes + " | " if notes else "") + "Excel-DB nicht gefunden; keine Löschung."
            else:
                try:
                    del_rows = delete_rows_in_excel_for_facility(xlsx_path, name, code, lat, lon)
                except Exception as e:
                    notes = (notes + " | " if notes else "") + f"Löschung fehlgeschlagen: {e}"
        if status == "KO": ko_count += 1

        if del_rows > 0:
            deleted_records.append({
                name_c: name, code_c: code, lat_c: lat, lon_c: lon,
                'MaxSinglePolyOutsidePA_m2': max_single_m2,
                'TotalOutsidePA_m2': total_out_m2,
                'RowsDeletedInExcel': del_rows
            })

        rows.append({
            name_c: name, code_c: code, lat_c: lat, lon_c: lon,
            "SecondCheck_Status": status,
            "SecondCheck_MaxSinglePolyOutsidePA_m2": round(max_single_m2,2) if isinstance(max_single_m2,(int,float)) else max_single_m2,
            "SecondCheck_TotalOutsidePA_m2": round(total_out_m2,2) if isinstance(total_out_m2,(int,float)) else total_out_m2,
            "SecondCheck_FreePolygonsUsed": n_polys,
            "SecondCheck_Radius_m": radius_m,
            "SecondCheck_Notes": notes,
            "RowsDeletedInExcel": del_rows,
        })

        processed += 1
        print(_progress_line(processed, total, ko_count, start_ts, label=name or ''), end='\r', flush=True)

    print(_progress_line(total, total, ko_count, start_ts, label='done'), end='\n', flush=True)

    out_df = pd.DataFrame(rows); OUT_CSV.parent.mkdir(parents=True, exist_ok=True); out_df.to_csv(OUT_CSV, index=False)

    # GPKG layers
    if deleted_records:
        del_df = pd.DataFrame(deleted_records)
        del_df['geometry'] = del_df.apply(lambda r: Point(float(r[lon_c]), float(r[lat_c])), axis=1)
        deleted_gdf = gpd.GeoDataFrame(del_df, geometry='geometry', crs='EPSG:4326')
    else:
        deleted_gdf = gpd.GeoDataFrame(columns=[name_c, code_c, lat_c, lon_c,
                                                'MaxSinglePolyOutsidePA_m2','TotalOutsidePA_m2',
                                                'RowsDeletedInExcel','geometry'], geometry=[], crs='EPSG:4326')
    xlsx_used = detect_excel_path()
    if xlsx_used:
        remaining_gdf = read_excel_remaining_as_gdf(xlsx_used)
    else:
        remaining_gdf = remaining_from_cache_after_deletions(df, out_df, lat_c, lon_c, name_c, code_c)
    write_gpkg(deleted_gdf, remaining_gdf, GPKG_PATH)
    
    # ---- SYNC: Alle Sheets mit General Data synchronisieren ----
    if xlsx_used and ko_count > 0:
        print("\n[sync] Synchronisiere alle Sheets mit General Data...")
        try:
            book = pd.read_excel(xlsx_used, sheet_name=None, engine="openpyxl")
            main_sheet = "General Data"
            key_col = "UWWTD Code"
            
            if main_sheet in book and key_col in book[main_sheet].columns:
                general_codes = set(book[main_sheet][key_col].astype(str).str.strip())
                synced = False
                
                for sheet_name, df in book.items():
                    if sheet_name != main_sheet and key_col in df.columns:
                        before = len(df)
                        book[sheet_name] = df[df[key_col].astype(str).str.strip().isin(general_codes)].copy()
                        after = len(book[sheet_name])
                        if before != after:
                            print(f"[sync] {sheet_name}: {before} → {after} Zeilen")
                            synced = True
                
                if synced:
                    with pd.ExcelWriter(xlsx_used, engine="openpyxl", mode="w") as writer:
                        for sheet_name in book.keys():
                            book[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
                    print("[sync] ✓ Excel aktualisiert")
        except Exception as e:
            print(f"[sync] Warnung: Sync fehlgeschlagen: {e}")

    print(f"Done. Candidates: {total} | KO: {ko_count} | ERROR: {err_count} | Report: {OUT_CSV}\n"
          f"GPKG: {GPKG_PATH} (layers: deleted, remaining)")
    if DEBUG_EXPORT_KO:
        print(f"[debug] KO debug packages: {DEBUG_DIR}")

if __name__ == "__main__":
    main()
