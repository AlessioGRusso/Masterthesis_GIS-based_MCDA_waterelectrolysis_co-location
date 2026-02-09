

import json
import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import numpy as np
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
import threading


@dataclass
class MonteCarloConfig:
    """Konfiguration für die Monte Carlo Analyse - hier kann man alles einstellen"""
    iterations: int = 1000
    confidence_level: float = 0.95
    random_seed: Optional[int] = 42
    run_name: str = ""


def auto_col_width(path: str, sheet_name: str, min_w: float = 8.0, max_w: float = 80.0) -> None:
    """Passt Spaltenbreiten automatisch an - macht die Excel-Dateien schöner lesbar"""
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            s_len = max(len(part) for part in s.splitlines()) if "\n" in s else len(s)
            max_len = max(max_len, s_len)
        width = min(max_w, max(min_w, max_len + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width
    wb.save(path)
    wb.close()


class MonteCarloWindow(tk.Toplevel):
    """Monte Carlo Konfigurationsfenster"""
    
    def __init__(self, parent, project_config_path: str):
        super().__init__(parent)
        self.title("Monte Carlo Analysis - WWTP")
        self.geometry("500x400")
        self.resizable(False, False)
        self.project_config_path = project_config_path
        
        # Fenster zentrieren
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        # Variablen
        self.var_iterations = tk.IntVar(value=10000)
        self.var_confidence = tk.DoubleVar(value=0.95)
        self.var_seed = tk.IntVar(value=42)
        self.var_use_seed = tk.BooleanVar(value=True)
        self.var_run_name = tk.StringVar(value="")
        
        self.setup_ui()
        
    def setup_ui(self):
        """UI erstellen"""
        self.main_frame = ttk.Frame(self, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame = self.main_frame
        
        # Titel
        ttk.Label(main_frame, text="Monte Carlo Analysis Configuration", 
                 font=("Arial", 12, "bold")).pack(pady=(0, 20))
        
        # Parameter
        param_frame = ttk.Frame(main_frame)
        param_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(param_frame, text="Anzahl Iterationen:").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Entry(param_frame, textvariable=self.var_iterations, width=15).grid(row=0, column=1, sticky="w", padx=(10, 0))
        ttk.Label(param_frame, text="(100-50000)").grid(row=0, column=2, sticky="w", padx=(5, 0))
        
        ttk.Label(param_frame, text="Konfidenzintervall:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(param_frame, textvariable=self.var_confidence, width=15).grid(row=1, column=1, sticky="w", padx=(10, 0))
        ttk.Label(param_frame, text="(0.90-0.99)").grid(row=1, column=2, sticky="w", padx=(5, 0))
        
        ttk.Label(param_frame, text="Lauf-Name:").grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(param_frame, textvariable=self.var_run_name, width=15).grid(row=2, column=1, sticky="w", padx=(10, 0))
        ttk.Label(param_frame, text="(optional)").grid(row=2, column=2, sticky="w", padx=(5, 0))
        
        # Seed
        seed_frame = ttk.Frame(param_frame)
        seed_frame.grid(row=3, column=0, columnspan=3, sticky="w", pady=10)
        ttk.Checkbutton(seed_frame, text="Random Seed (Reproduzierbarkeit):", 
                       variable=self.var_use_seed).pack(side=tk.LEFT)
        ttk.Entry(seed_frame, textvariable=self.var_seed, width=10).pack(side=tk.LEFT, padx=(10, 0))
        
        # Infos
        info_text = ("Monte Carlo Analyse sampelt Gewichte gleichmäßig innerhalb\n"
                    "der w_min und w_max Grenzen für jeden KPI und Kategorie.\n"
                    "Ergebnisse: Mittelwert, Standardabweichung und Rang-Statistiken.")
        ttk.Label(main_frame, text=info_text, font=("Arial", 8), 
                 foreground="gray", justify=tk.LEFT).pack(pady=(0, 20))
        
        # Status Label
        self.status_label = ttk.Label(main_frame, text="", font=("Arial", 9), foreground="blue")
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        self.cancel_button = ttk.Button(button_frame, text="Abbrechen", command=self.destroy)
        self.cancel_button.pack(side=tk.RIGHT, padx=(10, 0))
        self.start_button = ttk.Button(button_frame, text="Analyse starten", command=self.start_analysis)
        self.start_button.pack(side=tk.RIGHT)
    
    def update_status(self, message: str):
        """Status Label aktualisieren"""
        def _update():
            try:
                self.status_label.config(text=message)
                self.status_label.pack(pady=(10, 10))
                self.update()
            except Exception as e:
                print(f"Status update error: {e}")
        
        try:
            self.after(0, _update)
        except Exception as e:
            print(f"After scheduling error: {e}")
    
    def run_analysis_thread(self, config):
        """Analyse in separatem Thread ausführen"""
        print("DEBUG: run_analysis_thread started")
        try:
            print("DEBUG: Creating analyzer")
            # Analyse mit Status Callback ausführen
            analyzer = MonteCarloAnalyzer(self.project_config_path, config, status_callback=self.update_status)
            print("DEBUG: Starting analysis")
            analyzer.run_analysis()
            print("DEBUG: Analysis completed")
            
            # Erfolg - UI im Main Thread planen
            def _on_success():
                # Status aktualisieren
                self.status_label.config(text="Analyse erfolgreich abgeschlossen!", foreground="green")
                self.status_label.pack(pady=(10, 5))
                
                # Buttons wieder aktivieren
                self.start_button.config(state='normal')
                self.cancel_button.config(text='Schließen', command=self.destroy)
                
                # Fenster aktualisieren
                self.update_idletasks()
                self.update()
            
            self.after(0, _on_success)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print("=" * 80)
            print("MONTE CARLO ERROR:")
            print(error_details)
            print("=" * 80)
            
            # Fehlermeldung extrahieren
            error_msg = str(e)
            if "FileNotFoundError" in error_details:
                error_msg = f"Datei nicht gefunden:\n{str(e)}"
            elif "KeyError" in error_details:
                error_msg = f"Fehlende Spalte oder Schlüssel:\n{str(e)}"
            elif "ValueError" in error_details:
                error_msg = f"Ungültiger Wert:\n{str(e)}"
            
            # Fehlerbehandlung im Main Thread planen
            def _on_error():
                self.status_label.config(text="Analyse fehlgeschlagen!", foreground="red")
                self.start_button.config(state='normal')
                self.cancel_button.config(state='normal')
                
                # Show detailed error in scrollable text window
                error_window = tk.Toplevel(self)
                error_window.title("Analysis Error")
                error_window.geometry("600x400")
                
                # Center error window
                error_window.update_idletasks()
                w = error_window.winfo_width()
                h = error_window.winfo_height()
                sw = error_window.winfo_screenwidth()
                sh = error_window.winfo_screenheight()
                x = (sw - w) // 2
                y = (sh - h) // 2
                error_window.geometry(f"{w}x{h}+{x}+{y}")
                
                frame = ttk.Frame(error_window, padding="10")
                frame.pack(fill=tk.BOTH, expand=True)
                
                ttk.Label(frame, text="Analysis Failed", font=("Arial", 12, "bold"), foreground="red").pack(pady=(0, 10))
                ttk.Label(frame, text=error_msg, wraplength=550).pack(pady=(0, 10))
                
                # Scrollable text for full error
                text_frame = ttk.Frame(frame)
                text_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
                
                scrollbar = ttk.Scrollbar(text_frame)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, height=15)
                text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.config(command=text_widget.yview)
                
                text_widget.insert("1.0", "Full Error Details:\n\n" + error_details)
                text_widget.config(state=tk.DISABLED)
                
                ttk.Button(frame, text="Close", command=error_window.destroy).pack()
            
            self.after(0, _on_error)
    
    def start_analysis(self):
        """Start Monte Carlo analysis"""
        print("DEBUG: start_analysis called")
        
        # Validation
        iterations = self.var_iterations.get()
        confidence = self.var_confidence.get()
        
        print(f"DEBUG: iterations={iterations}, confidence={confidence}")
        
        if iterations < 100 or iterations > 50000:
            messagebox.showerror("Error", "Iterations must be between 100 and 50000")
            return
        
        if confidence < 0.9 or confidence > 0.99:
            messagebox.showerror("Error", "Confidence level must be between 0.90 and 0.99")
            return
        
        # Get and validate run name
        run_name = self.var_run_name.get().strip()
        if run_name:
            # Sanitize run name for filename
            import re
            run_name = re.sub(r'[<>:"/\\|?*]', '_', run_name)  # Replace invalid filename chars
            run_name = run_name[:50]  # Limit length
        
        config = MonteCarloConfig(
            iterations=iterations,
            confidence_level=confidence,
            random_seed=self.var_seed.get() if self.var_use_seed.get() else None,
            run_name=run_name
        )
        
        print("DEBUG: Config created, disabling buttons")
        
        # Buttons deaktivieren
        self.start_button.config(state='disabled')
        self.cancel_button.config(state='disabled')
        
        print("DEBUG: Starting thread")
        
        # Start analysis in separate thread
        thread = threading.Thread(target=self.run_analysis_thread, args=(config,), daemon=True)
        thread.start()
        
        print("DEBUG: Thread started")
    

class MonteCarloAnalyzer:
    """Monte Carlo Analyzer - Vollständige Integration"""
    
    def __init__(self, config_path: str, mc_config: MonteCarloConfig, status_callback=None):
        self.config_path = config_path
        self.mc_config = mc_config
        self.log_messages = []
        self.status_callback = status_callback
        
        # Seed setzen
        if mc_config.random_seed is not None:
            np.random.seed(mc_config.random_seed)
            self.log(f"Random seed gesetzt: {mc_config.random_seed}")
    
    def log(self, message: str):
        """Log-Nachricht hinzufügen"""
        try:
            print(message)
        except (OSError, UnicodeEncodeError):
            # Fallback für Windows-Konsolen-Probleme
            try:
                print(message.encode('utf-8', errors='replace').decode('utf-8'))
            except:
                pass  # Ignoriere Print-Fehler
        self.log_messages.append(message)
    
    def run_analysis(self):
        """Vollständige Monte Carlo Analyse durchführen - nur EHB"""
        self.log("=" * 80)
        self.log("MONTE CARLO ANALYSE START - EHB")
        self.log("=" * 80)
        
        # 1. Konfiguration laden
        self.log("\n1. Lade Konfiguration...")
        with open(self.config_path, 'r', encoding='utf-8') as f:
            self.project_config = json.load(f)
        self.log(f"   Konfiguration geladen: {self.config_path}")
        
        # Daten laden
        self.log("\n2. Lade Anlagendaten...")
        data_path = Path(self.project_config.get('excel_path', 'Output/UWWTD_TP_Database.xlsx'))
        if not data_path.exists():
            raise FileNotFoundError(f"Datenbank nicht gefunden: {data_path}")
        
        # Benötigte Sheets sammeln
        sheets_needed = set()
        categories = self.project_config.get('categories', {})
        for cat_data in categories.values():
            sheet = cat_data.get('sheet')
            if sheet:
                sheets_needed.add(sheet)
        
        # H2 Logistics Sheet hinzufügen
        sheets_needed.add('H2 Logistics')
        
        # Füge Grid Energy Connection Sheet hinzu (für Oxygen Usage Potential - Strompreis)
        sheets_needed.add('Grid Energy Connection')
        
        # Lade ID-Spalten
        id_columns = self.project_config.get('id_columns', ['UWWTD Code', 'Name'])
        
        # H2 Logistics als erstes Sheet laden
        first_sheet = 'H2 Logistics'
        
        if first_sheet not in sheets_needed:
            raise ValueError(f"Sheet '{first_sheet}' nicht in benötigten Sheets gefunden")
        
        self.log(f"   Lade Basis-Daten aus Sheet: {first_sheet}")
        self.plant_data = pd.read_excel(data_path, sheet_name=first_sheet)
        self.log(f"   Basis-Daten: {len(self.plant_data)} Anlagen")
        self.log(f"   Verfügbare Spalten: {list(self.plant_data.columns[:10])}...")  # Zeige erste 10 Spalten
        
        # UWWTD Code prüfen
        if 'UWWTD Code' not in self.plant_data.columns:
            raise ValueError(f"UWWTD Code Spalte nicht in {first_sheet} gefunden!")
        self.log(f"   ✓ UWWTD Code Spalte gefunden")
        
        # Merge alle anderen Sheets
        for sheet in sheets_needed:
            if sheet == first_sheet:
                continue
            self.log(f"   Merge Sheet: {sheet}")
            try:
                sheet_data = pd.read_excel(data_path, sheet_name=sheet)
                # Merge auf ID-Spalten
                merge_cols = [col for col in id_columns if col in self.plant_data.columns and col in sheet_data.columns]
                if merge_cols:
                    before_len = len(self.plant_data)
                    self.plant_data = self.plant_data.merge(sheet_data, on=merge_cols, how='left', suffixes=('', '_dup'))
                    after_len = len(self.plant_data)
                    
                    if after_len > before_len:
                        self.log(f"      WARNUNG: Merge hat Duplikate erzeugt ({before_len} → {after_len} Zeilen)")
                        self.log(f"      Entferne Duplikate basierend auf {merge_cols}")
                        # Duplikate entfernen
                        self.plant_data = self.plant_data.drop_duplicates(subset=merge_cols, keep='first')
                        self.log(f"      Nach Duplikat-Entfernung: {len(self.plant_data)} Zeilen")
                    
                    # Duplizierte Spalten entfernen
                    dup_cols = [col for col in self.plant_data.columns if col.endswith('_dup')]
                    if dup_cols:
                        self.plant_data = self.plant_data.drop(columns=dup_cols)
                else:
                    self.log(f"      WARNUNG: Keine gemeinsamen ID-Spalten gefunden für Sheet {sheet}")
            except Exception as e:
                self.log(f"      WARNUNG: Konnte Sheet {sheet} nicht laden: {e}")
        
        self.log(f"   Finale Daten: {len(self.plant_data)} Anlagen, {len(self.plant_data.columns)} Spalten")
        
        # Duplikate prüfen
        if available_id_cols := [col for col in id_columns if col in self.plant_data.columns]:
            duplicates = self.plant_data[available_id_cols].duplicated().sum()
            if duplicates > 0:
                self.log(f"   WARNUNG: {duplicates} Duplikate gefunden - werden entfernt")
                self.plant_data = self.plant_data.drop_duplicates(subset=available_id_cols, keep='first')
                self.log(f"   Nach Duplikat-Entfernung: {len(self.plant_data)} Anlagen")
        
        # ID-Spalten extrahieren
        self.id_columns = self.project_config.get('id_columns', ['Site_ID', 'Name'])
        self.log(f"   ID-Spalten: {self.id_columns}")
        
        # 2b. Lade Built Flags und erstelle Varianten-Masken
        self.log("\n2b. Lade Built Flags für Varianten...")
        self._load_built_flags()
        
        # EHB Analyse
        variant = 'EHB'
        self.log("\n" + "=" * 80)
        self.log(f"VARIANTE: {variant}")
        self.log("=" * 80)
        
        if self.status_callback:
            self.status_callback(f"Processing variant: {variant}")
        
        # Speichere aktuelle Variante für Abnehmer-Score-Berechnung
        self.current_variant = variant
        # Logging-Flag zurücksetzen
        self._offtaker_logged = False
            
        # 3. KPI-Struktur aufbauen (varianten-spezifisch)
        self.log(f"\n3. Baue KPI-Struktur auf für {variant}...")
        self.kpi_structure = self._build_kpi_structure(variant=variant)
        self.log(f"   Gefundene KPIs: {len(self.kpi_structure)}")
        
        # Log varianten-spezifische Spalten-Anpassungen
        logistic_kpis = [kpi for kpi in self.kpi_structure if kpi['category'] == 'Logistic']
        if logistic_kpis:
            for kpi in logistic_kpis:
                self.log(f"      Logistic KPI: '{kpi['name']}' → Spalte: '{kpi['column_name']}'")
        
        # Filtere Daten nach Built Flag
        self.log(f"\n4. Filtere Daten für Variante {variant}...")
        variant_mask = self.variant_masks[variant]
        num_plants = variant_mask.sum()
        self.log(f"   Anlagen mit Built Flag = 1: {num_plants}")
        
        if num_plants == 0:
            self.log(f"   WARNUNG: Keine Anlagen für Variante {variant} - Analyse abgebrochen")
            return
        
        # Erstelle gefilterte Kopie der Daten
        self.current_plant_data = self.plant_data[variant_mask].copy().reset_index(drop=True)
        
        # Nach ID-Spalten sortieren
        available_id_cols = [col for col in self.id_columns if col in self.current_plant_data.columns]
        if available_id_cols:
            self.current_plant_data = self.current_plant_data.sort_values(available_id_cols).reset_index(drop=True)
            self.log(f"   Daten sortiert nach {available_id_cols} für Reproduzierbarkeit")
        
        # 5. Daten vorbereiten (Imputation, Winsorizing, Normalisierung)
        self.log(f"\n5. Bereite Daten vor für {variant}...")
        self.prepared_data = self._prepare_data()
        self.log(f"   Daten vorbereitet: {self.prepared_data.shape}")
        
        # 6. Risiko-Daten laden
        self.log(f"\n6. Lade Risiko-Daten für {variant}...")
        self._load_risk_data()
        self.log(f"   Risiko-Daten geladen")
        
        # 7. Monte Carlo Simulation
        self.log(f"\n7. Führe Monte Carlo Simulation durch für {variant} ({self.mc_config.iterations} Iterationen)...")
        if self.status_callback:
            self.status_callback(f"{variant}: Starting {self.mc_config.iterations} iterations...")
        results = self._monte_carlo_simulation(variant=variant, variant_idx=0)
        
        # 8. Erstelle sortierte Datenbank mit Rankings (nur gefilterte Anlagen)
        self.log(f"\n8. Erstelle sortierte Datenbank für {variant}...")
        if self.status_callback:
            self.status_callback(f"{variant}: Creating ranked database...")
        self._create_ranked_database(results, variant=variant)
        
        # 10. Log speichern
        self._save_log()
        
        if self.status_callback:
            self.status_callback("Analysis completed! Saving results...")
        
        self.log("\n" + "=" * 80)
        self.log("MONTE CARLO ANALYSE ABGESCHLOSSEN - EHB")
        self.log("=" * 80)
    
    def _load_built_flags(self):
        """Lädt Built Flags aus H2 Logistics Sheet und erstellt EHB-Maske"""
        # Suche nach EHB Built Flag Spalte
        ehb_built_flag_col = None
        
        # Mögliche Spaltennamen für EHB Built Flag
        for col in self.plant_data.columns:
            col_lower = col.lower().strip()
            if 'built' in col_lower and 'scenario' in col_lower:
                if 'ehb' in col_lower or 'scenario 1' in col_lower:
                    ehb_built_flag_col = col
                    self.log(f"   EHB Built Flag: {col}")
                    break
        
        # Erstelle Maske für EHB
        self.variant_masks = {}
        if ehb_built_flag_col and ehb_built_flag_col in self.plant_data.columns:
            # Built Flag = 1 bedeutet "bauen"
            mask = pd.to_numeric(self.plant_data[ehb_built_flag_col], errors='coerce').fillna(0) == 1
            self.variant_masks['EHB'] = mask
            self.log(f"   EHB: {mask.sum()} Anlagen mit Built Flag = 1")
        else:
            self.log(f"   WARNUNG: Built Flag für EHB nicht gefunden - verwende alle Anlagen")
            self.variant_masks['EHB'] = pd.Series([True] * len(self.plant_data), index=self.plant_data.index)
    
    def _build_kpi_structure(self, variant: str = None) -> List[Dict]:
        """
        Baut flache KPI-Liste auf, inkl. Energy-Subkategorien.
        Jeder Eintrag: {name, column_name, direction, weight, w_min, w_max, category, subcategory}
        
        Nur für EHB-Variante - keine Spalten-Anpassung mehr nötig.
        """
        kpi_list = []
        categories = self.project_config.get('categories', {})
        
        for cat_name, cat_data in categories.items():
            if cat_name == "Energy" and cat_data.get('method') == 'composite':
                # Energy Subkategorien
                subcats = cat_data.get('subcategories', {})
                for subcat_name, subcat_data in subcats.items():
                    for kpi in subcat_data.get('kpis', []):
                        column_name = self._adapt_column_for_variant(kpi.get('column'), variant)
                        kpi_list.append({
                            'name': kpi.get('name'),
                            'column_name': column_name,
                            'direction': kpi.get('direction', 'benefit'),
                            'weight': kpi.get('weight'),
                            'w_min': kpi.get('w_min'),
                            'w_max': kpi.get('w_max'),
                            'category': cat_name,
                            'subcategory': subcat_name
                        })
            else:
                # Normale Kategorie
                for kpi in cat_data.get('kpis', []):
                    column_name = self._adapt_column_for_variant(kpi.get('column'), variant)
                    kpi_list.append({
                        'name': kpi.get('name'),
                        'column_name': column_name,
                        'direction': kpi.get('direction', 'benefit'),
                        'weight': kpi.get('weight'),
                        'w_min': kpi.get('w_min'),
                        'w_max': kpi.get('w_max'),
                        'category': cat_name,
                        'subcategory': None
                    })
        
        return kpi_list
    
    def _adapt_column_for_variant(self, column_name: str, variant: str = None) -> str:
        """
        Gibt Spalten-Namen unverändert zurück - nur EHB wird unterstützt.
        """
        return column_name
    
    def _prepare_data(self) -> pd.DataFrame:
        """
        Bereitet Daten vor: Imputation → Winsorizing → 0-1-Normalisierung → Richtungsflip
        
        Berechnet auch aggregierte Abnehmer-Scores on-the-fly
        """
        df = self.current_plant_data.copy()
        
        # Berechne aggregierte Abnehmer-Scores (vor Normalisierung)
        self._calculate_offtaker_scores(df)
        
        # Berechne Oxygen Usage Potential (vor Normalisierung)
        self._calculate_oxygen_potential(df)
        
        # Konvertiere alle KPI-Spalten zu numerischen Werten
        self.log("   Konvertiere Spalten zu numerischen Werten...")
        for kpi in self.kpi_structure:
            col = kpi['column_name']
            if col not in df.columns:
                self.log(f"   WARNUNG: Spalte '{col}' (KPI '{kpi['name']}') nicht gefunden - übersprungen")
                continue
            
            # Zu numerisch konvertieren
            df[col] = pd.to_numeric(df[col], errors='coerce')
            non_numeric_count = df[col].isna().sum()
            if non_numeric_count > 0:
                self.log(f"      {col}: {non_numeric_count} nicht-numerische Werte zu NaN konvertiert")
        
        # Imputation (außer Risk-KPIs)
        if self.project_config.get('imputation_enabled', True):
            method = self.project_config.get('imputation_method', 'median')
            self.log(f"   Imputation: {method}")
            
            for kpi in self.kpi_structure:
                col = kpi['column_name']
                if col not in df.columns:
                    continue
                
                # Risk-KPIs überspringen
                if kpi['category'] == 'Risk':
                    continue
                
                # Skip WaterAvailabilityIndex - removed from analysis
                # if col == 'WaterAvailabilityIndex':  # Removed
                #     continue
                
                if df[col].isna().any():
                    if method == 'median':
                        fill_val = df[col].median()
                    else:  # mean
                        fill_val = df[col].mean()
                    
                    df[col].fillna(fill_val, inplace=True)
                    self.log(f"      {col}: {df[col].isna().sum()} NaN mit {fill_val:.2f} gefüllt")
        
        # Winsorizing
        wins_pct = self.project_config.get('winsorize_pct')
        if wins_pct and wins_pct > 0:
            self.log(f"   Winsorizing: {wins_pct*100}%")
            for kpi in self.kpi_structure:
                col = kpi['column_name']
                if col not in df.columns:
                    continue
                
                # Skip Risk-KPIs
                if kpi['category'] == 'Risk':
                    continue
                
                # District Heating KPIs überspringen
                if kpi['category'] == 'District Heating':
                    self.log(f"      {col}: übersprungen (District Heating - kein Winsorizing)")
                    continue
                
                # Skip WaterAvailabilityIndex - removed from analysis
                # if col == 'WaterAvailabilityIndex':  # Removed
                #     continue
                
                lower = df[col].quantile(wins_pct)
                upper = df[col].quantile(1 - wins_pct)
                df[col] = df[col].clip(lower, upper)
                self.log(f"      {col}: geclippt auf [{lower:.2f}, {upper:.2f}]")
        
        # 0-1 Normalisierung (außer Risk-KPIs)
        self.log("   Normalisierung: 0-1 Skalierung")
        for kpi in self.kpi_structure:
            col = kpi['column_name']
            if col not in df.columns:
                continue
            
            # Risk-KPIs überspringen
            if kpi['category'] == 'Risk':
                continue
            
            # Skip WaterAvailabilityIndex - removed from analysis
            # if col == 'WaterAvailabilityIndex':  # Removed
            #     self.log(f"      {col}: übersprungen (bereits heuristisch bewertet)")
            #     continue
            
            col_min = df[col].min()
            col_max = df[col].max()
            
            if col_max - col_min > 1e-9:
                df[col] = (df[col] - col_min) / (col_max - col_min)
            else:
                df[col] = 0.5  # Konstante Spalte
                self.log(f"      {col}: konstant, auf 0.5 gesetzt")
        
        # Richtungsflip für "cost" / "minimize" (außer Risk-KPIs)
        self.log("   Richtungsflip für 'cost'-KPIs")
        for kpi in self.kpi_structure:
            col = kpi['column_name']
            if col not in df.columns:
                continue
            
            # Skip Risk-KPIs
            if kpi['category'] == 'Risk':
                continue
            
            # Skip WaterAvailabilityIndex - removed from analysis
            # if col == 'WaterAvailabilityIndex':  # Removed
            #     continue
            
            if kpi['direction'] in ['cost', 'minimize']:
                df[col] = 1.0 - df[col]
                self.log(f"      {col}: geflippt (war 'cost')")
        
        return df
    
    def _calculate_offtaker_scores(self, df: pd.DataFrame):
        """
        Berechnet aggregierten Abnehmer-Score on-the-fly (binär):
        
        Offtakers Potential = max(Local Score, Pipeline Score)
        
        Wobei:
        - Local Score = (Industry in 10km? × 3) + (Ports? × 2) + (Airports? × 1)
        - Pipeline Score = (Industry along Pipeline? × 3) + (Ports? × 2) + (Airports? × 1)
        
        Nimmt den besseren der beiden Werte - die beste Abnehmer-Situation zählt.
        
        Binäre Bewertung: Abnehmer vorhanden (1) oder nicht (0), unabhängig von der Anzahl
        
        Gewichtung pro Abnehmer-Typ:
        - Industrie: 3.0 (wichtigster Abnehmer)
        - Häfen: 2.0 (Transport-Hub)
        - Flughäfen: 1.0 (kleinerer Abnehmer)
        
        Maximaler Score: 6.0 (alle drei Abnehmer-Typen vorhanden)
        """
        # Flag für einmaliges Logging
        if not hasattr(self, '_offtaker_logged'):
            self._offtaker_logged = False
        
        # Gewichte
        weights = {
            'Industry': 3.0,
            'Ports': 2.0,
            'Airports': 1.0
        }
        
        # 1. Local Offtakers Score (10km Radius) - BINÄR
        local_cols = {
            'Industry in 10km': weights['Industry'],
            'Ports in 10km': weights['Ports'],
            'Airports in 10km': weights['Airports']
        }
        
        local_score = pd.Series(0.0, index=df.index)
        
        for col, weight in local_cols.items():
            if col in df.columns:
                # Konvertiere zu numerisch und prüfe ob > 0 (binär)
                values = pd.to_numeric(df[col], errors='coerce').fillna(0)
                has_offtaker = (values > 0).astype(float)  # 1 wenn vorhanden, 0 sonst
                local_score += has_offtaker * weight
        
        # 2. Pipeline Offtakers Score (entlang EHB Pipeline) - BINÄR
        variant = 'EHB'
        
        pipeline_cols = {
            f'Industry along Pipeline ({variant})': weights['Industry'],
            f'Ports along Pipeline ({variant})': weights['Ports'],
            f'Airports along Pipeline ({variant})': weights['Airports']
        }
        
        pipeline_score = pd.Series(0.0, index=df.index)
        
        for col, weight in pipeline_cols.items():
            if col in df.columns:
                # Konvertiere zu numerisch und prüfe ob > 0 (binär)
                values = pd.to_numeric(df[col], errors='coerce').fillna(0)
                has_offtaker = (values > 0).astype(float)  # 1 wenn vorhanden, 0 sonst
                pipeline_score += has_offtaker * weight
        
        # 3. Kombinierter Score: Maximum von Local und Pipeline
        # Nimmt den besseren der beiden Werte (beste Abnehmer-Situation)
        df['Offtakers Potential'] = pd.concat([local_score, pipeline_score], axis=1).max(axis=1)
        
        # Logging (nur beim ersten Aufruf)
        if not self._offtaker_logged:
            self.log(f"   Abnehmer-Score berechnet (BINÄR: vorhanden=1, nicht vorhanden=0)")
            self.log(f"   Gewichtung: Industrie={weights['Industry']}, Häfen={weights['Ports']}, Flughäfen={weights['Airports']}")
            self.log(f"   Formel: Offtakers Potential = max(Local, Pipeline) - beste Abnehmer-Situation zählt")
            
            # Local Stats
            local_mean = local_score.mean()
            local_max = local_score.max()
            local_nonzero = (local_score > 0).sum()
            self.log(f"      Local: Ø {local_mean:.2f}, Max {local_max:.0f}, {local_nonzero} Anlagen")
            
            # Pipeline Stats
            pipeline_mean = pipeline_score.mean()
            pipeline_max = pipeline_score.max()
            pipeline_nonzero = (pipeline_score > 0).sum()
            self.log(f"      Pipeline: Ø {pipeline_mean:.2f}, Max {pipeline_max:.0f}, {pipeline_nonzero} Anlagen")
            
            # Combined Stats
            combined_mean = df['Offtakers Potential'].mean()
            combined_max = df['Offtakers Potential'].max()
            combined_nonzero = (df['Offtakers Potential'] > 0).sum()
            combined_dist = df['Offtakers Potential'].value_counts().sort_index()
            self.log(f"      Kombiniert: Ø {combined_mean:.2f}, Max {combined_max:.0f}, {combined_nonzero} Anlagen mit Abnehmern")
            self.log(f"         Verteilung: {dict(list(combined_dist.items())[:10])}")  # Zeige erste 10
            
            self._offtaker_logged = True
    
    def _calculate_oxygen_potential(self, df: pd.DataFrame):
        """
        Berechnet Oxygen Usage Potential als Produkt von:
        - Capacity/PE (Anlagengröße)
        - Annual Average Price [ct/kWh] 2024 (Strompreis)
        
        Formel: Oxygen_Usage_Potential = Capacity/PE × Price
        
        Interpretation: Höhere Werte = größeres Potenzial
        (große Anlage + hoher Strompreis = lohnt sich mehr für Sauerstoffproduktion)
        """
        capacity_col = 'Capacity/PE'
        price_col = 'Annual Average Price [ct/kWh] 2024'
        
        # Spalten prüfen
        missing_cols = []
        if capacity_col not in df.columns:
            missing_cols.append(capacity_col)
        if price_col not in df.columns:
            missing_cols.append(price_col)
        
        if missing_cols:
            self.log(f"   WARNUNG: Oxygen Usage Potential kann nicht berechnet werden - fehlende Spalten: {missing_cols}")
            # Dummy-Spalte erstellen
            df['Oxygen_Usage_Potential'] = np.nan
            return
        
        # Konvertiere zu numerisch
        capacity = pd.to_numeric(df[capacity_col], errors='coerce')
        price = pd.to_numeric(df[price_col], errors='coerce')
        
        # Produkt berechnen
        df['Oxygen_Usage_Potential'] = capacity * price
        
        # Logging
        valid_count = df['Oxygen_Usage_Potential'].notna().sum()
        if valid_count > 0:
            self.log(f"   Oxygen Usage Potential berechnet: {valid_count} Anlagen")
            self.log(f"      Formel: Capacity/PE × Price [ct/kWh]")
            self.log(f"      Min: {df['Oxygen_Usage_Potential'].min():.2f}, Max: {df['Oxygen_Usage_Potential'].max():.2f}, Ø: {df['Oxygen_Usage_Potential'].mean():.2f}")
        else:
            self.log(f"   WARNUNG: Oxygen Usage Potential konnte für keine Anlage berechnet werden")
    
    def _load_risk_data(self):
        """
        Lädt Risiko-Daten und berechnet Risk-Scores mit linearer Interpolation zwischen safe und critical.
        
        Logik (wie vorher):
        - Distanz-KPIs: > safe = 0 (kein Risiko), < critical = 1 (hohes Risiko), dazwischen linear
        - Area-KPI: > sufficient = 0 (kein Risiko), = critical = 1 (hohes Risiko), dazwischen linear
        - ABER: Wir invertieren am Ende, damit höher = besser (für Konsistenz mit anderen KPIs)
        
        WICHTIG: Risk-Daten müssen mit current_plant_data synchronisiert werden (gleiche Zeilen-Reihenfolge)!
        """
        risk_sheet = self.project_config.get('risk_sheet', 'Risks')
        risk_cols = self.project_config.get('risk_cols', {})
        risk_thresholds = self.project_config.get('risk_thresholds', {})
        
        self.log("   Berechne Risk-KPIs mit linearer Interpolation...")
        
        # Risiko-Daten laden
        try:
            risk_data = pd.read_excel(
                Path(self.project_config.get('excel_path', 'Output/UWWTD_TP_Database.xlsx')),
                sheet_name=risk_sheet
            )
        except Exception as e:
            self.log(f"   WARNUNG: Risiko-Sheet '{risk_sheet}' nicht gefunden: {e}")
            self.log("   Risiko-KPIs werden übersprungen")
            return
        
        # WICHTIG: Merge risk_data mit current_plant_data basierend auf ID-Spalten
        # um sicherzustellen, dass die Zeilen-Reihenfolge übereinstimmt
        available_id_cols = [col for col in self.id_columns if col in self.current_plant_data.columns and col in risk_data.columns]
        
        if not available_id_cols:
            self.log(f"   WARNUNG: Keine gemeinsamen ID-Spalten zwischen current_plant_data und risk_data gefunden")
            self.log(f"   ID-Spalten in config: {self.id_columns}")
            self.log(f"   Verfügbar in current_plant_data: {list(self.current_plant_data.columns[:10])}")
            self.log(f"   Verfügbar in risk_data: {list(risk_data.columns[:10])}")
            return
        
        # Merge auf ID-Spalten - nur die Risk-Spalten übernehmen
        risk_columns_to_merge = [col for col in risk_cols.values() if col and col in risk_data.columns]
        if not risk_columns_to_merge:
            self.log(f"   WARNUNG: Keine Risk-Spalten in risk_data gefunden")
            return
        
        # Erstelle temporären DataFrame mit ID + Risk-Spalten
        risk_data_subset = risk_data[available_id_cols + risk_columns_to_merge].copy()
        
        # DEBUG: Log erste paar IDs
        if len(self.current_plant_data) > 0 and available_id_cols:
            first_id_col = available_id_cols[0]
            self.log(f"   DEBUG: Erste 3 IDs in current_plant_data: {list(self.current_plant_data[first_id_col].head(3))}")
            if first_id_col in risk_data_subset.columns:
                self.log(f"   DEBUG: Erste 3 IDs in risk_data: {list(risk_data_subset[first_id_col].head(3))}")
        
        # Merge mit current_plant_data (left join, um Reihenfolge beizubehalten)
        # WICHTIG: reset_index() um sicherzustellen, dass der Index mit current_plant_data übereinstimmt
        merged = self.current_plant_data[available_id_cols].merge(
            risk_data_subset, 
            on=available_id_cols, 
            how='left'
        ).reset_index(drop=True)
        
        self.log(f"   Risk-Daten gemerged: {len(merged)} Zeilen")
        
        # Merge prüfen
        if len(merged) != len(self.current_plant_data):
            self.log(f"   WARNUNG: Merge hat Zeilen-Anzahl geändert! {len(self.current_plant_data)} → {len(merged)}")
        
        # DEBUG: Prüfe auf NaN-Werte in Risk-Spalten
        for col in risk_columns_to_merge:
            if col in merged.columns:
                nan_count = merged[col].isna().sum()
                if nan_count > 0:
                    self.log(f"   WARNUNG: {nan_count} NaN-Werte in Risk-Spalte '{col}' nach Merge")
        
        # Risk-Scores berechnen
        
        # 1. Flood Risk
        col = risk_cols.get('flood', '')
        if col and col in merged.columns:
            safe = float(risk_thresholds.get('flood', {}).get('safe', 500))
            crit = float(risk_thresholds.get('flood', {}).get('critical', 100))
            
            # WICHTIG: Behandle "> 5000" Strings als 5000m (nicht als NaN/0)
            dist_raw = merged[col].astype(str)
            dist = pd.Series(index=dist_raw.index, dtype=float)
            for idx, val in dist_raw.items():
                if '>' in val or 'greater' in val.lower():
                    # "> 5000" oder ähnlich → verwende 5000
                    dist.loc[idx] = 5000.0
                else:
                    try:
                        dist.loc[idx] = float(val)
                    except:
                        dist.loc[idx] = 0.0  # Fallback für ungültige Werte
            
            # DEBUG: Log einige Beispielwerte
            if len(dist) > 0:
                self.log(f"      Flood Distanzen - Min: {dist.min():.1f}m, Max: {dist.max():.1f}m, Median: {dist.median():.1f}m")
                greater_than_count = (dist >= 5000).sum()
                if greater_than_count > 0:
                    self.log(f"      Flood: {greater_than_count} Anlagen mit ≥5000m (ursprünglich '> 5000')")
            
            # Linear: dist > safe → risk_score = 0
            # dist < crit → risk_score = 1
            risk_score = np.clip((safe - dist) / (safe - crit), 0, 1)
            
            # Invertieren: höher = besser (1 = kein Risiko, 0 = hohes Risiko)
            inverted_score = 1.0 - risk_score
            
            # DEBUG: Log Score-Verteilung
            self.log(f"      Flood Scores - Min: {inverted_score.min():.3f}, Max: {inverted_score.max():.3f}, Mean: {inverted_score.mean():.3f}")
            
            self.prepared_data[col] = inverted_score.values  # .values um Index-Probleme zu vermeiden
            self.log(f"      Flood: safe={safe}m, critical={crit}m → invertiert (höher = besser)")
            self.log(f"      Flood: Spalte '{col}' zu prepared_data hinzugefügt ({len(inverted_score)} Werte)")
        
        # 2. Residential Risk
        col = risk_cols.get('residential', '')
        if col and col in merged.columns:
            safe = float(risk_thresholds.get('residential', {}).get('safe', 500))
            crit = float(risk_thresholds.get('residential', {}).get('critical', 200))
            dist = pd.to_numeric(merged[col], errors='coerce').fillna(0)
            
            if len(dist) > 0:
                self.log(f"      Residential Distanzen - Min: {dist.min():.1f}m, Max: {dist.max():.1f}m, Median: {dist.median():.1f}m")
            
            risk_score = np.clip((safe - dist) / (safe - crit), 0, 1)
            inverted_score = 1.0 - risk_score
            
            self.log(f"      Residential Scores - Min: {inverted_score.min():.3f}, Max: {inverted_score.max():.3f}, Mean: {inverted_score.mean():.3f}")
            
            self.prepared_data[col] = inverted_score.values
            self.log(f"      Residential: safe={safe}m, critical={crit}m → invertiert (höher = besser)")
            self.log(f"      Residential: Spalte '{col}' zu prepared_data hinzugefügt ({len(inverted_score)} Werte)")
        
        # 3. Protected Area Risk
        col = risk_cols.get('protected', '')
        if col and col in merged.columns:
            safe = float(risk_thresholds.get('protected', {}).get('safe', 500))
            crit = float(risk_thresholds.get('protected', {}).get('critical', 200))
            dist = pd.to_numeric(merged[col], errors='coerce').fillna(0)
            
            if len(dist) > 0:
                self.log(f"      Protected Distanzen - Min: {dist.min():.1f}m, Max: {dist.max():.1f}m, Median: {dist.median():.1f}m")
            
            risk_score = np.clip((safe - dist) / (safe - crit), 0, 1)
            inverted_score = 1.0 - risk_score
            
            self.log(f"      Protected Scores - Min: {inverted_score.min():.3f}, Max: {inverted_score.max():.3f}, Mean: {inverted_score.mean():.3f}")
            
            self.prepared_data[col] = inverted_score.values
            self.log(f"      Protected: safe={safe}m, critical={crit}m → invertiert (höher = besser)")
            self.log(f"      Protected: Spalte '{col}' zu prepared_data hinzugefügt ({len(inverted_score)} Werte)")
        
        # Risk-Spalten in prepared_data prüfen
        self.log(f"\n   DEBUG: Risk-Spalten in prepared_data:")
        for kpi in self.kpi_structure:
            if kpi['category'] == 'Risk':
                col = kpi['column_name']
                if col in self.prepared_data.columns:
                    self.log(f"      ✓ {col}: vorhanden")
                else:
                    self.log(f"      ✗ {col}: FEHLT!")
        

    
    def _monte_carlo_simulation(self, variant: str = None, variant_idx: int = 0) -> Dict:
        """
        Führt Monte Carlo Simulation durch.
        Für jede Iteration:
          1. Sample Gewichte (±uncertainty)
          2. Berechne Energy-Score aus Subkategorien
          3. Berechne Gesamt-Score (gewichtete Summe aller Kategorien inkl. Risk)
          4. Berechne Ränge
        
        Rückgabe: Dict mit all_scores_df, rankings, statistics
        """
        n_iter = self.mc_config.iterations
        n_plants = len(self.prepared_data)
        
        # Seed erneut setzen
        if self.mc_config.random_seed is not None:
            # Verwende unterschiedliche Seeds für verschiedene Varianten
            seed = self.mc_config.random_seed + variant_idx * 10000
            np.random.seed(seed)
            self.log(f"   Random seed für Simulation gesetzt: {seed}")
        
        # Matrix: Zeilen = Iterationen, Spalten = Anlagen
        all_base_scores = np.zeros((n_iter, n_plants))
        all_final_scores = np.zeros((n_iter, n_plants))
        all_rankings = np.zeros((n_iter, n_plants), dtype=int)
        
        # Kategorie-Scores speichern
        categories_config = self.project_config.get('categories', {})
        category_names = list(categories_config.keys())
        all_category_scores = {cat_name: np.zeros((n_iter, n_plants)) for cat_name in category_names}
        
        for iteration in range(n_iter):
            if iteration % 100 == 0:
                self.log(f"   Iteration {iteration}/{n_iter}...")
                if self.status_callback:
                    self.status_callback(f"{variant}: Iteration {iteration}/{n_iter}")
            
            # 1. Sample Gewichte (KPIs und Kategorien)
            sampled_kpi_weights, sampled_cat_weights = self._sample_weights()
            
            # 2. Berechne Scores für alle Anlagen (inkl. Risk als Kategorie)
            final_scores, category_scores = self._calculate_scores_for_iteration(sampled_kpi_weights, sampled_cat_weights, debug=(iteration == 0), return_category_scores=True)
            
            # Speichere Kategorie-Scores für diese Iteration
            for cat_name, cat_score in category_scores.items():
                all_category_scores[cat_name][iteration, :] = cat_score
            
            # Debug: Zeige Score-Range in erster Iteration
            if iteration == 0:
                self.log(f"   DEBUG: Score Range: min={final_scores.min():.4f}, max={final_scores.max():.4f}, mean={final_scores.mean():.4f}")
                self.log(f"   DEBUG: Anzahl Kategorien mit Gewichten: {len(sampled_cat_weights)}")
                self.log(f"   DEBUG: Kategorie-Gewichte: {sampled_cat_weights}")
                self.log(f"   DEBUG: Anzahl KPI-Gewichte: {len(sampled_kpi_weights)}")
                self.log(f"   DEBUG: Erste 5 KPI-Gewichte: {dict(list(sampled_kpi_weights.items())[:5])}")
            
            # 3. Speichern (base_scores = final_scores, da kein Penalty mehr)
            all_base_scores[iteration, :] = final_scores
            all_final_scores[iteration, :] = final_scores
            
            # Ränge berechnen
            sorted_indices = np.argsort(-final_scores)  # Minus für absteigend
            ranks = np.empty_like(sorted_indices)
            ranks[sorted_indices] = np.arange(1, n_plants + 1)
            all_rankings[iteration, :] = ranks
        
        self.log(f"   {n_iter} Iterationen abgeschlossen")
        
        # Berechne Konfidenzintervalle für Rankings
        confidence = self.mc_config.confidence_level
        lower_percentile = (1 - confidence) / 2 * 100  # z.B. 2.5 für 95% CI
        upper_percentile = (1 - (1 - confidence) / 2) * 100  # z.B. 97.5 für 95% CI
        
        rank_ci_lower = np.percentile(all_rankings, lower_percentile, axis=0)
        rank_ci_upper = np.percentile(all_rankings, upper_percentile, axis=0)
        
        # Berechne auch Konfidenzintervalle für Scores
        score_ci_lower = np.percentile(all_final_scores, lower_percentile, axis=0)
        score_ci_upper = np.percentile(all_final_scores, upper_percentile, axis=0)
        
        self.log(f"   Konfidenzintervalle berechnet ({confidence*100:.1f}%)")
        self.log(f"      Rank CI: [{lower_percentile:.1f}%, {upper_percentile:.1f}%] Perzentile")
        
        # Statistiken berechnen
        results = {
            'all_base_scores': all_base_scores,
            'all_final_scores': all_final_scores,
            'all_rankings': all_rankings,
            'mean_final_score': all_final_scores.mean(axis=0),
            'mean_rank': all_rankings.mean(axis=0),
            'std_rank': all_rankings.std(axis=0),
            'top20_frequency': self._calculate_top20_freq(all_rankings),
            # Konfidenzintervalle
            'rank_ci_lower': rank_ci_lower,
            'rank_ci_upper': rank_ci_upper,
            'score_ci_lower': score_ci_lower,
            'score_ci_upper': score_ci_upper,
            'confidence_level': confidence,
            # Kategorie-Scores
            'mean_category_scores': {cat_name: scores.mean(axis=0) for cat_name, scores in all_category_scores.items()}
        }
        
        self.log(f"   Statistiken berechnet (inkl. Kategorie-Scores und Konfidenzintervalle)")
        
        return results
    
    def _sample_weights(self) -> Tuple[Dict[str, float], Dict[str, float]]:
        """
        Sampelt Gewichte gleichverteilt zwischen w_min und w_max.
        Renormalisiert innerhalb jeder Kategorie (Summe = 1).
        Risk ist jetzt eine normale Kategorie (nicht mehr separates Penalty).
        
        Returns:
            Tuple[Dict[str, float], Dict[str, float]]: (kpi_weights, category_weights)
        """
        sampled_kpis = {}
        sampled_categories = {}
        
        # 1. Sample Kategorie-Gewichte (inkl. Risk)
        categories_config = self.project_config.get('categories', {})
        missing_cat_weights = []
        
        for cat_name, cat_data in categories_config.items():
            if cat_name == "Energy":
                continue  # Energy hat keine direkte Gewichtung
            
            # Kategorie Gewichte prüfen
            w_min = cat_data.get('w_min')
            w_max = cat_data.get('w_max')
            
            # FEHLER: Wenn w_min oder w_max fehlt, Analyse abbrechen
            if w_min is None or w_max is None:
                missing_cat_weights.append(cat_name)
                continue
            
            # Sample gleichverteilt zwischen w_min und w_max
            new_weight = np.random.uniform(w_min, w_max)
            new_weight = max(0.01, new_weight)
            sampled_categories[cat_name] = new_weight
        
        # Prüfe ob Gewichte fehlen
        if missing_cat_weights:
            error_msg = f"Fehlende Kategorie-Gewichte (w_min und w_max müssen beide gesetzt sein):\n" + "\n".join(f"  - {cat}" for cat in missing_cat_weights)
            self.log(f"\nFEHLER: {error_msg}")
            raise ValueError(error_msg)
        
        # Renormalisiere Kategorie-Gewichte (Summe = 1)
        if sampled_categories:
            total = sum(sampled_categories.values())
            if total > 0:
                sampled_categories = {k: v/total for k, v in sampled_categories.items()}
        else:
            raise ValueError("Keine Kategorie-Gewichte definiert. Bitte w_min und w_max für alle Kategorien setzen.")
        
        # 2. Sample KPI-Gewichte (innerhalb jeder Kategorie, inkl. Risk)
        # Gruppiere KPIs nach Kategorie/Subkategorie
        categories = {}
        missing_kpi_weights = []
        
        # Zähle KPIs pro Kategorie (für Single-KPI-Kategorien)
        kpis_per_category = {}
        for kpi in self.kpi_structure:
            cat_key = kpi['category']
            if cat_key not in kpis_per_category:
                kpis_per_category[cat_key] = []
            kpis_per_category[cat_key].append(kpi)
        
        for kpi in self.kpi_structure:
            cat_key = kpi['category']
            is_single_kpi_category = len(kpis_per_category.get(cat_key, [])) == 1
            
            # Bei Single-KPI-Kategorien: Verwende Kategorie-Gewichte statt KPI-Gewichte
            if is_single_kpi_category:
                # Hole Kategorie-Gewichte aus categories_config
                cat_data = categories_config.get(cat_key, {})
                cat_w_min = cat_data.get('w_min')
                cat_w_max = cat_data.get('w_max')
                
                # Übertrage Kategorie-Gewichte auf KPI
                if cat_w_min is not None and cat_w_max is not None:
                    kpi['w_min'] = cat_w_min
                    kpi['w_max'] = cat_w_max
                else:
                    # Kategorie-Gewichte fehlen (sollte bereits oben geprüft worden sein)
                    kpi_full_name = f"{kpi['category']}::{kpi['name']}"
                    missing_kpi_weights.append(f"{kpi_full_name} (verwende Kategorie-Gewichte)")
                    continue
            else:
                # Multi-KPI-Kategorie: Prüfe KPI-Gewichte
                w_min = kpi.get('w_min')
                w_max = kpi.get('w_max')
                
                if w_min is None or w_max is None:
                    kpi_full_name = f"{kpi['category']}::{kpi['name']}"
                    missing_kpi_weights.append(kpi_full_name)
                    continue
            
            key = (kpi['category'], kpi['subcategory'])
            if key not in categories:
                categories[key] = []
            categories[key].append(kpi)
        
        # Prüfe ob KPI-Gewichte fehlen
        if missing_kpi_weights:
            error_msg = f"Fehlende KPI-Gewichte (w_min und w_max müssen beide gesetzt sein):\n" + "\n".join(f"  - {kpi}" for kpi in missing_kpi_weights)
            self.log(f"\nFEHLER: {error_msg}")
            raise ValueError(error_msg)
        
        # Debug: Zeige welche KPIs gefunden wurden (nur beim ersten Aufruf)
        if not hasattr(self, '_weights_logged'):
            self._weights_logged = False
        
        if not self._weights_logged:
            total_kpis = sum(len(kpis) for kpis in categories.values())
            self.log(f"   Gefundene KPIs mit Gewichten: {total_kpis}")
            for (cat, subcat), kpis in categories.items():
                cat_name = f"{cat}::{subcat}" if subcat else cat
                self.log(f"      {cat_name}: {len(kpis)} KPIs")
            self._weights_logged = True
        
        # Sample und renormalisiere pro Kategorie
        for (cat, subcat), kpis in categories.items():
            weights = []
            names = []
            
            for kpi in kpis:
                w_min = kpi.get('w_min')
                w_max = kpi.get('w_max')
                
                # Sample gleichverteilt zwischen w_min und w_max
                new_weight = np.random.uniform(w_min, w_max)
                
                new_weight = max(0.01, new_weight)  # Mindestens positiv
                
                weights.append(new_weight)
                names.append(kpi['column_name'])
            
            # Renormalisierung (Summe = 1)
            total = sum(weights)
            if total > 0:
                weights = [w / total for w in weights]
            
            for name, weight in zip(names, weights):
                sampled_kpis[name] = weight
        
        return sampled_kpis, sampled_categories
    
    def _calculate_scores_for_iteration(self, kpi_weights: Dict[str, float], cat_weights: Dict[str, float], debug: bool = False, return_category_scores: bool = False):
        """
        Berechnet base_scores für eine Iteration.
        
        Logik:
        1. Für jede normale Kategorie: gewichtete Summe der KPIs
        2. Für Energy: 
           a) Berechne Score für Grid Electricity (gewichtete Summe)
           b) Berechne Score für EE Potential (gewichtete Summe)
           c) Aggregiere zu einem Energy-Score gemäß sub_agg_method
        3. Gewichte Kategorie-Scores mit Kategorie-Gewichten
        4. Summiere alle gewichteten Kategorie-Scores
        
        Args:
            return_category_scores: Wenn True, gibt (total_score, category_scores) zurück, sonst nur total_score
        """
        n_plants = len(self.prepared_data)
        category_scores = {}
        
        categories_config = self.project_config.get('categories', {})
        
        for cat_name, cat_data in categories_config.items():
            if cat_name == "Energy" and cat_data.get('method') == 'composite':
                # Energy mit Subkategorien
                subcat_scores = {}
                
                for subcat_name, subcat_data in cat_data.get('subcategories', {}).items():
                    # Nur noch custom (weighted sum)
                    subcat_scores[subcat_name] = self._calculate_weighted_sum_score(subcat_name, kpi_weights)
                
                # Aggregiere Subkategorien zu Energy-Score
                agg_method = cat_data.get('sub_agg_method', 'weighted_average')
                agg_weights = cat_data.get('sub_agg_weights', {})
                
                if agg_method == 'weighted_average':
                    w_grid = agg_weights.get('Grid Electricity', 1.0)
                    w_ee = agg_weights.get('EE Potential', 1.0)
                    total_w = w_grid + w_ee
                    
                    energy_score = (
                        (subcat_scores.get('Grid Electricity', 0) * w_grid +
                         subcat_scores.get('EE Potential', 0) * w_ee) / total_w
                    )
                elif agg_method == 'average':
                    energy_score = np.mean([subcat_scores.get('Grid Electricity', 0),
                                           subcat_scores.get('EE Potential', 0)], axis=0)
                else:  # sum
                    energy_score = (subcat_scores.get('Grid Electricity', 0) +
                                   subcat_scores.get('EE Potential', 0))
                
                category_scores[cat_name] = energy_score
            
            else:
                # Normale Kategorie - nur noch custom (weighted sum)
                category_scores[cat_name] = self._calculate_weighted_sum_score(cat_name, kpi_weights, debug=debug)
        
        # Debug: Zeige Kategorie-Scores
        if debug:
            self.log(f"   DEBUG: Kategorie-Scores (vor Gewichtung):")
            for cat_name, score in category_scores.items():
                if isinstance(score, np.ndarray):
                    self.log(f"      {cat_name}: min={score.min():.4f}, max={score.max():.4f}, mean={score.mean():.4f}")
                else:
                    self.log(f"      {cat_name}: {score}")
        
        # Gewichte und summiere Kategorie-Scores
        total_score = np.zeros(n_plants)
        
        # Kategorie-Gewichte müssen vorhanden sein (wurde bereits in _sample_weights geprüft)
        if not cat_weights:
            raise ValueError("Keine Kategorie-Gewichte vorhanden. Dies sollte nicht passieren.")
        
        # Mit Kategorie-Gewichten
        for cat_name, score in category_scores.items():
            cat_weight = cat_weights.get(cat_name)
            if cat_weight is None:
                raise ValueError(f"Kategorie-Gewicht für '{cat_name}' fehlt. Dies sollte nicht passieren.")
            
            if isinstance(score, np.ndarray):
                weighted_score = score * cat_weight
                total_score += weighted_score
                if debug:
                    self.log(f"      {cat_name} (gewichtet mit {cat_weight:.4f}): min={weighted_score.min():.4f}, max={weighted_score.max():.4f}")
            else:
                weighted_score = np.array(score) * cat_weight
                total_score += weighted_score
                if debug:
                    self.log(f"      {cat_name} (gewichtet mit {cat_weight:.4f}): {weighted_score}")
        
        # Stelle sicher, dass der Score zwischen 0 und 1 liegt
        # Dies sollte bereits der Fall sein, aber wir clippen zur Sicherheit
        total_score = np.clip(total_score, 0.0, 1.0)
        
        if return_category_scores:
            return total_score, category_scores
        else:
            return total_score
    
    def _calculate_weighted_sum_score(self, category_or_subcat: str, weights: Dict[str, float], debug: bool = False) -> np.ndarray:
        """Berechnet gewichtete Summe für eine Kategorie/Subkategorie"""
        score = np.zeros(len(self.prepared_data))
        kpi_count = 0
        
        if debug:
            self.log(f"      Berechne Score für Kategorie: {category_or_subcat}")
        
        # Spezielle Logik für Grid Electricity: Dynamische Gewichtung basierend auf Entfernungen
        if category_or_subcat == "Grid Electricity":
            score = self._apply_grid_electricity_logic(weights, debug)
            return score
        
        for kpi in self.kpi_structure:
            # Prüfe ob KPI zu dieser Kategorie/Subkategorie gehört
            if kpi['subcategory']:
                if kpi['subcategory'] != category_or_subcat:
                    continue
            else:
                if kpi['category'] != category_or_subcat:
                    continue
            
            # KPI gehört zu dieser Kategorie
            col = kpi['column_name']
            if col not in self.prepared_data.columns:
                if debug:
                    self.log(f"         WARNUNG: KPI '{kpi['name']}' - Spalte '{col}' nicht in prepared_data gefunden")
                continue
            
            weight = weights.get(col, 0)
            if weight == 0:
                if debug:
                    self.log(f"         WARNUNG: KPI '{kpi['name']}' (col='{col}') hat Gewicht 0")
                continue
            
            kpi_values = self.prepared_data[col].values
            weighted_values = weight * kpi_values
            score += weighted_values
            kpi_count += 1
            
            if debug:
                self.log(f"         KPI '{kpi['name']}' (col='{col}'): weight={weight:.4f}, values: min={kpi_values.min():.4f}, max={kpi_values.max():.4f}, mean={kpi_values.mean():.4f}, weighted: min={weighted_values.min():.4f}, max={weighted_values.max():.4f}")
        
        if debug:
            self.log(f"      {category_or_subcat}: {kpi_count} KPIs verarbeitet, Score: min={score.min():.4f}, max={score.max():.4f}, mean={score.mean():.4f}")
        
        # District Heating: Conditional Check
        if category_or_subcat == "District Heating":
            score = self._apply_district_heating_conditional(score)
        
        return score
    
    def _apply_grid_electricity_logic(self, weights: Dict[str, float], debug: bool = False) -> np.ndarray:
        """
        Wendet dynamische Gewichtungslogik für Grid Electricity an:
        
        Wenn Entfernung zum Umspannwerk (UW) <= Entfernung zur Trasse (oder nur minimal weiter):
        - Gewicht für UW = 1.0
        - Gewicht für Trasse = 0.0
        
        Sonst:
        - Verwende die gesampelten Gewichte (normalisiert)
        
        "Minimal weiter" = bis zu 10% weiter als zur Trasse
        """
        # Flag für einmaliges Logging
        if not hasattr(self, '_grid_elec_logged'):
            self._grid_elec_logged = False
        
        # Finde die beiden KPIs
        col_trasse = None
        col_uw = None
        
        for kpi in self.kpi_structure:
            if kpi.get('subcategory') == 'Grid Electricity':
                col = kpi['column_name']
                if 'Power' in col or 'power' in col or 'Line' in col or 'line' in col:
                    col_trasse = col
                elif 'Substation' in col or 'substation' in col:
                    col_uw = col
        
        if col_trasse is None or col_uw is None:
            if not self._grid_elec_logged:
                self.log(f"   WARNUNG: Grid Electricity KPIs nicht gefunden (Trasse: {col_trasse}, UW: {col_uw})")
                self.log(f"   Verwende Standard-Gewichtung")
                self._grid_elec_logged = True
            # Fallback: Standard-Gewichtung
            return self._calculate_weighted_sum_score_standard('Grid Electricity', weights, debug)
        
        # Prüfe ob Spalten in prepared_data vorhanden sind
        if col_trasse not in self.prepared_data.columns or col_uw not in self.prepared_data.columns:
            if not self._grid_elec_logged:
                self.log(f"   WARNUNG: Grid Electricity Spalten nicht in prepared_data gefunden")
                self.log(f"   Trasse: {col_trasse in self.prepared_data.columns}, UW: {col_uw in self.prepared_data.columns}")
                self._grid_elec_logged = True
            return self._calculate_weighted_sum_score_standard('Grid Electricity', weights, debug)
        
        # Hole die ORIGINAL-Entfernungen (vor Normalisierung) aus current_plant_data
        # Die prepared_data enthält normalisierte Werte (0-1), wir brauchen die Original-Distanzen
        if col_trasse not in self.current_plant_data.columns or col_uw not in self.current_plant_data.columns:
            if not self._grid_elec_logged:
                self.log(f"   WARNUNG: Grid Electricity Spalten nicht in current_plant_data gefunden")
                self._grid_elec_logged = True
            return self._calculate_weighted_sum_score_standard('Grid Electricity', weights, debug)
        
        dist_trasse_orig = pd.to_numeric(self.current_plant_data[col_trasse], errors='coerce').fillna(999).values
        dist_uw_orig = pd.to_numeric(self.current_plant_data[col_uw], errors='coerce').fillna(999).values
        
        # Hole normalisierte Werte (0-1, geflippt für cost-KPIs)
        values_trasse = self.prepared_data[col_trasse].values
        values_uw = self.prepared_data[col_uw].values
        
        # Berechne dynamische Gewichte für jede Anlage
        # Schwellwert: UW darf bis zu 10% weiter sein als Trasse
        threshold_factor = 1.10
        
        # Maske: UW ist näher oder nur minimal weiter als Trasse
        mask_uw_preferred = dist_uw_orig <= (dist_trasse_orig * threshold_factor)
        
        # Initialisiere Gewichte-Arrays
        weight_trasse_array = np.zeros(len(self.prepared_data))
        weight_uw_array = np.zeros(len(self.prepared_data))
        
        # Hole gesampelte Gewichte
        w_trasse_sampled = weights.get(col_trasse, 0.5)
        w_uw_sampled = weights.get(col_uw, 0.5)
        
        # Normalisiere gesampelte Gewichte (Summe = 1)
        total_sampled = w_trasse_sampled + w_uw_sampled
        if total_sampled > 0:
            w_trasse_norm = w_trasse_sampled / total_sampled
            w_uw_norm = w_uw_sampled / total_sampled
        else:
            w_trasse_norm = 0.5
            w_uw_norm = 0.5
        
        # Setze Gewichte basierend auf Maske
        weight_uw_array[mask_uw_preferred] = 1.0
        weight_trasse_array[mask_uw_preferred] = 0.0
        
        weight_uw_array[~mask_uw_preferred] = w_uw_norm
        weight_trasse_array[~mask_uw_preferred] = w_trasse_norm
        
        # Berechne gewichteten Score
        score = weight_trasse_array * values_trasse + weight_uw_array * values_uw
        
        # Logging (nur beim ersten Aufruf)
        if not self._grid_elec_logged:
            num_uw_preferred = mask_uw_preferred.sum()
            num_trasse_preferred = (~mask_uw_preferred).sum()
            
            self.log(f"   Grid Electricity: Dynamische Gewichtung aktiviert")
            self.log(f"      Schwellwert: UW darf bis zu {(threshold_factor-1)*100:.0f}% weiter sein als Trasse")
            self.log(f"      {num_uw_preferred} Anlagen: UW bevorzugt (Gewicht UW=1.0, Trasse=0.0)")
            self.log(f"      {num_trasse_preferred} Anlagen: Standard-Gewichtung (UW={w_uw_norm:.3f}, Trasse={w_trasse_norm:.3f})")
            
            # Beispiele
            if num_uw_preferred > 0:
                idx_example = np.where(mask_uw_preferred)[0][0]
                self.log(f"         Beispiel UW-bevorzugt: Dist_Trasse={dist_trasse_orig[idx_example]:.2f}km, Dist_UW={dist_uw_orig[idx_example]:.2f}km")
            
            if num_trasse_preferred > 0:
                idx_example = np.where(~mask_uw_preferred)[0][0]
                self.log(f"         Beispiel Standard: Dist_Trasse={dist_trasse_orig[idx_example]:.2f}km, Dist_UW={dist_uw_orig[idx_example]:.2f}km")
            
            self._grid_elec_logged = True
        
        if debug:
            self.log(f"      Grid Electricity: Score berechnet mit dynamischer Gewichtung")
            self.log(f"         Score: min={score.min():.4f}, max={score.max():.4f}, mean={score.mean():.4f}")
        
        return score
    
    def _calculate_weighted_sum_score_standard(self, category_or_subcat: str, weights: Dict[str, float], debug: bool = False) -> np.ndarray:
        """Standard gewichtete Summe ohne spezielle Logik (Fallback)"""
        score = np.zeros(len(self.prepared_data))
        kpi_count = 0
        
        for kpi in self.kpi_structure:
            # Prüfe ob KPI zu dieser Kategorie/Subkategorie gehört
            if kpi['subcategory']:
                if kpi['subcategory'] != category_or_subcat:
                    continue
            else:
                if kpi['category'] != category_or_subcat:
                    continue
            
            col = kpi['column_name']
            if col not in self.prepared_data.columns:
                continue
            
            weight = weights.get(col, 0)
            if weight == 0:
                continue
            
            kpi_values = self.prepared_data[col].values
            weighted_values = weight * kpi_values
            score += weighted_values
            kpi_count += 1
        
        return score
    
    def _apply_district_heating_conditional(self, score: np.ndarray) -> np.ndarray:
        """
        Wendet komplexe District Heating Logik an:
        
        Basierend auf Network und Connection Status:
        - Network = "None" → Score = 0
        - Network = "Study" → Score = 0.2
        - Network ∈ {"Planned", "Existing"}:
            - Connection ∈ {"Planned", "Existing"} → Score = 1.0 (Topkandidat)
            - Connection = "None":
                - Distance ≤ 5 km → Score = 0.4 + 0.5 * S_raw (basierend auf Entfernung & Temperatur)
                - Distance > 5 km → Score = 0.4 (Basis-Score)
        
        S_raw wird berechnet aus:
        - S_dist = max(0, 1 - (distance_km / 5)²)  [0 km → 1.0, ≥5 km → 0.0, quadratisch]
        - S_temp = (T_max - T_net) / (T_max - T_min)  [niedrigere Temp = besser]
        - S_raw = w_dist * S_dist + w_temp * S_temp
        """
        # Flag für einmaliges Logging (nur beim ersten Aufruf)
        if not hasattr(self, '_dh_logged'):
            self._dh_logged = False
        
        try:
            # Verwende die Spalten direkt aus prepared_data (wurden bereits gemerged)
            col_map = {
                'network': 'Network',
                'connection': 'Connection',
                'distance': 'Distance to network',  # in km (oder 'Distance to network [km]')
                'temperature': 'Max. Flow temperature'  # Vorlauftemperatur (oder 'Max. Flow tempature')
            }
            
            # Prüfe ob alle erforderlichen Spalten in prepared_data vorhanden sind
            missing_cols = []
            for key, col_name in col_map.items():
                if col_name not in self.prepared_data.columns:
                    # Versuche alternative Schreibweisen
                    found = False
                    for col in self.prepared_data.columns:
                        if col.lower().replace(' ', '').replace('[km]', '') == col_name.lower().replace(' ', '').replace('[km]', ''):
                            col_map[key] = col
                            found = True
                            break
                    if not found:
                        missing_cols.append(col_name)
            
            if missing_cols:
                self.log(f"   WARNUNG: Fehlende Spalten im District Heating Sheet: {missing_cols}")
                self.log(f"   Verwende Standard-Score ohne District Heating Logik")
                return score
            
            if 'network' not in col_map or 'connection' not in col_map:
                self.log(f"   WARNUNG: Network oder Connection Spalte nicht gefunden - verwende Standard-Score")
                return score
            
            # Initialisiere DH-Score Array mit der richtigen Länge
            dh_score = np.zeros(len(self.prepared_data))
            
            # Normalisiere Status-Werte (case-insensitive)
            network_status_raw = self.prepared_data[col_map['network']].fillna('None').astype(str).str.strip()
            connection_status_raw = self.prepared_data[col_map['connection']].fillna('None').astype(str).str.strip()
            network_status = network_status_raw.str.lower()
            connection_status = connection_status_raw.str.lower()
            
            # Fix common typos: "planed" → "planned"
            network_status = network_status.replace('planed', 'planned')
            connection_status = connection_status.replace('planed', 'planned')
            
            # Debug: Zeige unique Werte
            if not self._dh_logged:
                self.log(f"   District Heating Spalten gefunden:")
                self.log(f"      Network unique values: {sorted(network_status_raw.unique())}")
                self.log(f"      Connection unique values: {sorted(connection_status_raw.unique())}")
            
            # Lade Entfernung für alle Anlagen
            if 'distance' in col_map:
                distance_km = pd.to_numeric(self.prepared_data[col_map['distance']], errors='coerce').fillna(999).values
            else:
                distance_km = np.full(len(self.prepared_data), 0)  # Fallback: keine Distanz-Info
            
            # 1. Network = "None" → Score = 0
            mask_none = network_status == 'none'
            dh_score[mask_none] = 0.0
            
            # 3. Network ∈ {"Planned", "Existing"} - ZUERST definieren
            mask_network_exists = network_status.isin(['planned', 'existing'])
            
            # Berechne maximale Entfernung für dynamische Schwellenwerte
            max_distance = distance_km[mask_network_exists & (distance_km > 0)].max() if (mask_network_exists & (distance_km > 0)).any() else 10.0
            
            # 2. Network = "Study" → Score = 0.15, ABER nur wenn ≤ max_distance
            mask_study = network_status == 'study'
            mask_study_in_range = mask_study & (distance_km <= max_distance)
            mask_study_too_far = mask_study & (distance_km > max_distance)
            dh_score[mask_study_in_range] = 0.15
            dh_score[mask_study_too_far] = 0.0  # Studie, aber zu weit weg
            
            if mask_network_exists.any():
                # 3a. Connection ∈ {"Planned", "Existing"} → Score = 1.0
                mask_connected = mask_network_exists & connection_status.isin(['planned', 'existing'])
                dh_score[mask_connected] = 1.0
                
                # 3b. Connection = "None" → Score basierend auf Entfernung & Temperatur
                mask_not_connected = mask_network_exists & (connection_status == 'none')
                
                if mask_not_connected.any() and 'distance' in col_map and 'temperature' in col_map:
                    # Lade Temperatur (in °C) - Distanz wurde bereits oben geladen
                    temperature = pd.to_numeric(self.prepared_data[col_map['temperature']], errors='coerce').fillna(0).values
                    
                    # WICHTIG: Wenn Entfernung > max_distance → Score = 0.3 (Basis-Score)
                    mask_too_far = distance_km > max_distance
                    mask_in_range = mask_not_connected & ~mask_too_far
                    
                    # Setze Score = 0.3 für Anlagen die zu weit weg sind
                    dh_score[mask_not_connected & mask_too_far] = 0.3
                    
                    # Berechne Score nur für Anlagen in Reichweite (≤ max_distance)
                    if mask_in_range.any():
                        # Berechne S_dist: max(0, 1 - (distance_km / max_distance)²) - QUADRATISCH
                        # 0 km → 1.0, max_distance km → 0.0
                        s_dist = np.maximum(0, 1 - (distance_km / max_distance)**2)
                        
                        # Berechne S_temp: (T_max - T_net) / (T_max - T_min)
                        # Niedrigere Temperatur = besser
                        # Nur für Anlagen mit Netz (um T_min und T_max zu bestimmen)
                        temp_subset = temperature[mask_network_exists & (temperature > 0)]
                        if len(temp_subset) > 0 and temp_subset.max() > temp_subset.min():
                            t_min = temp_subset.min()
                            t_max = temp_subset.max()
                            s_temp_raw = (t_max - temperature) / (t_max - t_min)
                            s_temp = np.clip(s_temp_raw, 0, 1)
                            if not self._dh_logged:
                                self.log(f"      Temperatur-Range: {t_min:.1f}°C - {t_max:.1f}°C")
                        else:
                            # Fallback: wenn keine Temperaturvariation, setze auf 0.5
                            s_temp = np.full_like(temperature, 0.5, dtype=float)
                            if not self._dh_logged:
                                self.log(f"      Temperatur: keine Variation, verwende 0.5")
                        
                        # Gewichte für Distanz und Temperatur
                        w_dist = 0.6
                        w_temp = 0.4
                        
                        # S_raw = w_dist * S_dist + w_temp * S_temp
                        s_raw = w_dist * s_dist + w_temp * s_temp
                        
                        # KPI_DH_base = 0.3 + 0.6 * S_raw (Bereich 0.3 bis 0.9)
                        kpi_dh_base = 0.3 + 0.6 * s_raw
                        
                        # Setze Score für nicht verbundene Anlagen in Reichweite
                        dh_score[mask_in_range] = kpi_dh_base[mask_in_range]
                        
                        # Logging für Debugging (nur beim ersten Aufruf)
                        if not self._dh_logged:
                            num_too_far = (mask_not_connected & mask_too_far).sum()
                            num_in_range = mask_in_range.sum()
                            avg_dist = distance_km[mask_in_range].mean() if num_in_range > 0 else 0
                            avg_temp = temperature[mask_in_range & (temperature > 0)].mean() if (mask_in_range & (temperature > 0)).any() else 0
                            avg_score = dh_score[mask_in_range].mean() if num_in_range > 0 else 0
                            self.log(f"      Max. Entfernung (dynamisch): {max_distance:.2f} km")
                            self.log(f"      In Reichweite (≤{max_distance:.1f}km): {num_in_range}, Zu weit (>{max_distance:.1f}km): {num_too_far}")
                            if num_in_range > 0:
                                self.log(f"      Ø Distanz: {avg_dist:.2f} km, Ø Temp: {avg_temp:.1f}°C, Ø Score: {avg_score:.3f}")
                
                elif mask_not_connected.any():
                    # Fallback: wenn Entfernung/Temperatur fehlen, setze auf 0.6 (Mittelwert zwischen 0.3 und 0.9)
                    dh_score[mask_not_connected] = 0.6
            
            # Logging (nur beim ersten Aufruf)
            if not self._dh_logged:
                num_none = mask_none.sum()
                num_study_in_range = mask_study_in_range.sum()
                num_study_too_far = mask_study_too_far.sum()
                num_connected = mask_network_exists.sum() and (connection_status.isin(['planned', 'existing']) & mask_network_exists).sum()
                num_not_connected = mask_network_exists.sum() and ((connection_status == 'none') & mask_network_exists).sum()
                
                self.log(f"   District Heating Logik (wird für alle Iterationen angewendet):")
                self.log(f"      Network=None: {num_none} Anlagen → Score = 0.0")
                self.log(f"      Network=Study (≤{max_distance:.1f}km): {num_study_in_range} Anlagen → Score = 0.15")
                self.log(f"      Network=Study (>{max_distance:.1f}km): {num_study_too_far} Anlagen → Score = 0.0")
                self.log(f"      Network exists + Connected: {num_connected} Anlagen → Score = 1.0")
                self.log(f"      Network exists + Not connected (≤{max_distance:.1f}km): {num_in_range} Anlagen → Score = 0.3-0.9")
                self.log(f"      Network exists + Not connected (>{max_distance:.1f}km): {num_too_far} Anlagen → Score = 0.3")
                
                # Debug: Zeige Details für Anlagen mit Score = 0 aber Network exists
                mask_zero_with_network = (dh_score == 0.0) & mask_network_exists
                if mask_zero_with_network.any():
                    self.log(f"      DEBUG: {mask_zero_with_network.sum()} Anlagen mit Network exists aber Score = 0:")
                    for idx in np.where(mask_zero_with_network)[0][:5]:  # Zeige max. 5 Beispiele
                        name = self.prepared_data.iloc[idx].get('Name', 'Unknown')
                        net = network_status.iloc[idx]
                        conn = connection_status.iloc[idx]
                        dist = distance_km[idx]
                        self.log(f"         - {name}: Network={net}, Connection={conn}, Distance={dist:.2f}km")
                
                self._dh_logged = True
            
            return dh_score
            
        except Exception as e:
            self.log(f"   WARNUNG: District Heating Logik fehlgeschlagen: {e}")
            import traceback
            self.log(f"   {traceback.format_exc()}")
            return score
    

    def _calculate_top20_freq(self, all_rankings: np.ndarray) -> np.ndarray:
        """Berechnet Häufigkeit in Top-20"""
        return (all_rankings <= 20).mean(axis=0)
    
    def _create_ranked_database(self, results: Dict, variant: str = None):
        """
        Erstellt sortierte Datenbank mit NUR den Anlagen der aktuellen Variante (Built Flag = 1)
        """
        try:
            # Output-Pfad mit optionalem Run Name Suffix
            base_name = f"UWWTD_TP_Database_ranked_{variant}" if variant else "UWWTD_TP_Database_ranked"
            if self.mc_config.run_name:
                filename = f"{base_name}_{self.mc_config.run_name}.xlsx"
            else:
                filename = f"{base_name}.xlsx"
            
            output_path = Path(f"Output/MCA/Ranked/{filename}")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            original_path = Path(self.project_config.get('excel_path', 'Output/UWWTD_TP_Database.xlsx'))
            
            if not original_path.exists():
                self.log(f"   WARNUNG: Original-Datenbank nicht gefunden: {original_path}")
                return
            
            # Erstelle Rankings DataFrame
            available_id_cols = [col for col in self.id_columns if col in self.current_plant_data.columns]
            if not available_id_cols:
                result_df = pd.DataFrame({'Plant_Index': range(len(self.current_plant_data))})
                id_col = 'Plant_Index'
                self.log(f"   WARNUNG: Keine ID-Spalten gefunden, verwende Index")
            else:
                # Verwende nur die ERSTE ID-Spalte für das Calculated Values Sheet
                # (um Duplikate in der View zu vermeiden)
                id_col = available_id_cols[0]
                result_df = self.current_plant_data[[id_col]].copy()
                self.log(f"   Verwende ID-Spalte für Calculated Values: {id_col}")
            
            # Füge Scores und Statistiken hinzu
            result_df['Final Score'] = results['mean_final_score']
            result_df['Score CI Lower'] = results['score_ci_lower']
            result_df['Score CI Upper'] = results['score_ci_upper']
            result_df['Mean Rank'] = results['mean_rank']
            result_df['Rank CI Lower'] = results['rank_ci_lower']
            result_df['Rank CI Upper'] = results['rank_ci_upper']
            result_df['Rank Std Dev'] = results['std_rank']
            result_df['Rank CI Width'] = results['rank_ci_upper'] - results['rank_ci_lower']
            result_df['Top 20 Frequency'] = results['top20_frequency']
            
            # WICHTIG: Füge Kategorie-Scores VOR der Sortierung hinzu!
            mean_category_scores = results.get('mean_category_scores', {})
            if mean_category_scores:
                for cat_name, cat_scores in mean_category_scores.items():
                    result_df[f'{cat_name} Score'] = cat_scores
            
            # Entferne Risk-Distanz-Spalten (wir brauchen nur die Score-Spalten)
            risk_distance_cols = [
                'Distance to possible flood area [m]',
                'Distance to residential area [m]',
                'Distance to protected area [m]',
                'PA_Type'
            ]
            cols_to_drop = [col for col in risk_distance_cols if col in result_df.columns]
            if cols_to_drop:
                result_df = result_df.drop(columns=cols_to_drop)
                self.log(f"   Entferne Risk-Distanz-Spalten: {cols_to_drop}")
            
            # Sortiere nach Final Score (absteigend) - Zeile 1 = Rank 1, Zeile 2 = Rank 2, etc.
            result_df = result_df.sort_values('Final Score', ascending=False).reset_index(drop=True)
            
            # Füge Rank-Spalte hinzu (1 = beste Anlage)
            result_df['Rank'] = range(1, len(result_df) + 1)
            
            # Erstelle Rank-Mapping für Filterung
            rank_mapping = dict(zip(result_df[id_col], result_df['Rank']))
            
            self.log(f"   Erstelle Datenbank mit {len(rank_mapping)} Anlagen (nur Built Flag = 1)...")
            
            # Lade alle Sheets aus Original-Datenbank
            xls = pd.ExcelFile(original_path)
            
            # Erstelle Writer für Output
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            
            # Filtere und sortiere jedes Sheet
            for sheet_name in xls.sheet_names:
                self.log(f"      Filtere Sheet: {sheet_name}")
                
                # Lade Sheet
                df = pd.read_excel(original_path, sheet_name=sheet_name)
                
                # Finde ID-Spalte in diesem Sheet
                id_col_in_sheet = None
                if id_col in df.columns:
                    id_col_in_sheet = id_col
                else:
                    # Suche nach alternativen ID-Spalten
                    for col in df.columns:
                        col_lower = str(col).lower().strip().replace('_', '').replace(' ', '')
                        id_col_lower = str(id_col).lower().strip().replace('_', '').replace(' ', '')
                        if col_lower == id_col_lower:
                            id_col_in_sheet = col
                            break
                
                if id_col_in_sheet is None:
                    # Kein ID gefunden, überspringe Sheet
                    self.log(f"         Keine ID-Spalte gefunden, überspringe Sheet")
                    continue
                
                # Filtere nur Anlagen die in rank_mapping sind (Built Flag = 1)
                df_filtered = df[df[id_col_in_sheet].isin(rank_mapping.keys())].copy()
                
                if df_filtered.empty:
                    self.log(f"         Keine Anlagen nach Filterung, überspringe Sheet")
                    continue
                
                # Entferne duplizierte Spalten (die mit _dup enden)
                dup_cols = [col for col in df_filtered.columns if col.endswith('_dup')]
                if dup_cols:
                    df_filtered = df_filtered.drop(columns=dup_cols)
                    self.log(f"         Entferne duplizierte Spalten: {dup_cols}")
                
                # Füge Rank hinzu und sortiere
                df_filtered['_Rank'] = df_filtered[id_col_in_sheet].map(rank_mapping)
                df_filtered = df_filtered.sort_values('_Rank').reset_index(drop=True)
                df_filtered = df_filtered.drop(columns=['_Rank'])
                
                self.log(f"         {len(df_filtered)} Anlagen nach Filterung")
                
                # Schreibe gefiltertes und sortiertes Sheet
                df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Erstelle "Calculated Values" Sheet mit allen berechneten Werten
            self.log(f"      Erstelle Sheet: Calculated Values")
            
            # Kategorie-Scores wurden bereits vor der Sortierung hinzugefügt
            if mean_category_scores:
                self.log(f"         {len(mean_category_scores)} Kategorie-Scores enthalten (Durchschnitt über {self.mc_config.iterations} Iterationen)")
            
            # Schreibe Calculated Values Sheet
            result_df.to_excel(writer, sheet_name='Calculated Values', index=False)
            
            # Konfiguration Sheet mit Gewichtungen
            config_data = {
                'Parameter': [
                    'Run Name',
                    'Variant',
                    'Iterations',
                    'Confidence Level',
                    'Random Seed',
                    'Number of Plants (filtered)',
                    'Number of KPIs',
                    'Imputation',
                    'Winsorizing',
                    'Weight Sampling'
                ],
                'Value': [
                    self.mc_config.run_name if self.mc_config.run_name else 'Default',
                    variant if variant else 'All',
                    self.mc_config.iterations,
                    self.mc_config.confidence_level,
                    self.mc_config.random_seed if self.mc_config.random_seed else 'None',
                    len(self.current_plant_data),
                    len(self.kpi_structure),
                    f"{self.project_config.get('imputation_method', 'median')} (enabled: {self.project_config.get('imputation_enabled', True)})",
                    f"{self.project_config.get('winsorize_pct', 0)*100}%",
                    'Uniform distribution between w_min and w_max'
                ]
            }
            
            # Erstelle separates Sheet für Gewichtungen
            self._create_weights_sheet(writer)
            
            # Schreibe Configuration Sheet
            pd.DataFrame(config_data).to_excel(writer, sheet_name='Configuration', index=False)
            
            # Speichere Workbook
            writer.close()
            
            # Spaltenbreiten automatisch anpassen für alle Sheets
            self.log(f"   Passe Spaltenbreiten an...")
            wb = load_workbook(output_path)
            for sheet_name in wb.sheetnames:
                auto_col_width(str(output_path), sheet_name)
            
            self.log(f"   ✓ Gefilterte und sortierte Datenbank gespeichert: {output_path}")
            self.log(f"      - Nur {len(result_df)} Anlagen mit Built Flag = 1 für {variant}")
            if self.mc_config.run_name:
                self.log(f"      - Run Name: '{self.mc_config.run_name}'")
            self.log(f"      - Alle Sheets gefiltert und nach Ranking sortiert")
            self.log(f"      - Sheet 'Calculated Values' mit Final Score, Rankings und Kategorie-Scores")
            self.log(f"      - Sheet 'Configuration' mit Monte Carlo Parametern")
            self.log(f"      - Sheet 'Weights Configuration' mit allen Gewichtungs-Einstellungen")
            
        except Exception as e:
            self.log(f"   FEHLER beim Erstellen der Datenbank: {e}")
            import traceback
            self.log(f"   {traceback.format_exc()}")
    
    def _create_weights_sheet(self, writer):
        """Erstellt ein separates Sheet mit allen Gewichtungs-Konfigurationen"""
        try:
            weights_data = []
            
            # Kategorie-Gewichte
            categories_config = self.project_config.get('categories', {})
            weights_data.append(['=== CATEGORY WEIGHTS ===', '', '', ''])
            weights_data.append(['Category', 'w_min', 'w_max', 'Description'])
            
            for cat_name, cat_data in categories_config.items():
                if cat_name == "Energy":
                    continue  # Energy hat keine direkte Gewichtung
                
                w_min = cat_data.get('w_min', 'Not set')
                w_max = cat_data.get('w_max', 'Not set')
                description = f"Weight range for {cat_name} category"
                weights_data.append([cat_name, w_min, w_max, description])
            
            # Leerzeile
            weights_data.append(['', '', '', ''])
            
            # KPI-Gewichte nach Kategorien gruppiert
            weights_data.append(['=== KPI WEIGHTS ===', '', '', ''])
            weights_data.append(['KPI Name', 'w_min', 'w_max', 'Category'])
            
            # Gruppiere KPIs nach Kategorien
            kpis_by_category = {}
            for kpi in self.kpi_structure:
                cat_key = kpi['category']
                if cat_key not in kpis_by_category:
                    kpis_by_category[cat_key] = []
                kpis_by_category[cat_key].append(kpi)
            
            # Zähle KPIs pro Kategorie für Single-KPI-Kategorien
            kpis_per_category = {cat: len(kpis) for cat, kpis in kpis_by_category.items()}
            
            for cat_name in sorted(kpis_by_category.keys()):
                kpis = kpis_by_category[cat_name]
                is_single_kpi_category = kpis_per_category.get(cat_name, 0) == 1
                
                # Kategorie-Header
                weights_data.append([f'--- {cat_name} ---', '', '', ''])
                
                for kpi in kpis:
                    kpi_name = kpi['name']
                    
                    if is_single_kpi_category:
                        # Bei Single-KPI-Kategorien: Verwende Kategorie-Gewichte
                        cat_data = categories_config.get(cat_name, {})
                        w_min = cat_data.get('w_min', 'Not set')
                        w_max = cat_data.get('w_max', 'Not set')
                        note = f"{cat_name} (uses category weights)"
                    else:
                        # Multi-KPI-Kategorie: Verwende KPI-Gewichte
                        w_min = kpi.get('w_min', 'Not set')
                        w_max = kpi.get('w_max', 'Not set')
                        note = cat_name
                    
                    weights_data.append([kpi_name, w_min, w_max, note])
            
            # Leerzeile
            weights_data.append(['', '', '', ''])
            
            # Risk-Gewichte und Thresholds
            weights_data.append(['=== RISK CONFIGURATION ===', '', '', ''])
            weights_data.append(['Risk Type', 'Safe Threshold', 'Critical Threshold', 'Weight Range'])
            
            risk_thresholds = self.project_config.get('risk_thresholds', {})
            risk_weights = self.project_config.get('risk_weights', {})
            
            for risk_type in ['flood', 'residential', 'protected']:
                thresholds = risk_thresholds.get(risk_type, {})
                weights = risk_weights.get(risk_type, {})
                
                if risk_type == 'area':
                    safe_val = thresholds.get('sufficient', 'Not set')
                    crit_val = thresholds.get('critical', 'Not set')
                else:
                    safe_val = thresholds.get('safe', 'Not set')
                    crit_val = thresholds.get('critical', 'Not set')
                
                w_min = weights.get('w_min', 'Not set')
                w_max = weights.get('w_max', 'Not set')
                weight_info = f"w_min: {w_min}, w_max: {w_max}" if w_min != 'Not set' else 'Not set'
                
                weights_data.append([risk_type.title(), safe_val, crit_val, weight_info])
            
            # Erstelle DataFrame und schreibe ins Excel
            weights_df = pd.DataFrame(weights_data, columns=['Parameter', 'Min Value', 'Max Value', 'Notes'])
            weights_df.to_excel(writer, sheet_name='Weights Configuration', index=False)
            
            self.log(f"      - Sheet 'Weights Configuration' mit allen Gewichtungs-Einstellungen erstellt")
            
        except Exception as e:
            self.log(f"   WARNUNG: Konnte Weights Configuration Sheet nicht erstellen: {e}")
            import traceback
            self.log(f"   {traceback.format_exc()}")

    def _save_log(self):
        """Speichert Log-Datei"""
        if self.mc_config.run_name:
            log_filename = f"monte_carlo_log_{self.mc_config.run_name}.txt"
        else:
            log_filename = "monte_carlo_log.txt"
        
        log_path = Path(f"Output/MCA/{log_filename}")
        log_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.log_messages))
        
        self.log(f"\nLog gespeichert: {log_path}")


# Hilfsfunktion für Integration in GUI
def add_monte_carlo_button_to_main_gui():
    """
    Diese Funktion wird von 01_potential_grading.py importiert.
    Die Integration ist bereits in der GUI implementiert.
    """
    pass


if __name__ == "__main__":
    # Test der Monte Carlo Analyse
    root = tk.Tk()
    root.withdraw()
    
    config_path = Path("Output/MCA/Weights_MC/kpi_config.json")
    if config_path.exists():
        mc_window = MonteCarloWindow(root, str(config_path))
        root.mainloop()
    else:
        print(f"Konfigurationsdatei nicht gefunden: {config_path}")
        print("Bitte zuerst die GUI (01_potential_grading.py) ausführen und Konfiguration exportieren.")
