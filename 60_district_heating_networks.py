import os
from pathlib import Path
from typing import List, Tuple, Dict, Optional, Set

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# Pfade
BASE_DIR = Path(__file__).resolve().parent
FERNWAERME_PATH = BASE_DIR / "Daten" / "DistrictHeating_Database.xlsx"
UWWTD_PATH = BASE_DIR / "Output" / "UWWTD_TP_Database.xlsx"
OUTPUT_PATH = BASE_DIR / "Output" / "UWWTD_TP_Database.xlsx"

# Einstellungen
HEADER_ROW = 1
NEW_SHEET_TITLE_BASE = "District Heating"

# Spaltennamen
ID_COL_CANDIDATES = ["code", "anlage code", "anlagen code", "plant code", "id", "uwwtd code"]
NAME_COL_CANDIDATES = ["name der anlage", "anlagenname", "name", "plant name"]
DIST_COL_CANDIDATES = ["entfernung klasse", "entfernungsklasse", "distanz klasse", "distance class"]
SOURCE_COL_CANDIDATES = ["quellen", "quelle", "source", "sources"]


def norm(s: object) -> str:
    """Normalisiert Strings für Matching"""
    if s is None:
        return ""
    return " ".join(str(s).strip().lower().split())


def build_header_map(ws: Worksheet, header_row: int = 1) -> Tuple[Dict[str, int], List[str]]:
    """Baut Header-Map auf"""
    headers_raw: List[str] = []
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        v_str = "" if v is None else str(v)
        headers_raw.append(v_str)
        n = norm(v_str)
        if n and n not in header_map:
            header_map[n] = col
    return header_map, headers_raw


def find_col_by_candidates(header_map: Dict[str, int], candidates: List[str]) -> Optional[int]:
    """Findet Spalte anhand Kandidaten-Namen"""
    for cand in candidates:
        c = norm(cand)
        if c in header_map:
            return header_map[c]
    for key, col in header_map.items():
        for cand in candidates:
            if key.startswith(norm(cand)):
                return col
    return None


def unique_sheet_title(wb: Workbook, base_title: str) -> str:
    title = base_title[:31]
    if title not in wb.sheetnames:
        return title
    i = 2
    while True:
        suf = f" ({i})"
        t = (base_title[:31 - len(suf)]) + suf
        if t not in wb.sheetnames:
            return t
        i += 1


# Kompatibilität
def delete_rows_by_values(ws: Worksheet, header_row: int, col_idx: int, to_delete: Set[str]) -> int:
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        if str(v).strip() in to_delete:
            rows.append(r)
    for r in reversed(rows):
        ws.delete_rows(r, 1)
    return len(rows)


# Kompatibilität
def collect_uwwtd_inventory_and_order(wb_u: Workbook) -> Tuple[Set[str], Set[str], Dict[str, int], Dict[str, int]]:
    codes_set: Set[str] = set()
    names_set: Set[str] = set()
    code_order: Dict[str, int] = {}
    name_order: Dict[str, int] = {}

    ref_ws: Optional[Worksheet] = None
    if "General Data" in wb_u.sheetnames:
        ref_ws = wb_u["General Data"]
    else:
        for sn in wb_u.sheetnames:
            ws = wb_u[sn]
            header_map, _ = build_header_map(ws, header_row=HEADER_ROW)
            if find_col_by_candidates(header_map, ID_COL_CANDIDATES) or \
               find_col_by_candidates(header_map, NAME_COL_CANDIDATES):
                ref_ws = ws
                break

    if ref_ws is None:
        rank = 0
        for sn in wb_u.sheetnames:
            ws = wb_u[sn]
            header_map, _ = build_header_map(ws, header_row=HEADER_ROW)
            id_col = find_col_by_candidates(header_map, ID_COL_CANDIDATES)
            name_col = find_col_by_candidates(header_map, NAME_COL_CANDIDATES)
            if id_col is None and name_col is None:
                continue
            for r in range(HEADER_ROW + 1, ws.max_row + 1):
                cval = ws.cell(row=r, column=id_col).value if id_col else None
                nval = ws.cell(row=r, column=name_col).value if name_col else None
                code = str(cval).strip() if cval is not None else ""
                name = str(nval).strip() if nval is not None else ""
                if code:
                    if code not in codes_set:
                        codes_set.add(code)
                        code_order[code] = rank
                        rank += 1
                if name:
                    if name not in names_set:
                        names_set.add(name)
                        name_order[name] = rank
                        rank += 1
        return codes_set, names_set, code_order, name_order

    header_map, _ = build_header_map(ref_ws, header_row=HEADER_ROW)
    id_col = find_col_by_candidates(header_map, ID_COL_CANDIDATES)
    name_col = find_col_by_candidates(header_map, NAME_COL_CANDIDATES)

    rank = 0
    for r in range(HEADER_ROW + 1, ref_ws.max_row + 1):
        cval = ref_ws.cell(row=r, column=id_col).value if id_col else None
        nval = ref_ws.cell(row=r, column=name_col).value if name_col else None
        code = str(cval).strip() if cval is not None else ""
        name = str(nval).strip() if nval is not None else ""
        if code:
            if code not in codes_set:
                codes_set.add(code)
                code_order[code] = rank
                rank += 1
        if name:
            if name not in names_set:
                names_set.add(name)
                name_order[name] = rank
                rank += 1

    return codes_set, names_set, code_order, name_order


def main():
    # ======= Dateien prüfen =======
    if not FERNWAERME_PATH.exists():
        raise FileNotFoundError(f"Fernwärme-Datei nicht gefunden: {FERNWAERME_PATH}")
    if not UWWTD_PATH.exists():
        raise FileNotFoundError(f"UWWTD-Datei nicht gefunden: {UWWTD_PATH}")
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # ======= Fernwärme laden =======
    wb_fw = load_workbook(FERNWAERME_PATH, data_only=False, keep_links=True)
    ws_fw = wb_fw[wb_fw.sheetnames[0]]  # erstes Sheet

    header_map_fw, headers_fw_raw = build_header_map(ws_fw, header_row=HEADER_ROW)

    id_col_fw = find_col_by_candidates(header_map_fw, ID_COL_CANDIDATES)
    name_col_fw = find_col_by_candidates(header_map_fw, NAME_COL_CANDIDATES)
    # Distanzspalte optional
    dist_col_fw = find_col_by_candidates(header_map_fw, DIST_COL_CANDIDATES)
    src_col_fw = find_col_by_candidates(header_map_fw, SOURCE_COL_CANDIDATES)

    if id_col_fw is None and name_col_fw is None:
        raise RuntimeError("Konnte weder 'Code' noch 'Name der Anlage' in der Fernwärme-DB finden.")

    # ======= UWWTD laden um vorhandene Site_IDs zu ermitteln =======
    wb_u = load_workbook(UWWTD_PATH, data_only=False, keep_links=True)
    
    # Sammle alle vorhandenen Site_IDs aus "General Data"
    valid_site_ids: Set[str] = set()
    if "General Data" in wb_u.sheetnames:
        ws_general = wb_u["General Data"]
        header_map_general, _ = build_header_map(ws_general, header_row=HEADER_ROW)
        id_col_general = find_col_by_candidates(header_map_general, ID_COL_CANDIDATES)
        
        if id_col_general:
            for r in range(HEADER_ROW + 1, ws_general.max_row + 1):
                site_id_val = ws_general.cell(row=r, column=id_col_general).value
                if site_id_val is not None:
                    site_id = str(site_id_val).strip()
                    if site_id:
                        valid_site_ids.add(site_id)
            print(f"→ Gefundene Site_IDs in General Data: {len(valid_site_ids)}")
        else:
            print("⚠ WARNUNG: Keine ID-Spalte in General Data gefunden - kopiere alle Zeilen")
    else:
        print("⚠ WARNUNG: Sheet 'General Data' nicht gefunden - kopiere alle Zeilen")
    
    # Fernwärme-Zeilen filtern
    rows_to_keep: List[int] = []
    
    if valid_site_ids and id_col_fw:
        for r in range(HEADER_ROW + 1, ws_fw.max_row + 1):
            site_id_val = ws_fw.cell(row=r, column=id_col_fw).value
            if site_id_val is not None:
                site_id = str(site_id_val).strip()
                if site_id in valid_site_ids:
                    rows_to_keep.append(r)
        print(f"→ Zeilen mit passenden Site_IDs: {len(rows_to_keep)} von {ws_fw.max_row - HEADER_ROW}")
    else:
        # Fallback: alle Zeilen kopieren wenn keine Filterung möglich
        rows_to_keep = list(range(HEADER_ROW + 1, ws_fw.max_row + 1))
        print(f"→ Übernahme aller Zeilen (kein Filter möglich): {len(rows_to_keep)}")

    # Keine Löschung/Filterung

    # Neues Sheet mit allen Anlagen
    new_sheet_title = unique_sheet_title(wb_u, NEW_SHEET_TITLE_BASE)
    ws_new = wb_u.create_sheet(title=new_sheet_title)

    # Spaltenreihenfolge: Code, Name, dann alle weiteren Spalten
    ordered_src_cols: List[int] = []
    ordered_headers_new: List[str] = []

    code_header_text = "UWWTD Code"
    name_header_text = "Name"

    if id_col_fw:
        ordered_src_cols.append(id_col_fw)
        ordered_headers_new.append(code_header_text)
    if name_col_fw:
        ordered_src_cols.append(name_col_fw)
        ordered_headers_new.append(name_header_text)

    for idx, hdr in enumerate(headers_fw_raw, start=1):
        if idx in ordered_src_cols:
            continue
        ordered_src_cols.append(idx)
        ordered_headers_new.append(hdr)

    # Header schreiben
    for c, hdr in enumerate(ordered_headers_new, start=1):
        ws_new.cell(row=1, column=c, value=hdr)

    # Daten kopieren, Hyperlinks ab "Quellen" erhalten
    hyperlink_start_src_col = src_col_fw if src_col_fw else None

    for out_row_idx, src_row in enumerate(rows_to_keep, start=2):
        for out_col_idx, src_col in enumerate(ordered_src_cols, start=1):
            src_cell = ws_fw.cell(row=src_row, column=src_col)
            dst_cell = ws_new.cell(row=out_row_idx, column=out_col_idx)

            dst_cell.value = src_cell.value

            if hyperlink_start_src_col is not None and src_col >= hyperlink_start_src_col:
                if src_cell.hyperlink:
                    try:
                        dst_cell.hyperlink = src_cell.hyperlink.target
                    except Exception:
                        dst_cell.hyperlink = src_cell.hyperlink
                    try:
                        dst_cell.style = "Hyperlink"
                    except Exception:
                        pass

            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format

    # Auto-Breite
    for col_idx in range(1, ws_new.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for r in range(1, ws_new.max_row + 1):
            v = ws_new.cell(row=r, column=col_idx).value
            if v is None:
                continue
            v_str = str(v)
            max_len = max(max_len, len(v_str))
        ws_new.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    # Speichern
    wb_u.save(OUTPUT_PATH)
    print(f"✓ {len(rows_to_keep)} plants with district heating data → {OUTPUT_PATH.name}")



if __name__ == "__main__":
    main()
