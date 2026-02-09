import os
import sys
import re
import math
import time
import argparse
import pandas as pd
import geopandas as gpd
from shapely.geometry import Point, LineString, Polygon
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Konsolen-Encoding (für Windows)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# Pfade und Einstellungen

# Excel-Datei
EXCEL_DIR = os.path.join("Output", "UWWTD_TP_Database")
EXCEL_FALLBACK_FILE = os.path.join("Output", "UWWTD_TP_Database.xlsx")
EXCEL_SHEET_IN  = "General Data"
EXCEL_SHEET_OUT = "Grid Energy Connection"
CACHE_DIR = os.path.join('Output', 'Energy Network')
CACHE_CSV = os.path.join(CACHE_DIR, 'energy_network_cache.csv')

# Spaltennamen
CODE_COL = "UWWTD Code"
NAME_COL = "Name"
LAT_COL  = "Latitude"
LON_COL  = "Longitude"

# Overpass API Endpoints
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.openstreetmap.ru/api/interpreter",
]

# ArcGIS Fallback
FEATURESERVER_LINES = (
    "https://services-eu1.arcgis.com/"
    "zci5bUiJ8olAal7N/arcgis/rest/services/"
    "OpenStreetMap_Power_Lines_for_Europe/FeatureServer/0/query"
)

# Suchradius
RADIUS_KM = 20.0  # Reduziert von 30 auf 20 km

# Parallelisierung (konservativ)
WORKERS = 1  # Sequenziell, um Overpass nicht zu überlasten

# Leitungen: Primär Overpass, Fallback ArcGIS
LINES_SOURCE = "overpass"  # nicht ändern, wir fallbacken automatisch auf ArcGIS

# -------- Substation-Filter (NUR Anschluss-Punkte behalten) --------
MIN_CONNECT_KV = 110  # Mindestens 110 kV für Substations
GOOD_SUBSTATION_ROLES = {"transmission", "distribution", "switching"}
BAD_TOKENS = {
    "traction","rail","bahn","industrial","converter","hvdc","synchronous_condenser",
    "plant","generator","solar","pv","photovoltaic","wind","biogas","biomass","geothermal"
}
ANCHOR_TOKENS = {"transformer","switchgear","busbar","bay"}
CONNECT_REQUIRE_AREA = True  # Fläche oder Anker-Tags erforderlich

# -------- Leitungs-Filter --------
MIN_LINE_KV = 110  # Mindestens 110 kV für Leitungen

# -------- Ausgabe-Spalten --------
COL_TRASSE_DIST = "Distance to HV Power Line [km]"
COL_TRASSE_TYPE = "Power Network Type"
COL_LINE_KV     = "Nearest Line Voltage [kV]"

COL_SUB_DIST    = "Distance to HV Substation [km]"
COL_SUB_NAME    = "Nearest Substation Name"
COL_SUB_KV      = "Substation Voltage [kV]"
# -------- Tile-/Kachel-Cache für Overpass --------
TILE_DEG = 0.25                 # Kachelgröße in Grad (0.25° ~ 25–30 km)
CACHE_TTL_DAYS = 7              # Cache-Gültigkeit
CACHE_DIR = os.path.join(".cache", "osm_power")
os.makedirs(CACHE_DIR, exist_ok=True)

# --------------------------- HTTP Session ---------------------------

def get_session() -> requests.Session:
    sess = requests.Session()
    retries = Retry(
        total=8, connect=8, read=8, status=8,
        backoff_factor=2.0,  # Längere Wartezeiten zwischen Retries
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET", "POST"]),
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retries,
                          pool_connections=WORKERS*2 if 'WORKERS' in globals() else 16,
                          pool_maxsize=WORKERS*2 if 'WORKERS' in globals() else 16)
    sess.mount("https://", adapter)
    sess.mount("http://", adapter)
    return sess

SESSION = get_session()

# --------------------------- Excel-Helfer ---------------------------

def auto_col_width(path: str, sheet_name: str, min_w: float = 8.0, max_w: float = 80.0) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            s_len = max(len(part) for part in s.splitlines()) if "\n" in s else len(s)
            max_len = max(max_len, s_len)
        width = min(max_w, max(min_w, max_len + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width
    wb.save(path)
    wb.close()

# --------------------------- Progress (eine Zeile) ---------------------------

def _pb_chars():
    enc = (sys.stdout.encoding or "").upper()
    return ("█","─") if "UTF" in enc else ("#","-")

def progress_line(proc, total, ok_tr, ok_sub,
                  *, phase="", trasse=None, sub=None,
                  width=32, end=False):
    full, dash = _pb_chars()
    pct = 0.0 if total == 0 else proc / max(total, 1)
    filled = int(width * pct)
    bar = full * filled + dash * (width - filled)

    parts = []
    if phase: parts.append(phase)
    if trasse is not None: parts.append(f"Trasse:{'OK' if trasse else '--'}")
    if sub is not None:    parts.append(f"Sub:{'OK' if sub else '--'}")
    extra = ("  |  " + "  ".join(parts)) if parts else ""
    msg = f"[{bar}] {pct*100:5.1f}%  {proc}/{total}  OK_Trasse:{ok_tr}  OK_Sub:{ok_sub}{extra}"
    print("\r" + msg + " " * 8, end=("\n" if end else ""), flush=True)

# --------------------------- Overpass + Tile-Cache ---------------------------

def _tile_key(lat: float, lon: float):
    lat0 = math.floor(lat / TILE_DEG) * TILE_DEG
    lon0 = math.floor(lon / TILE_DEG) * TILE_DEG
    return f"{lat0:.2f}_{lon0:.2f}"

def _bbox_for_tile_with_buffer(lat: float, lon: float, radius_km: float):
    # Kachel um den Punkt
    lat0 = math.floor(lat / TILE_DEG) * TILE_DEG
    lat1 = lat0 + TILE_DEG
    lon0 = math.floor(lon / TILE_DEG) * TILE_DEG
    lon1 = lon0 + TILE_DEG
    # Puffer (grob) in Grad
    dlat = radius_km / 111.0
    clat = max(math.cos(math.radians(lat)), 0.2)
    dlon = radius_km / (111.0 * clat)
    south = lat0 - dlat
    north = lat1 + dlat
    west  = lon0 - dlon
    east  = lon1 + dlon
    return south, west, north, east, lat0, lon0

def _cache_path(qtype: str, lat0: float, lon0: float, radius_km: float):
    return os.path.join(CACHE_DIR, f"{qtype}_{lat0:.2f}_{lon0:.2f}_{int(radius_km)}.json")

def _cache_fresh(path: str) -> bool:
    if not os.path.exists(path):
        return False
    age_days = (time.time() - os.path.getmtime(path)) / 86400.0
    return age_days <= CACHE_TTL_DAYS

def _overpass_query(ql: str) -> dict:
    last_err = None
    headers = {"Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
    for url in OVERPASS_ENDPOINTS:
        try:
            r = SESSION.post(url, data=ql.encode("utf-8"), headers=headers, timeout=180)  # 180s statt 90s
            if r.status_code == 429:
                time.sleep(3.0)  # Längere Pause bei Rate Limiting
                continue
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_err = e
            time.sleep(2.0)  # Pause vor nächstem Mirror
            continue
    raise RuntimeError(f"Overpass request failed on all mirrors: {last_err}")

def _overpass_bbox_cached(qtype: str, lat: float, lon: float, radius_km: float, body_builder) -> dict:
    south, west, north, east, lat0, lon0 = _bbox_for_tile_with_buffer(lat, lon, radius_km)
    cpath = _cache_path(qtype, lat0, lon0, radius_km)
    if _cache_fresh(cpath):
        try:
            import json
            with open(cpath, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass  # wenn kaputt, neu laden

    ql = f"[out:json][timeout:180];({body_builder(south, west, north, east)});out body geom;"  # 180s timeout
    time.sleep(1.0)  # Pause vor jeder Overpass-Anfrage
    js = _overpass_query(ql)

    try:
        import json
        with open(cpath, "w", encoding="utf-8") as f:
            json.dump(js, f)
    except Exception:
        pass
    return js

def _elements_to_gdf(elements) -> gpd.GeoDataFrame:
    feats = []
    for el in elements:
        etype = el.get("type")
        tags  = el.get("tags", {}) or {}
        props = {**tags, "_osm_type": etype, "_osm_id": el.get("id")}
        geom = None

        if etype == "node":
            lon, lat = el.get("lon"), el.get("lat")
            if lon is not None and lat is not None:
                geom = Point(float(lon), float(lat))
        else:
            if "geometry" in el and el["geometry"]:
                coords = [(p["lon"], p["lat"]) for p in el["geometry"] if "lon" in p and "lat" in p]
                if len(coords) >= 2:
                    if len(coords) >= 4 and coords[0] == coords[-1]:
                        try:
                            geom = Polygon(coords)
                        except Exception:
                            geom = LineString(coords)
                    else:
                        try:
                            geom = LineString(coords)
                        except Exception:
                            geom = None
            if geom is None and "center" in el:
                c = el["center"]
                if "lon" in c and "lat" in c:
                    geom = Point(float(c["lon"]), float(c["lat"]))

        if geom is not None:
            feats.append({"type": "Feature", "properties": props, "geometry": geom.__geo_interface__})

    if not feats:
        return gpd.GeoDataFrame(geometry=[], crs=4326)
    return gpd.GeoDataFrame.from_features(feats, crs=4326)

# --------------------------- ArcGIS (Fallback für Leitungen) ---------------------------

def _arcgis_geom_to_geojson(arcgeom):
    if not arcgeom:
        return None
    if "x" in arcgeom and "y" in arcgeom:
        return {"type": "Point", "coordinates": [arcgeom["x"], arcgeom["y"]]}
    if "paths" in arcgeom:
        paths = arcgeom["paths"] or []
        if not paths:
            return None
        if len(paths) == 1:
            return {"type": "LineString", "coordinates": paths[0]}
        return {"type": "MultiLineString", "coordinates": paths}
    if "rings" in arcgeom:
        return {"type": "Polygon", "coordinates": arcgeom["rings"]}
    return None

def _query_feature_server_point(fs_url: str, lon: float, lat: float, radius_km: float, where: str) -> gpd.GeoDataFrame:
    base = {
        "where": where,
        "outFields": "*",
        "returnGeometry": "true",
        "geometry": f"{lon},{lat}",
        "geometryType": "esriGeometryPoint",
        "inSR": 4326, "outSR": 4326,
        "spatialRel": "esriSpatialRelIntersects",
        "distance": radius_km,
        "units": "esriSRUnit_Kilometer"
    }
    params = {"f": "geojson", **base}
    r = SESSION.get(fs_url, params=params, timeout=60)
    if r.ok:
        try:
            data = r.json()
            if isinstance(data, dict) and data.get("type") == "FeatureCollection":
                return gpd.GeoDataFrame.from_features(data, crs=4326)
        except ValueError:
            pass
    params = {"f": "json", **base}
    r = SESSION.get(fs_url, params=params, timeout=60)
    r.raise_for_status()
    data = r.json()
    feats = []
    for ftr in data.get("features", []):
        geom = _arcgis_geom_to_geojson(ftr.get("geometry"))
        if geom is None:
            continue
        feats.append({"type": "Feature", "properties": ftr.get("attributes", {}), "geometry": geom})
    if not feats:
        return gpd.GeoDataFrame(geometry=[], crs=4326)
    return gpd.GeoDataFrame.from_features(feats, crs=4326)

def _query_lines_arcgis_robust(lon: float, lat: float, radius_km: float) -> gpd.GeoDataFrame:
    try:
        gdf = _query_feature_server_point(FEATURESERVER_LINES, lon, lat, radius_km, "1=1")
        if not gdf.empty: return gdf
    except Exception:
        pass
    for delay in (0.25, 0.75):
        try:
            time.sleep(delay)
            gdf = _query_feature_server_point(FEATURESERVER_LINES, lon, lat, radius_km, "1=1")
            if not gdf.empty: return gdf
        except Exception:
            continue
    for rad in (45.0, 60.0):
        try:
            gdf = _query_feature_server_point(FEATURESERVER_LINES, lon, lat, rad, "1=1")
            if not gdf.empty: return gdf
        except Exception:
            continue
    return gpd.GeoDataFrame(geometry=[], crs=4326)

# --------------------------- Voltage / Extraktion ---------------------------

_VOL_NUM_RE = re.compile(r'(\d{2,6})', re.I)
_VOL_FROM_TAGS_RE = re.compile(
    r'voltage(?::primary|:secondary)?\s*[:=]\s*["\']?(\d{2,6})',
    re.I
)

def _kv_from_voltage_field(s: str):
    if not s:
        return None
    nums = []
    for token in re.split(r'[;,\|]', str(s)):
        m = _VOL_NUM_RE.search(token)
        if not m: continue
        n = int(m.group(1))
        nums.append(n // 1000 if n >= 2000 else n)
    return max(nums) if nums else None

def _kv_from_tags_blob(s: str):
    if not s:
        return None
    nums = [int(x) for x in _VOL_FROM_TAGS_RE.findall(s)]
    kvs = [(n // 1000 if n >= 2000 else n) for n in nums]
    return max(kvs) if kvs else None

def _extract_max_voltage_kv(val):
    if val is None:
        return None
    s = str(val)
    nums, cur = [], ""
    for ch in s:
        if ch.isdigit(): cur += ch
        else:
            if cur:
                nums.append(int(cur)); cur = ""
    if cur: nums.append(int(cur))
    if not nums: return None
    kvs = [(n // 1000 if n >= 2000 else n) for n in nums]
    return max(kvs) if kvs else None

# --------------------------- Leitungen: Overpass (mit Filter) + Fallback ---------------------------

def query_power_lines_overpass_cached(lon: float, lat: float, radius_km: float = RADIUS_KM) -> gpd.GeoDataFrame:
    """
    Holt Leitungen aus OSM via Overpass im Kachel-BBox-Modus:
      power = line | minor_line | cable   UND   "voltage" vorhanden
    """
    def body(s,w,n,e):
        return f'nwr["power"~"^(line|minor_line|cable)$"]["voltage"]({s},{w},{n},{e});'
    js = _overpass_bbox_cached("lines", lat, lon, radius_km, body)
    gdf = _elements_to_gdf(js.get("elements", []))
    if gdf.empty:
        return gdf
    # Filter: voltage >= 110 kV
    if "voltage" in gdf.columns:
        mask = gdf["voltage"].apply(_kv_from_voltage_field).fillna(0).astype(int) >= MIN_LINE_KV
    else:
        mask = pd.Series(False, index=gdf.index)
    gdf = gdf[mask]
    return gdf

def query_power_near_point(lon: float, lat: float) -> gpd.GeoDataFrame:
    """
    Primär Overpass (gefiltert), falls leer -> Fallback ArcGIS.
    """
    g_over = query_power_lines_overpass_cached(lon, lat, RADIUS_KM)
    if not g_over.empty:
        return g_over
    g_arc = _query_lines_arcgis_robust(lon, lat, RADIUS_KM)
    return g_arc

# --------------------------- Substations via Overpass (Strikt anschlussfähig) ---------------------------

def _is_substation_like(row: pd.Series) -> bool:
    return str(row.get("power", "")).lower() == "substation" or "substation" in " ".join(str(v).lower() for v in row.dropna().values)

def _has_bad_tokens(row: pd.Series) -> bool:
    blob = " ".join(str(v).lower() for v in row.dropna().values)
    return any(b in blob for b in BAD_TOKENS)

def _role_ok(row: pd.Series) -> bool:
    for key in ("substation","tags_substation"):
        if key in row.index and pd.notna(row.get(key)):
            val = str(row.get(key)).lower()
            if any(r in val for r in GOOD_SUBSTATION_ROLES):
                return True
    for key in ("tags","other_tags"):
        if key in row.index and pd.notna(row.get(key)):
            val = str(row.get(key)).lower()
            if any(f"substation={r}" in val or f"substation:{r}" in val for r in GOOD_SUBSTATION_ROLES):
                return True
    return False

def _substation_voltage_kv(row: pd.Series):
    for c in ("voltage", "voltage:primary", "voltage:secondary", "tags_voltage"):
        if c in row.index and pd.notna(row.get(c)):
            kv = _kv_from_voltage_field(str(row.get(c)))
            if kv:
                return kv
    for c in ("tags", "other_tags"):
        if c in row.index and pd.notna(row.get(c)):
            kv = _kv_from_tags_blob(str(row.get(c)))
            if kv:
                return kv
    return None

def _has_anchor_tokens(row: pd.Series) -> bool:
    blob = " ".join(str(v).lower() for v in row.dropna().values)
    return any(tok in blob for tok in ANCHOR_TOKENS)

def _filter_substations_connection_only(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    if gdf.empty:
        return gdf
    keep = []
    for idx, r in gdf.iterrows():
        if not _is_substation_like(r):
            continue
        if _has_bad_tokens(r):
            continue
        if not _role_ok(r):
            continue
        kv = _substation_voltage_kv(r)
        if kv is None or kv < MIN_CONNECT_KV:
            continue
        geom = r.geometry
        is_area = getattr(geom, "geom_type", "") in ("Polygon","MultiPolygon")
        if CONNECT_REQUIRE_AREA:
            if not (is_area or _has_anchor_tokens(r)):
                continue
        else:
            if not (is_area or _has_anchor_tokens(r)):
                continue
        keep.append(idx)
    return gdf.loc[keep] if keep else gpd.GeoDataFrame(geometry=[], crs=gdf.crs)

def query_substations_overpass_cached(lon: float, lat: float, radius_km: float = RADIUS_KM) -> gpd.GeoDataFrame:
    def body(s,w,n,e):
        return f'nwr["power"="substation"]({s},{w},{n},{e});'
    js = _overpass_bbox_cached("subs", lat, lon, radius_km, body)
    gdf = _elements_to_gdf(js.get("elements", []))
    return _filter_substations_connection_only(gdf)

def nearest_substation_distance_and_voltage(pt_wgs84: Point):
    subs = query_substations_overpass_cached(pt_wgs84.x, pt_wgs84.y, RADIUS_KM)
    if subs.empty:
        for rad in (60.0, 120.0):
            subs = query_substations_overpass_cached(pt_wgs84.x, pt_wgs84.y, rad)
            if not subs.empty:
                break
    if subs.empty:
        return None, None, None

    pt_3035 = gpd.GeoSeries([pt_wgs84], crs=4326).to_crs(3035).iloc[0]
    subs_3035 = subs.to_crs(3035)
    is_poly = subs_3035.geom_type.isin(["Polygon","MultiPolygon"])
    dists = subs_3035.geometry.boundary.distance(pt_3035) if is_poly.any() else subs_3035.geometry.distance(pt_3035)
    idxmin = dists.idxmin()
    dist_km = float(dists.loc[idxmin]) / 1000.0

    name = None
    for c in ("name","ref","station","_osm_id"):
        if c in subs.columns and pd.notna(subs.at[idxmin, c]):
            name = str(subs.at[idxmin, c]); break

    sub_kv = _substation_voltage_kv(subs.loc[idxmin])
    return round(dist_km, 3), (name or "substation"), sub_kv

# --------------------------- Leitungen: Distanz & Spannung ---------------------------

def nearest_distance_and_type(pt_wgs84: Point):
    lines = query_power_near_point(pt_wgs84.x, pt_wgs84.y)
    if lines.empty:
        return None, None
    power_col = next((c for c in ["power", "tags_power", "type", "fclass"] if c in lines.columns), None)
    pt_3035 = gpd.GeoSeries([pt_wgs84], crs=4326).to_crs(3035).iloc[0]
    lines_3035 = lines.to_crs(3035)
    dists = lines_3035.geometry.distance(pt_3035)
    idxmin = dists.idxmin()
    dmin_m = float(dists.loc[idxmin])
    ptype = None
    if power_col is not None:
        val = lines_3035.loc[idxmin, power_col]
        if pd.notna(val):
            ptype = str(val)
    return round(dmin_m / 1000.0, 3), ptype

def nearest_line_voltage_kv(pt_wgs84: Point):
    lines = query_power_near_point(pt_wgs84.x, pt_wgs84.y)
    if lines.empty:
        return None
    pt_3035 = gpd.GeoSeries([pt_wgs84], crs=4326).to_crs(3035).iloc[0]
    lines_3035 = lines.to_crs(3035)
    idxmin = lines_3035.geometry.distance(pt_3035).idxmin()
    for c in ("voltage", "tags_voltage", "voltages", "other_voltage", "other_tags"):
        if c in lines.columns and pd.notna(lines.at[idxmin, c]):
            return _extract_max_voltage_kv(lines.at[idxmin, c])
    return None

# --------------------------- Excel laden (ohne GPKG) ---------------------------

def _resolve_excel_path() -> str:
    """Sucht eine .xlsx im Ordner EXCEL_DIR. Fällt zurück auf EXCEL_FALLBACK_FILE."""
    # Wenn EXCEL_DIR ein existierendes File ist, direkt nehmen
    if os.path.isfile(EXCEL_DIR):
        return EXCEL_DIR
    # Falls Ordner existiert, die *erste* .xlsx nehmen (deterministisch sortiert)
    if os.path.isdir(EXCEL_DIR):
        xs = [os.path.join(EXCEL_DIR, p) for p in os.listdir(EXCEL_DIR) if p.lower().endswith('.xlsx')]
        xs.sort()
        if xs:
            return xs[0]
    # Fallback auf eine bekannte Datei
    return EXCEL_FALLBACK_FILE

def load_points_from_excel() -> gpd.GeoDataFrame:
    path = _resolve_excel_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel nicht gefunden: {path}")
    df = pd.read_excel(path, sheet_name=EXCEL_SHEET_IN, engine="openpyxl")
    df.rename(columns=lambda c: str(c).strip(), inplace=True)
    required = [CODE_COL, NAME_COL, LAT_COL, LON_COL]
    miss = [c for c in required if c not in df.columns]
    if miss:
        raise ValueError(f"Im Sheet '{EXCEL_SHEET_IN}' fehlen Spalten: {miss}")
    df[LAT_COL] = pd.to_numeric(df[LAT_COL], errors="coerce")
    df[LON_COL] = pd.to_numeric(df[LON_COL], errors="coerce")
    df = df.dropna(subset=[LAT_COL, LON_COL]).copy()
    gdf = gpd.GeoDataFrame(
        df,
        geometry=[Point(xy) for xy in zip(df[LON_COL], df[LAT_COL])],
        crs=4326
    )
    return gdf

# --------------------------- Punktberechnung ---------------------------

def _compute_point_row(code, name, pt):
    # Trasse
    tr_ok = False
    try:
        dist_km, ptype = nearest_distance_and_type(pt)
        line_kv  = nearest_line_voltage_kv(pt)
        tr_ok = (dist_km is not None)
    except Exception:
        dist_km, ptype, line_kv = None, None, None

    # Substation (strikt anschlussfähig)
    sub_ok = False
    try:
        dist_sub_km, sub_name, sub_kv = nearest_substation_distance_and_voltage(pt)
        sub_ok = (dist_sub_km is not None)
    except Exception:
        dist_sub_km, sub_name, sub_kv = None, None, None

    row = {
        CODE_COL: code,
        NAME_COL: name,

        # Trasse
        COL_TRASSE_DIST: dist_km,
        COL_TRASSE_TYPE: ptype,
        COL_LINE_KV:     line_kv,

        # Substation
        COL_SUB_DIST:  dist_sub_km,
        COL_SUB_NAME:  sub_name,
        COL_SUB_KV:    sub_kv,
            }
    return row, tr_ok, sub_ok

# --------------------------- Hauptlauf: berechnen & schreiben ---------------------------

def _write_energy_sheet_from_df(df: pd.DataFrame, excel_out_path: str) -> None:
    """Write DF into Excel sheet 'Grid Energy Connection' and autosize columns."""
    os.makedirs(os.path.dirname(excel_out_path), exist_ok=True)
    with pd.ExcelWriter(excel_out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        df.to_excel(xw, sheet_name=EXCEL_SHEET_OUT, index=False)
    auto_col_width(excel_out_path, EXCEL_SHEET_OUT)

def compute_and_write_energy_sheet(gdf_points_wgs84: gpd.GeoDataFrame, excel_out_path: str) -> pd.DataFrame:
    total = len(gdf_points_wgs84)
    processed = 0
    ok_tr = ok_sub = 0

    progress_line(processed, total, ok_tr, ok_sub, phase="Init")
    rows = []

    if WORKERS <= 1:
        # Sequenziell mit Phasen-Updates pro Anlage
        for _, r in gdf_points_wgs84.iterrows():
            code = r.get(CODE_COL); name = r.get(NAME_COL); pt = r.geometry

            progress_line(processed, total, ok_tr, ok_sub, phase=f"{code or ''} Trasse", trasse=False)
            item, tr_ok, sub_ok = _compute_point_row(code, name, pt)
            progress_line(processed, total, ok_tr, ok_sub,
                          phase=f"{code or ''} Sub", trasse=tr_ok, sub=sub_ok)

            rows.append(item)
            ok_tr  += int(bool(tr_ok))
            ok_sub += int(bool(sub_ok))
            processed += 1

            progress_line(processed, total, ok_tr, ok_sub, phase="Weiter …", trasse=tr_ok, sub=sub_ok)
    else:
        # Parallel: konservativ; nur Gesamt-Fortschritt
        lock = threading.Lock()
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            futs = []
            for _, r in gdf_points_wgs84.iterrows():
                futs.append(ex.submit(_compute_point_row, r.get(CODE_COL), r.get(NAME_COL), r.geometry))
            for fut in as_completed(futs):
                item, tr_ok, sub_ok = fut.result()
                rows.append(item)
                with lock:
                    processed += 1
                    ok_tr  += int(bool(tr_ok))
                    ok_sub += int(bool(sub_ok))
                    progress_line(processed, total, ok_tr, ok_sub, phase=f"{item.get(CODE_COL) or ''}")

    progress_line(processed, total, ok_tr, ok_sub, phase="Fertig", end=True)

    df = pd.DataFrame(rows)

    # Schreiben: Sheet 'Grid Energy Connection' ersetzen
    os.makedirs(os.path.dirname(excel_out_path), exist_ok=True)
    with pd.ExcelWriter(excel_out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        df.to_excel(xw, sheet_name=EXCEL_SHEET_OUT, index=False)
    auto_col_width(excel_out_path, EXCEL_SHEET_OUT)
    return df

# --------------------------- main() ---------------------------


def main():
    ap = argparse.ArgumentParser(description="Netz-KPIs (ohne Preise) -> Sheet 'Grid Energy Connection'")
    args = ap.parse_args()

    excel_in = _resolve_excel_path()
    print(f">> Lese Koordinaten aus Excel → {excel_in} | Sheet='{EXCEL_SHEET_IN}'")

    if os.path.exists(CACHE_CSV):
        print(f">> Cache gefunden → {CACHE_CSV}. Schreibe nur in Excel (kein Neuberechnen)…")
        df_out = pd.read_csv(CACHE_CSV)
        excel_out = excel_in
        _write_energy_sheet_from_df(df_out, excel_out)
    else:
        gdf = load_points_from_excel()
        print(f"   {len(gdf)} Punkte geladen.")
        if WORKERS <= 1:
            print(f">> Berechne Trasse/Sub im {RADIUS_KM:.0f}-km-Radius je Punkt … (sequenziell)")
        else:
            print(f">> Berechne Trasse/Sub … (parallel: WORKERS={WORKERS})")
        excel_out = excel_in
        df_out = compute_and_write_energy_sheet(gdf, excel_out)
        
        # Cache speichern
        os.makedirs(CACHE_DIR, exist_ok=True)
        df_out.to_csv(CACHE_CSV, index=False)
        print(f">> Cache gespeichert → {CACHE_CSV}")

    print(f"✓ Grid Energy Connection data saved → Sheet '{EXCEL_SHEET_OUT}'")


if __name__ == "__main__":
    main()


