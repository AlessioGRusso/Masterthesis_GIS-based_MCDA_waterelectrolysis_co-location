from pathlib import Path
import re
import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
import sys
import warnings

# Warnungen unterdrücken - die sind nur nervig
warnings.filterwarnings(
    "ignore",
    ".*DataFrame concatenation with empty or all-NA entries is deprecated.*",
    category=FutureWarning
)

# Wichtige Einstellungen
THRESHOLD = 300000          # Filter-Grenze: nur Anlagen über 300k PE
LAYER = None                # None = nimmt das erste Layer aus der Datei

# Spalten, in der richtigen Reihenfolge
SELECT_COLS = [
    "uwwCode", "uwwName", "uwwLatitude", "uwwLongitude",
    "uwwCapacity", "uwwOzonation", "uwwWasteWaterTreated", "uwwBODIncomingMeasured",
    "uwwCODIncomingMeasured", "uwwNIncomingMeasured", "uwwBODDischargeMeasured",
    "uwwCODDischargeMeasured", "uwwNDischargeMeasured"
]

# Pfade zu den Dateien
BASE = Path(__file__).resolve().parent
GPKG = BASE / "Daten" / "UWWTD_TreatmentPlants2024.gpkg"
OUT  = BASE / "Output" / "UWWTD_TP_Database_EU.xlsx"

# Zusätzliche Datenbanken
GPKG_1 = BASE / "Daten" / "UWWTD_TreatmentPlants2018.gpkg"
OUT_1  = BASE / "Output" / "UWWTD_TP_Database_EU_2019.xlsx"   
OUT_EUNISC = BASE / "Output" / "UWWTD_TP_Database_EUNISC.xlsx"

# HydroWASTE Datenbank
HYDROWASTE_CSV = BASE / "Daten" / "HydroWASTE_v10.csv"
OUT_HYDROWASTE = BASE / "Output" / "UWWTD_TP_Database_HydroWASTE.xlsx"

# Prüfen ob die Hauptdatei da ist
if not GPKG.exists():
    raise SystemExit(f"❌ Datei nicht gefunden: {GPKG}")

def parse_number(val):
    """Parst Zahlen aus verschiedenen Formaten - europäische Zahlen sind manchmal tricky"""
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()
    if s == "":
        return pd.NA
    s = re.sub(r"[ \u00A0\u202F']", "", s)  # Verschiedene Leerzeichen und Apostrophe entfernen
    has_dot, has_com = "." in s, "," in s
    if has_dot and has_com:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_com:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return pd.NA

# robuste Auto-Breiten-Funktion ----------
def _text_len(val) -> int:
    s = str(val)
    return max(len(part) for part in s.splitlines()) if "\n" in s else len(s)

def autosize_excel(path: Path, min_w: float = 8.0, max_w: float = 80.0,
                   only_sheets: set[str] | None = None, exclude: set[str] | None = None) -> None:
    """
    Passt Spaltenbreite in allen (oder ausgewählten) Sheets an den längsten Eintrag an.
    """
    wb = load_workbook(path)
    targets = []
    if only_sheets:
        for name in only_sheets:
            if name in wb.sheetnames:
                targets.append(wb[name])
    else:
        targets = list(wb.worksheets)

    if exclude:
        targets = [ws for ws in targets if ws.title not in exclude]

    for ws in targets:
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
            max_len = 0
            for c in col:
                if c.value is not None:
                    max_len = max(max_len, _text_len(c.value))
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, max_len + 2))
    wb.save(path)
    wb.close()
# --------------------------------------------------------

# 1) GPKG lesen
gdf = gpd.read_file(GPKG, layer=LAYER)
df = gdf.drop(columns="geometry", errors="ignore").copy()

# 2) Kapazität parsen & filtern
if "uwwCapacity" not in df.columns:
    raise SystemExit(f"❌ Column 'uwwCapacity' missing. Columns: {list(df.columns)}")

df["uwwCapacity_num"] = df["uwwCapacity"].map(parse_number)
filtered = df[df["uwwCapacity_num"] > THRESHOLD].drop(columns=["uwwCapacity_num"])

# 3) Nur gewünschte Spalten behalten
existing = [c for c in SELECT_COLS if c in filtered.columns]
missing  = [c for c in SELECT_COLS if c not in filtered.columns]
if not existing:
    raise SystemExit("❌ None of the required columns found. Check SELECT_COLS.")
final_df = filtered[existing].copy()

# 4) Speichern
OUT.parent.mkdir(parents=True, exist_ok=True)

# === Spalten umbenennen  ===
RENAME_COLS = {
    "uwwLatitude": "Latitude",
    "uwwLongitude": "Longitude",
    "uwwCapacity": "Capacity/PE",
    "uwwName": "Name",
    "uwwOzonation": "Ozonation",
    "uwwWasteWaterTreated": "Waste Water Treated [m³/a]",
    "uwwCode": "UWWTD Code",
    "uwwBODIncomingMeasured": "BODIncomingMeasured [t/a]",
    "uwwCODIncomingMeasured": "CODIncomingMeasured [t/a]",
    "uwwNIncomingMeasured": "NIncomingMeasured [t/a]",
    "uwwPIncomingMeasured": "PIncomingMeasured [t/a]",
    "uwwBODDischargeMeasured": "BODDischargeMeasured [t/a]",
    "uwwCODDischargeMeasured": "CODDischargeMeasured [t/a]",
    "uwwNDischargeMeasured": "NDischargeMeasured [t/a]",
    "uwwPDischargeMeasured": "PDischargeMeasured [t/a]"

}
final_df = final_df.rename(columns=RENAME_COLS)
final_df.to_excel(OUT, index=False)
autosize_excel(OUT)  # <<— fix: Auto-Breite

print(f"✓ {len(final_df)} plants saved → {OUT.name}")

# ===== Zusätzliche Verarbeitung: zweite DB + gemeinsame EUNISC =====
# --- zweite GPKG lesen und wie gehabt verarbeiten ---
if not GPKG_1.exists():
    raise SystemExit(f"❌ File not found: {GPKG_1}")

gdf2 = gpd.read_file(GPKG_1, layer=LAYER)
df2 = gdf2.drop(columns="geometry", errors="ignore").copy()

if "uwwCapacity" not in df2.columns:
    raise SystemExit(f"❌ Column 'uwwCapacity' missing in {GPKG_1}. Columns: {list(df2.columns)}")

df2["uwwCapacity_num"] = df2["uwwCapacity"].map(parse_number)
filtered2 = df2[df2["uwwCapacity_num"] > THRESHOLD].drop(columns=["uwwCapacity_num"])

existing2 = [c for c in SELECT_COLS if c in filtered2.columns]
if not existing2:
    raise SystemExit("❌ (File 2) None of the required columns found. Check SELECT_COLS.")
final_df2 = filtered2[existing2].copy()

# Speichern zweite Einzel-Excel — als *_EU_2019.xlsx
OUT_1.parent.mkdir(parents=True, exist_ok=True)
final_df2_renamed = final_df2.rename(columns=RENAME_COLS)
final_df2_renamed.to_excel(OUT_1, index=False)
autosize_excel(OUT_1)

print(f"✓ {len(final_df2_renamed)} plants saved → {OUT_1.name}")

# --- Gemeinsame Datei: (alte) EU-Daten + (UKNI|UKSC) aus Datei 2 ---
# final_df ist bereits umbenannt (nach erster Speicherung)
eu_df = final_df.copy()

mask_uk = final_df2.get("uwwCode", pd.Series([], dtype=str)).astype(str).str.startswith(("UKNI", "UKSC"), na=False)
ukni_sc = final_df2.loc[mask_uk].rename(columns=RENAME_COLS).copy()

# Spalten wie in EU-Datei ausrichten
ukni_sc = ukni_sc.reindex(columns=eu_df.columns, fill_value=pd.NA)

# Minimaler Fix gegen FutureWarning: concat nur, wenn nicht leer
if ukni_sc.empty:
    combined = eu_df.copy()
else:
    combined = pd.concat([eu_df, ukni_sc], ignore_index=True)

combined.to_excel(OUT_EUNISC, index=False)
autosize_excel(OUT_EUNISC)
print(f"✓ {len(combined)} plants saved → {OUT_EUNISC.name}")

# ===== HydroWASTE Verarbeitung =====
# Zielländer: Schweiz, Albanien, Serbien, Montenegro, Bosnien und Herzegowina, Kosovo, Nordmazedonien
TARGET_COUNTRIES = ['CHE', 'ALB', 'SRB', 'MNE', 'BIH', 'XKX', 'MKD']

if HYDROWASTE_CSV.exists():
    hw_df = pd.read_csv(HYDROWASTE_CSV, encoding='latin1')
    
    # Filtern nach Zielländern
    hw_filtered = hw_df[hw_df['CNTRY_ISO'].isin(TARGET_COUNTRIES)].copy()
    
    # DESIGN_CAP numerisch parsen, falls leer POP_SERVED als Fallback nutzen
    hw_filtered['DESIGN_CAP_num'] = pd.to_numeric(hw_filtered['DESIGN_CAP'], errors='coerce')
    hw_filtered['POP_SERVED_num'] = pd.to_numeric(hw_filtered['POP_SERVED'], errors='coerce')
    
    # Capacity: DESIGN_CAP wenn vorhanden, sonst POP_SERVED
    hw_filtered['Capacity'] = hw_filtered['DESIGN_CAP_num'].fillna(hw_filtered['POP_SERVED_num'])
    
    # Filtern nach Kapazität > 300.000
    hw_filtered = hw_filtered[hw_filtered['Capacity'] > THRESHOLD].copy()
    
    # AIRE-Anlagen ausschließen (Duplikate)
    hw_filtered = hw_filtered[~hw_filtered['WWTP_NAME'].astype(str).str.contains('AIRE', case=False, na=False)].copy()
    
    # Spalten mappen auf  Schema
    hw_mapped = pd.DataFrame({
        'UWWTD Code': 'HW_' + hw_filtered['WASTE_ID'].astype(str),
        'Name': hw_filtered['WWTP_NAME'],
        'Latitude': hw_filtered['LAT_WWTP'],
        'Longitude': hw_filtered['LON_WWTP'],
        'Capacity/PE': hw_filtered['Capacity'],
        'Ozonation': pd.NA,  # nicht verfügbar in HydroWASTE
    })
    
    # Speichern als separate Datei
    hw_mapped.to_excel(OUT_HYDROWASTE, index=False)
    autosize_excel(OUT_HYDROWASTE)
    print(f"✓ {len(hw_mapped)} plants (HydroWASTE) saved → {OUT_HYDROWASTE.name}")
else:
    print(f"⚠ HydroWASTE file not found: {HYDROWASTE_CSV}")
    hw_mapped = pd.DataFrame()  # leerer DataFrame als Fallback

# ===== Einstellungen =====
BASE = Path(__file__).resolve().parent

IN_ODS   = BASE / "Daten" / "UWWTR_Art15_24thOct2024.ods"     # Pfad zur ODS-Datei
SHEET    = "T_UWWTPS"                                   # Tabellensheet
CAP_COL  = "uwwCapacity"
MIN_CAP  = 300_000
KEEP_COLS = [
    "uwwCode", "uwwName",
    "uwwLatitude", "uwwLongitude", CAP_COL, "uwwOzonation"
]

OUT_XLSX = BASE / "Output" / "UWWTD_TP_Database_EN.xlsx"
# ==========================

def main():
    # 1) ODS laden
    df = pd.read_excel(IN_ODS, sheet_name=SHEET, dtype=str, engine="odf")

    # 2) uwwCapacity numerisch + filtern
    df[CAP_COL] = pd.to_numeric(
        df[CAP_COL].astype(str).str.replace(",", ".", regex=False),
        errors="coerce"
    )
    filtered = df[df[CAP_COL] > MIN_CAP].copy()

    # 3) nur gewünschte Spalten
    keep = [c for c in KEEP_COLS if c in filtered.columns]
    result = filtered[keep]

    # === Spalten umbenennen ===
    RENAME_COLS = {
        "uwwLatitude": "Latitude",
        "uwwLongitude": "Longitude",
        "uwwCapacity": "Capacity/PE",
        "uwwName" : "Name",
        "uwwOzonation": "Ozonation",
        "uwwCode" : "UWWTD Code"
    }
    result = result.rename(columns=RENAME_COLS)
    # ================================

    # 4) nach Excel speichern
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    result.to_excel(OUT_XLSX, index=False)
    autosize_excel(OUT_XLSX)



if __name__ == "__main__":
    main()

# --------- zweiter Teil: finale Master-Datenbank bauen & formatieren (ohne GPKG) ---------
try:
    import tkinter as tk
    from tkinter import messagebox
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False

BASE = Path(__file__).resolve().parent
EU_XLSX = BASE / "Output" / "UWWTD_TP_Database_EUNISC.xlsx"
UK_XLSX = BASE / "Output" / "UWWTD_TP_Database_EN.xlsx"
HW_XLSX = BASE / "Output" / "UWWTD_TP_Database_HydroWASTE.xlsx"
OUT_XLSX = BASE / "Output" / "UWWTD_TP_Database.xlsx"

GENERAL_COLS = [
    "UWWTD Code", "Name", "Latitude", "Longitude", "Capacity/PE",
]

TECHNICAL_COLS = [
    "UWWTD Code", "Name",
    "Capacity/PE", "Ozonation", "Waste Water Treated [m³/a]",
    "BODIncomingMeasured [t/a]", "CODIncomingMeasured [t/a]", "NIncomingMeasured [t/a]",
    "BODDischargeMeasured [t/a]", "CODDischargeMeasured [t/a]", "NDischargeMeasured [t/a]"
]

FALLBACK_KEYS = ["Name", "CountryCode", "Latitude", "Longitude"]

URL_RE = re.compile(r"(https?://\S+)", re.IGNORECASE)

def notify(title: str, message: str, error: bool = False):
    if TK_AVAILABLE:
        try:
            root = tk.Tk(); root.withdraw()
            (messagebox.showerror if error else messagebox.showinfo)(title, message)
            root.destroy(); return
        except Exception:
            pass
    print(f"{title}: {message}")

def read_first_sheet(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str)  # first sheet by default
    df = df.apply(lambda c: c.str.strip() if c.dtype == "object" else c)
    return df

def union_frames(eu: pd.DataFrame, uk: pd.DataFrame) -> pd.DataFrame:
    # Union columns, EU first
    all_cols = list(dict.fromkeys(list(eu.columns) + [c for c in uk.columns if c not in eu.columns]))
    eu_u = eu.reindex(columns=all_cols)
    uk_u = uk.reindex(columns=all_cols)
    merged = pd.concat([eu_u, uk_u], ignore_index=True)
    if "UWWTD Code" in merged.columns:
        merged = merged.drop_duplicates(subset=["UWWTD Code"], keep="first")
    else:
        keys = [c for c in FALLBACK_KEYS if c in merged.columns]
        if keys:
            merged = merged.drop_duplicates(subset=keys, keep="first")
    return merged

def ensure_columns(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = pd.NA
    return out[cols]

def build_sheets(merged: pd.DataFrame) -> dict:
    general = ensure_columns(merged, GENERAL_COLS)
    technical = ensure_columns(merged, TECHNICAL_COLS)
    return {
        "General Data": general,
        "Technical Data - Plant Metrics": technical,
    }

def make_summary_hyperlinks(path: Path, sheet_name: str = "Summary", col: int = 1):
    """Links in Summary klickbar machen (sichtbarer Text bleibt erhalten)."""
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path); return
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if isinstance(val, str):
            m = URL_RE.search(val)
            if m:
                url = m.group(1)
                cell.hyperlink = url
                cell.style = "Hyperlink"
    wb.save(path)

def main_final():
    try:
        if not EU_XLSX.exists() or not UK_XLSX.exists():
            missing = [str(p) for p in (EU_XLSX, UK_XLSX) if not p.exists()]
            raise FileNotFoundError("Not found:\n" + "\n".join(missing))

        eu = read_first_sheet(EU_XLSX)
        uk = read_first_sheet(UK_XLSX)

        # HydroWASTE Daten hinzufügen (falls vorhanden)
        if HW_XLSX.exists():
            hw = read_first_sheet(HW_XLSX)
            # Erst EU + UK mergen, dann HydroWASTE hinzufügen
            merged = union_frames(eu, uk)
            merged = union_frames(merged, hw)
        else:
            merged = union_frames(eu, uk)
        sheets = build_sheets(merged)

        OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
            summary = pd.DataFrame({
                "Info": [
                    "Database European WWTPs Electrolysis",
                    "Fundamental data sources:",
                    "EU source - https://www.eea.europa.eu/en/datahub/datahubitem-view/6244937d-1c2c-47f5-bdf1-33ca01ff1715?activeAccordion=1094402",
                    "UK source - https://www.data.gov.uk/dataset/d7e2c57b-110a-462b-97a0-9833e7d26cc2/wastewater-treatment-in-england",
                ]
            })
            summary.to_excel(writer, sheet_name="Summary", index=False)
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)

        # Hyperlinks in Summary setzen
        make_summary_hyperlinks(OUT_XLSX, sheet_name="Summary", col=1)

        # Auto-Spaltenbreite für ALLE Sheets der finalen Database
        autosize_excel(OUT_XLSX)

    except Exception as e:
        notify("Error", f"{type(e).__name__}: {e}", error=True)
        sys.exit(1)

if __name__ == "__main__":
    main_final()

def _cleanup_intermediate_files():
    """
    Löscht still (ohne Textausgabe) die im Skript erzeugten Zwischendateien.
    Die finale Datenbank "UWWTD_TP_Database.xlsx" bleibt erhalten.
    """
    from pathlib import Path

    base = Path(__file__).resolve().parent
    out_dir = base / "Output"
    keep = {
        out_dir / "UWWTD_TP_Database.xlsx",
    }

    # Bekannte Zwischendateien
    candidates = [
        out_dir / "UWWTD_TP_Database_EU.xlsx",
        out_dir / "UWWTD_TP_Database_EU_2019.xlsx",
        out_dir / "UWWTD_TP_Database_EUNISC.xlsx",
        out_dir / "UWWTD_TP_Database_EN.xlsx",
        out_dir / "UWWTD_TP_Database_HydroWASTE.xlsx",
    ]

    # Löschlogik: stillschweigend und robust
    for c in candidates:
        if c not in keep:
            try:
                c.unlink(missing_ok=True)
            except Exception:
                # Keine Ausgabe, einfach überspringen
                pass

    # Zusätzlich: Musterbasierte Bereinigung direkt im Output-Ordner
    try:
        if out_dir.exists():
            for f in out_dir.iterdir():
                if f.is_file() and f.suffix.lower() == ".xlsx":
                    name = f.name.lower()
                    if (
                        name.endswith("_en.xlsx")
                        or "_eunisc" in name
                        or "_eu_2019" in name
                        or (name.startswith("uwwtd_tp_database_") and name != "uwwtd_tp_database.xlsx")
                    ) and f not in keep:
                        try:
                            f.unlink(missing_ok=True)
                        except Exception:
                            pass
    except Exception:
        pass


if __name__ == "__main__":
    try:
        _cleanup_intermediate_files()
    except Exception:
        pass

print("✓ Database created")
